from __future__ import annotations

import ast
from collections import OrderedDict
import json
import logging
import os
import re
from datetime import datetime
from io import BytesIO
from typing import Any

from anthropic import Anthropic
import requests
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backend.ledger_scan import preprocess_image_bytes
from backend.services.anthropic_usage import (
    ANTHROPIC_MODEL_HAIKU,
    ANTHROPIC_MODEL_OPUS,
    ANTHROPIC_MODEL_SONNET,
    build_anthropic_usage_summary,
    get_next_anthropic_model_upgrade,
    get_ocr_confidence_threshold,
    get_ocr_max_retries,
    merge_anthropic_usage_summaries,
    normalize_anthropic_model_name,
    serialize_anthropic_response_debug,
    verify_anthropic_response_model,
    would_exceed_cost_limit,
)
from backend.services.ocr_confidence import calculate_structural_confidence

logger = logging.getLogger(__name__)

BYTEZ_API_BASE = "https://api.bytez.com/models/v2"
BYTEZ_DEFAULT_MODEL = "google/gemma-7b"
DEFAULT_ANTHROPIC_MODEL_FAST = ANTHROPIC_MODEL_HAIKU
DEFAULT_ANTHROPIC_MODEL_BALANCED = ANTHROPIC_MODEL_SONNET
DEFAULT_ANTHROPIC_MODEL_BEST = ANTHROPIC_MODEL_OPUS
MAX_RETRY = max(0, get_ocr_max_retries() - 1)

SYSTEM_PROMPT = (
    "You are a document table extraction expert.\n"
    "Your job is to extract a table from an image into strict JSON format.\n\n"
    "RULES YOU MUST FOLLOW:\n"
    "- Return ONLY valid JSON, no explanation, no markdown, no backticks\n"
    "- The JSON must be an object with two keys: \"headers\" and \"rows\"\n"
    "- \"headers\" must be an array of strings (column names)\n"
    "- \"rows\" must be an array of arrays (each row aligns to headers)\n"
    "- If a header row is missing, infer short headers like \"Column 1\", \"Column 2\"\n"
    "- If a cell is empty or blank, use null\n"
    "- Do NOT skip any rows from the image\n"
    "- Do NOT add rows that don't exist in the image\n"
    "- Preserve the exact text and order from the image cells\n"
    "- Ensure every row has the same number of columns as headers (pad with nulls)\n"
    "\n"
    "Example output:\n"
    "{\n"
    "  \"headers\": [\"Name\", \"ID\", \"Fee\"],\n"
    "  \"rows\": [\n"
    "    [\"Amit\", \"S-102\", \"12000\"],\n"
    "    [\"Sara\", \"S-103\", \"15000\"]\n"
    "  ]\n"
    "}"
)

USER_MESSAGE = "Extract the table from this image into JSON following the rules exactly."

RETRY_MESSAGE = (
    "Your previous response was not valid JSON. "
    "Return ONLY the JSON object with headers and rows, nothing else."
)

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

TEXT_FONT = Font(name="Arial", size=10)
TEXT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
NUMBER_ALIGN = Alignment(horizontal="right", vertical="center", wrap_text=True)
FOOTER_LABEL_FONT = Font(name="Arial", size=10, bold=True)
FOOTER_VALUE_FONT = Font(name="Arial", size=10)
TOTAL_FILL = PatternFill("solid", fgColor="E2F0D9")
TOTAL_FONT = Font(name="Arial", size=10, bold=True)
LOW_CONF_FILL = PatternFill("solid", fgColor="FFF2CC")  # Light yellow
ORANGE_CONF_FILL = PatternFill("solid", fgColor="FDE9D9") # Orange
VERY_LOW_CONF_FILL = PatternFill("solid", fgColor="F4CCCC") # Light red
CORRECTED_FILL = PatternFill("solid", fgColor="D9EAF7") # Light blue
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

_FORMULA_PREFIXES = ("=", "+", "-", "@")
_COLUMN_PRIORITY_GROUPS = (
    ("date", "day", "time", "timestamp"),
    ("invoice", "bill", "receipt", "voucher", "reference", "ref", "number", "no", "id"),
    ("item", "name", "description", "product", "material"),
    ("qty", "quantity", "unit", "units"),
    ("rate", "price", "amount", "debit", "credit", "balance", "total"),
)



def _get_client() -> Anthropic:
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        raise ValueError("ANTHROPIC_API_KEY is not set. Cannot use Anthropic as table scan provider.")
    return Anthropic(
        api_key=api_key,
        timeout=_provider_timeout_seconds(),
    )


def _provider_timeout_seconds() -> float:
    raw = (
        os.getenv("TABLE_SCAN_PROVIDER_TIMEOUT_SECONDS")
        or os.getenv("OCR_PROVIDER_TIMEOUT_SECONDS")
        or "15"
    ).strip()
    try:
        value = float(raw)
    except ValueError:
        value = 15.0
    return max(3.0, min(60.0, value))


def _table_scan_provider() -> str:
    configured = (os.getenv("TABLE_SCAN_PROVIDER") or os.getenv("LEDGER_SCAN_PROVIDER") or "").strip().lower()
    if configured:
        return configured
    if _has_provider_key("anthropic"):
        return "anthropic"
    if _has_provider_key("bytez"):
        return "bytez"
    return "tesseract"


def _normalize_provider(provider: str | None) -> str:
    key = (provider or "").strip().lower()
    if key in {"claude", "anthropic"}:
        return "anthropic"
    if key in {"local", "ocr", "tesseract"}:
        return "tesseract"
    if key == "bytez":
        return "bytez"
    return ""


def _table_scan_provider_chain(primary: str | None, *, allow_local_fallback: bool = True) -> list[str]:
    chain_env = os.getenv("TABLE_SCAN_PROVIDER_CHAIN")
    if chain_env:
        raw = [item.strip() for item in chain_env.split(",")]
        seen: set[str] = set()
        chain: list[str] = []
        for item in raw:
            normalized = _normalize_provider(item)
            if normalized and normalized not in seen:
                chain.append(normalized)
                seen.add(normalized)
        if chain:
            return chain
    first = _normalize_provider(primary) or _normalize_provider(_table_scan_provider()) or "tesseract"
    fallback_order = {
        "anthropic": ["tesseract", "bytez"],
        "bytez": ["tesseract", "anthropic"],
        "tesseract": ["anthropic", "bytez"],
    }
    chain = [first]
    for candidate in fallback_order.get(first, ["tesseract", "anthropic", "bytez"]):
        if candidate not in chain:
            chain.append(candidate)
    if allow_local_fallback:
        return chain
    return [provider for provider in chain if provider != "tesseract"]


def _has_provider_key(provider: str) -> bool:
    if provider == "tesseract":
        return True
    if provider == "anthropic":
        return bool((os.getenv("ANTHROPIC_API_KEY") or "").strip())
    if provider == "bytez":
        return bool((os.getenv("BYTEZ_API_KEY") or "").strip())
    return False


def _anthropic_model_name(model_tier: str | None = None, requested_model: str | None = None) -> str:
    normalized_requested = normalize_anthropic_model_name(requested_model)
    if normalized_requested:
        return normalized_requested
    normalized_tier = (model_tier or "").strip().lower()
    if normalized_tier == "best":
        return (
            os.getenv("TABLE_SCAN_ANTHROPIC_MODEL_BEST")
            or os.getenv("OCR_ANTHROPIC_MODEL_BEST")
            or DEFAULT_ANTHROPIC_MODEL_BEST
        ).strip()
    if normalized_tier == "balanced":
        return (
            os.getenv("TABLE_SCAN_ANTHROPIC_MODEL_BALANCED")
            or os.getenv("OCR_ANTHROPIC_MODEL_BALANCED")
            or os.getenv("TABLE_SCAN_ANTHROPIC_MODEL")
            or os.getenv("OCR_ANTHROPIC_MODEL")
            or os.getenv("ANTHROPIC_MODEL")
            or DEFAULT_ANTHROPIC_MODEL_BALANCED
        ).strip()
    return (
        os.getenv("TABLE_SCAN_ANTHROPIC_MODEL_FAST")
        or os.getenv("OCR_ANTHROPIC_MODEL_FAST")
        or os.getenv("TABLE_SCAN_ANTHROPIC_MODEL")
        or os.getenv("OCR_ANTHROPIC_MODEL")
        or os.getenv("ANTHROPIC_MODEL")
        or DEFAULT_ANTHROPIC_MODEL_FAST
    ).strip()


def _bytez_auth_header(api_key: str) -> str:
    lowered = api_key.lower()
    if lowered.startswith("key ") or lowered.startswith("bearer "):
        return api_key
    return api_key


def _resolve_prompts(system_prompt: str | None, user_message: str | None) -> tuple[str, str]:
    if system_prompt is None:
        system_prompt = os.getenv("TABLE_SCAN_SYSTEM_PROMPT") or SYSTEM_PROMPT
    if user_message is None:
        user_message = os.getenv("TABLE_SCAN_USER_MESSAGE") or USER_MESSAGE
    return system_prompt, user_message


def _call_bytez(base64_image: str, *, system_prompt: str, user_message: str) -> Any:
    api_key = os.getenv("BYTEZ_API_KEY")
    if not api_key:
        raise RuntimeError("BYTEZ_API_KEY is not set in the environment.")

    model_id = (
        os.getenv("TABLE_SCAN_MODEL_ID")
        or os.getenv("BYTEZ_MODEL_ID")
        or BYTEZ_DEFAULT_MODEL
    ).strip() or BYTEZ_DEFAULT_MODEL
    provider_key = os.getenv("TABLE_SCAN_PROVIDER_KEY") or os.getenv("BYTEZ_PROVIDER_KEY")

    headers = {
        "Authorization": _bytez_auth_header(api_key),
        "Content-Type": "application/json",
    }
    if provider_key:
        headers["provider-key"] = provider_key

    image_url = os.getenv("BYTEZ_IMAGE_URL")
    if image_url:
        image_url = image_url.strip()
    if image_url and not image_url.lower().startswith(("http://", "https://")):
        logger.warning("BYTEZ_IMAGE_URL ignored (must start with http/https).")
        image_url = None
    if image_url:
        image_kind = "url"
        payload_mode = "content_blocks_url"
        payload = {
            "messages": [
                {"role": "system", "content": system_prompt},
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": user_message},
                        {"type": "image", "url": image_url},
                    ],
                },
            ],
            "stream": False,
            "params": {"temperature": 0, "max_length": 2048},
        }
    else:
        image_kind = "base64"
        payload_mode = "top_level_base64_with_messages"
        combined_text = f"{system_prompt}\n\n{user_message}"
        payload = {
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": combined_text},
            ],
            "text": combined_text,
            "base64": base64_image,
            "stream": False,
            "params": {"temperature": 0, "max_length": 2048},
        }

    logger.info(
        "TableScan Bytez request: model=%s image_kind=%s payload_mode=%s base64_len=%s has_provider_key=%s",
        model_id,
        image_kind,
        payload_mode,
        len(base64_image) if image_kind == "base64" else "n/a",
        bool(provider_key),
    )

    response = requests.post(
        f"{BYTEZ_API_BASE}/{model_id}",
        headers=headers,
        json=payload,
        timeout=_provider_timeout_seconds(),
    )
    if response.status_code >= 400:
        raise RuntimeError(f"Bytez request failed ({response.status_code}): {response.text}")
    data = response.json()
    error = data.get("error")
    if error:
        raise RuntimeError(f"Bytez error: {error}")
    return data.get("output")


def _extract_text_from_bytez_output(output: Any) -> str | None:
    if not isinstance(output, dict):
        return None
    content = output.get("content")
    if isinstance(content, str):
        return content
    if isinstance(content, list):
        parts: list[str] = []
        for block in content:
            if isinstance(block, str):
                parts.append(block)
            elif isinstance(block, dict):
                text = block.get("text")
                if isinstance(text, str):
                    parts.append(text)
        if parts:
            return "".join(parts)
    text = output.get("text")
    if isinstance(text, str):
        return text
    return None


def _extract_text(response) -> str:
    parts = []
    for block in getattr(response, "content", []) or []:
        if getattr(block, "type", None) == "text":
            parts.append(block.text)
    if not parts and hasattr(response, "content"):
        raw = response.content
        if isinstance(raw, str):
            parts.append(raw)
    return "".join(parts).strip()


def _call_local_tesseract(image_bytes: bytes) -> dict[str, Any]:
    from backend.ocr_utils import extract_table_from_image as local_extract_table_from_image

    local_result = local_extract_table_from_image(
        image_bytes,
        columns=5,
        language="auto",
    )
    max_columns = max((len(row) for row in local_result.rows), default=0)
    headers = [f"Column {index}" for index in range(1, max_columns + 1)]
    normalized_rows = [
        list(row) + [""] * max(0, max_columns - len(row))
        for row in (local_result.rows or [])
    ]
    return {"headers": headers, "rows": normalized_rows}


def _call_claude(
    image_bytes: bytes,
    base64_image: str,
    *,
    system_prompt: str,
    user_message: str,
    model_tier: str | None = None,
    requested_model: str | None = None,
    history: list[dict] | None = None,
) -> dict[str, Any]:
    client = _get_client()
    model = _anthropic_model_name(model_tier, requested_model)

    if history:
        messages = history
    else:
        messages = [
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": user_message},
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": "image/jpeg",
                            "data": base64_image,
                        },
                    },
                ],
            }
        ]

    messages_api = client.messages
    response = messages_api.create(
        model=model,
        max_tokens=2048,
        temperature=0,
        system=[
            {
                "type": "text",
                "text": system_prompt,
                "cache_control": {"type": "ephemeral"},
            }
        ],
        messages=messages,
        timeout=_provider_timeout_seconds(),
    )

    text = _extract_text(response)
    next_history = messages + [
        {"role": "assistant", "content": text}
    ]
    actual_model = verify_anthropic_response_model(
        model,
        response,
        context="TableScan",
    )
    usage_summary = build_anthropic_usage_summary(actual_model, response)
    logger.info(
        "TableScan Claude response model=%s input_tokens=%s output_tokens=%s total_tokens=%s estimated_cost=%s",
        actual_model,
        usage_summary["input_tokens"],
        usage_summary["output_tokens"],
        usage_summary["total_tokens"],
        usage_summary["estimated_cost"],
    )
    return {
        "text": text,
        "history": next_history,
        "model_used": actual_model,
        "usage_summary": usage_summary,
        "debug_response": serialize_anthropic_response_debug(response),
    }


def _normalize_table(data: Any) -> dict | None:
    if not isinstance(data, dict):
        return None

    headers = data.get("headers")
    rows = data.get("rows")
    if not isinstance(rows, list):
        return None

    if not isinstance(headers, list):
        headers = []
    headers = ["" if h is None else str(h).strip() for h in headers]

    normalized_rows: list[list[Any]] = []
    max_cols = len(headers)

    for row in rows:
        if isinstance(row, dict) and headers:
            row_list = [row.get(header) for header in headers]
        elif isinstance(row, list):
            row_list = row
        else:
            row_list = [row]

        normalized_rows.append(row_list)
        max_cols = max(max_cols, len(row_list))

    if max_cols == 0:
        return {"headers": [], "rows": []}

    if len(headers) < max_cols:
        headers.extend([f"Column {index}" for index in range(len(headers) + 1, max_cols + 1)])

    for index, header in enumerate(headers):
        if not header:
            headers[index] = f"Column {index + 1}"

    final_rows: list[list[Any]] = []
    for row in normalized_rows:
        if len(row) < max_cols:
            row = row + [None] * (max_cols - len(row))
        elif len(row) > max_cols:
            row = row[:max_cols]

        cleaned: list[Any] = []
        for cell in row:
            if cell is None:
                cleaned.append(None)
            elif isinstance(cell, (int, float, bool)):
                cleaned.append(cell)
            elif isinstance(cell, (dict, list)):
                cleaned.append(json.dumps(cell, ensure_ascii=False))
            else:
                cleaned.append(str(cell).strip())
        final_rows.append(cleaned)

    return {"headers": headers, "rows": final_rows}


def _extract_json_candidate(raw: str) -> Any | None:
    if not raw:
        return None
    raw = raw.strip()
    if not raw:
        return None
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        pass
    try:
        return ast.literal_eval(raw)
    except (ValueError, SyntaxError):
        pass

    for opener, closer in (("[", "]"), ("{", "}")):
        depth = 0
        start_idx: int | None = None
        for idx, ch in enumerate(raw):
            if ch == opener:
                if depth == 0:
                    start_idx = idx
                depth += 1
            elif ch == closer and depth > 0:
                depth -= 1
                if depth == 0 and start_idx is not None:
                    candidate = raw[start_idx : idx + 1]
                    try:
                        return json.loads(candidate)
                    except json.JSONDecodeError:
                        try:
                            return ast.literal_eval(candidate)
                        except (ValueError, SyntaxError):
                            start_idx = None
        # try next opener
    return None


def _is_numeric(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return True
    try:
        cleaned = re.sub(r"[^0-9\.\-]", "", str(value))
        float(cleaned)
        return True
    except (ValueError, TypeError):
        return False


def _to_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        cleaned = re.sub(r"[^0-9\.\-]", "", str(value))
        return float(cleaned)
    except (ValueError, TypeError):
        return 0.0


def _clean_excel_header(value: Any, index: int) -> str:
    text = "" if value is None else str(value).strip()
    return text or f"Column {index}"


def _normalize_column_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()


def _column_priority(value: str) -> tuple[int, str]:
    normalized = _normalize_column_name(value)
    for priority, aliases in enumerate(_COLUMN_PRIORITY_GROUPS):
        if normalized in aliases or any(alias in normalized for alias in aliases):
            return priority, normalized
    return len(_COLUMN_PRIORITY_GROUPS), normalized


def _make_unique_headers(headers: list[str], required_count: int) -> list[str]:
    unique_headers: list[str] = []
    seen: dict[str, int] = {}
    for index in range(max(required_count, len(headers))):
        base = _clean_excel_header(headers[index] if index < len(headers) else "", index + 1)
        count = seen.get(base, 0)
        seen[base] = count + 1
        unique_headers.append(base if count == 0 else f"{base} ({count + 1})")
    return unique_headers


def _numeric_text_parts(value: Any) -> tuple[bool, str]:
    if value is None or isinstance(value, bool):
        return False, ""
    if isinstance(value, (int, float)):
        return True, str(value)
    text = str(value).strip()
    if not text:
        return False, ""
    compact = re.sub(r"[\s,₹$€£]", "", text)
    if compact.endswith("%"):
        compact = compact[:-1]
    if re.fullmatch(r"[+-]?\d+(\.\d+)?", compact):
        return True, compact
    return False, compact


def _is_display_numeric(value: Any) -> bool:
    matched, _ = _numeric_text_parts(value)
    return matched


def _is_summable_numeric(value: Any) -> bool:
    if isinstance(value, str) and "%" in value:
        return False
    matched, _ = _numeric_text_parts(value)
    return matched


def _excel_safe_value(value: Any) -> Any:
    """
    Phase 3: Enhanced to use normalized numeric values for Excel cells.
    
    If value is a Phase 3 cell object with 'normalized' field,
    use the numeric value to create true Excel numeric cells.
    Otherwise preserve display value as string.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (int, float)):
        return value
    
    # Phase 3: Check if this is a cell object with normalized numeric value
    if isinstance(value, dict):
        # If cell has normalized value, use it for Excel (true numeric cell)
        if "normalized" in value and value["normalized"] is not None:
            try:
                return float(value["normalized"])
            except (ValueError, TypeError):
                pass  # Fall through to use display value

        # If it's a cell object with display value, extract it
        if "value" in value:
            display_value = value["value"]
            matched, compact = _numeric_text_parts(display_value)
            if matched:
                try:
                    return float(compact)
                except (ValueError, TypeError):
                    pass
            value = display_value
        else:
            # Other dict/list -> serialize
            value = json.dumps(value, ensure_ascii=False, default=str)
    elif isinstance(value, list):
        value = json.dumps(value, ensure_ascii=False, default=str)
    
    text = str(value).strip()
    if not text:
        return ""
    if text[0] in _FORMULA_PREFIXES and not _is_display_numeric(text):
        return f"'{text}"
    return text


def _normalize_excel_table(table: dict[str, Any]) -> tuple[list[str], list[list[Any]]]:
    raw_headers = table.get("headers")
    raw_rows = table.get("rows")
    explicit_headers = [
        _clean_excel_header(header, index + 1)
        for index, header in enumerate(raw_headers)
    ] if isinstance(raw_headers, list) else []

    if not isinstance(raw_rows, list):
        return _make_unique_headers(explicit_headers, len(explicit_headers)), []

    dict_rows = [row for row in raw_rows if isinstance(row, dict)]
    if dict_rows:
        discovered = OrderedDict()
        for header in explicit_headers:
            discovered[header] = None
        for row in dict_rows:
            for key in row:
                header = _clean_excel_header(key, len(discovered) + 1)
                if header not in discovered:
                    discovered[header] = None
        extra_headers = [header for header in discovered if header not in explicit_headers]
        if explicit_headers:
            headers = explicit_headers + sorted(extra_headers, key=_column_priority)
        else:
            headers = sorted(list(discovered.keys()), key=_column_priority)
        headers = _make_unique_headers(headers, len(headers))
        normalized_rows = [
            [_excel_safe_value(row.get(header, "")) for header in headers]
            for row in dict_rows
        ]
        return headers, normalized_rows

    normalized_rows: list[list[Any]] = []
    max_columns = len(explicit_headers)
    for row in raw_rows:
        row_values = row if isinstance(row, list) else [row]
        normalized_row = [_excel_safe_value(cell) for cell in row_values]
        normalized_rows.append(normalized_row)
        max_columns = max(max_columns, len(normalized_row))
    headers = _make_unique_headers(explicit_headers, max_columns)
    for row in normalized_rows:
        if len(row) < len(headers):
            row.extend([""] * (len(headers) - len(row)))
        elif len(row) > len(headers):
            row[:] = row[: len(headers)]
    return headers, normalized_rows


def _build_totals_row(headers: list[str], rows: list[list[Any]]) -> list[Any] | None:
    if not headers or not rows:
        return None
    totals: list[Any] = [""] * len(headers)
    numeric_columns: list[int] = []
    for column_index in range(len(headers)):
        values = [row[column_index] for row in rows if column_index < len(row) and row[column_index] not in {"", None}]
        if values and all(_is_summable_numeric(value) for value in values):
            totals[column_index] = round(sum(_to_float(value) for value in values), 2)
            numeric_columns.append(column_index)
    if not numeric_columns:
        return None
    label_index = next((index for index in range(len(headers)) if index not in numeric_columns), None)
    if label_index is None and len(headers) > 1:
        label_index = 0
        totals[label_index] = "Total"
    elif label_index is not None:
        totals[label_index] = "Total"
    return totals


def _apply_widths(ws) -> None:
    for column_cells in ws.columns:
        max_len = 10
        for cell in column_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)) + 2)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max_len, 50)


def _cell_review_metadata(value: Any) -> tuple[float | None, str | None]:
    if not isinstance(value, dict):
        return None, None
    raw_confidence = value.get("confidence")
    confidence: float | None
    try:
        confidence = float(raw_confidence)
    except (TypeError, ValueError):
        confidence = None
    if confidence is not None and confidence > 1:
        confidence = confidence / 100.0
    if confidence is not None:
        confidence = max(0.0, min(1.0, confidence))
    source = str(value.get("source") or "").strip().lower() or None
    return confidence, source


def _apply_review_style(cell, value: Any) -> tuple[float | None, str | None]:
    confidence, source = _cell_review_metadata(value)
    if source == "corrected":
        cell.fill = CORRECTED_FILL
    elif confidence is not None and confidence < 0.5:
        cell.fill = VERY_LOW_CONF_FILL
    elif confidence is not None and confidence < 0.7:
        cell.fill = ORANGE_CONF_FILL
    elif confidence is not None and confidence < 0.9:
        cell.fill = LOW_CONF_FILL

    if confidence is not None or source:
        parts: list[str] = []
        if confidence is not None:
            parts.append(f"Confidence: {round(confidence * 100)}%")
        if source:
            parts.append(f"Source: {source}")
        cell.comment = Comment("\n".join(parts), "DPR OCR")
    return confidence, source


def _append_metadata_sheet(wb: Workbook, metadata_rows: list[tuple[int, int, str, float | None, str | None]]) -> None:
    if not metadata_rows:
        return
    ws = wb.create_sheet("OCR Metadata")
    headers = ["Row", "Column", "Value", "Confidence", "Source"]
    for column_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=column_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    for row_index, (sheet_row, sheet_col, value, confidence, source) in enumerate(metadata_rows, start=2):
        ws.cell(row=row_index, column=1, value=sheet_row).border = THIN_BORDER
        ws.cell(row=row_index, column=2, value=sheet_col).border = THIN_BORDER
        ws.cell(row=row_index, column=3, value=value).border = THIN_BORDER
        ws.cell(row=row_index, column=4, value=round(confidence * 100, 2) if confidence is not None else None).border = THIN_BORDER
        ws.cell(row=row_index, column=5, value=source).border = THIN_BORDER
    _apply_widths(ws)


def validate_table_data(table: dict) -> dict:
    headers = table.get("headers", [])
    rows = table.get("rows", [])
    if not rows:
        table["metadata"] = {
            "validation_warnings": [],
            "low_confidence_cells": [],
            "is_validated": True,
            "confidence": calculate_structural_confidence(
                {
                    "headers": headers if isinstance(headers, list) else [],
                    "rows": rows if isinstance(rows, list) else [],
                }
            ),
        }
        return table

    num_cols = len(headers)
    num_rows = len(rows)
    low_confidence_cells: list[list[int]] = [] # [row_idx, col_idx]
    validation_warnings: list[str] = []

    # 1. Detect Numeric Columns and perform Vertical Sum Check
    for col_idx in range(num_cols):
        # A column is likely numeric if > 60% of its cells (excluding the last one) are numeric
        numeric_count = sum(1 for r in range(num_rows - 1) if _is_numeric(rows[r][col_idx]))
        if num_rows > 2 and (numeric_count / (num_rows - 1)) > 0.6:
            # Possible numeric column. Check if the last row is the sum.
            column_sum = sum(_to_float(rows[r][col_idx]) for r in range(num_rows - 1))
            total_value = _to_float(rows[num_rows - 1][col_idx])
            
            # Use a small epsilon for float comparison if needed, but here we assume ledger-style integers/decimals
            if total_value > 0 and abs(column_sum - total_value) > 0.01:
                # Discrepancy detected in a likely "Total" column
                header_name = headers[col_idx] or f"Column {col_idx+1}"
                validation_warnings.append(
                    f"Column '{header_name}': Last row value ({total_value}) does not match sum of previous rows ({column_sum})."
                )
                low_confidence_cells.append([num_rows - 1, col_idx])

    table["metadata"] = {
        "validation_warnings": validation_warnings,
        "low_confidence_cells": low_confidence_cells,
        "is_validated": True,
        "confidence": calculate_structural_confidence(
            {
                "headers": headers if isinstance(headers, list) else [],
                "rows": rows if isinstance(rows, list) else [],
            }
        ),
    }
    return table


def extract_table_from_image(
    image_bytes: bytes,
    *,
    system_prompt: str | None = None,
    user_message: str | None = None,
    preprocess_profile: str | None = None,
    provider_preference: str | None = None,
    allow_local_fallback: bool = True,
    model_tier: str | None = None,
    requested_model: str | None = None,
) -> dict:
    profile = (
        preprocess_profile
        or os.getenv("TABLE_SCAN_PREPROCESS_PROFILE")
        or os.getenv("IMAGE_PREPROCESS_PROFILE")
    )
    base64_image = preprocess_image_bytes(image_bytes, profile=profile)
    system_prompt, user_message = _resolve_prompts(system_prompt, user_message)
    providers = _table_scan_provider_chain(provider_preference or _table_scan_provider(), allow_local_fallback=allow_local_fallback)
    failures: list[str] = []
    last_error: Exception | None = None

    def _run_provider(provider: str) -> dict:
        usage_summaries: list[dict[str, Any]] = []
        last_debug_response: dict[str, Any] | None = None
        model_used: str | None = None
        if provider == "tesseract":
            table = _call_local_tesseract(image_bytes)
            table["provider_used"] = "tesseract"
            table["provider_model"] = "local-tesseract"
            table["ai_applied"] = False
            return table
        if provider == "bytez":
            output = _call_bytez(base64_image, system_prompt=system_prompt, user_message=user_message)
            table = _normalize_table(output)
            if table is not None:
                table["provider_used"] = "bytez"
                table["provider_model"] = os.getenv("TABLE_SCAN_MODEL_ID") or os.getenv("BYTEZ_MODEL_ID") or BYTEZ_DEFAULT_MODEL
                table["ai_applied"] = True
                return table

            raw = _extract_text_from_bytez_output(output)
            if raw is None:
                raw = output if isinstance(output, str) else json.dumps(output)
        else:
            selected_model = _anthropic_model_name(model_tier, requested_model)
            current_model = selected_model
            confidence_threshold = get_ocr_confidence_threshold()
            accumulated_cost = 0.0
            best_table: dict[str, Any] | None = None
            best_score = -1.0
            last_parsed: Any = None
            last_raw = ""
            last_history: list[dict] | None = None

            for attempt_index in range(MAX_RETRY + 1):
                if attempt_index > 0 and would_exceed_cost_limit(accumulated_cost, current_model):
                    logger.warning(
                        "TableScan retry stopped before attempt %s because model=%s would exceed cost limit.",
                        attempt_index + 1,
                        current_model,
                    )
                    break

                if attempt_index == 0:
                    prompt_text = user_message
                    prompt_history = None
                else:
                    prompt_text = RETRY_MESSAGE
                    prompt_history = last_history + [{"role": "user", "content": RETRY_MESSAGE}] if last_history is not None else None
                result = _call_claude(
                    image_bytes,
                    base64_image,
                    system_prompt=system_prompt,
                    user_message=prompt_text,
                    model_tier=model_tier,
                    requested_model=current_model,
                    history=prompt_history,
                )
                raw = result["text"]
                history = result["history"]
                model_used = str(result.get("model_used") or current_model or "")
                last_raw = raw
                last_history = history
                if result.get("usage_summary"):
                    usage_summaries.append(result["usage_summary"])
                    accumulated_cost += float(result["usage_summary"].get("estimated_cost") or 0.0)
                if result.get("debug_response"):
                    last_debug_response = result["debug_response"]

                parsed = _extract_json_candidate(raw)
                last_parsed = parsed
                table = _normalize_table(parsed) if parsed is not None else None
                if table is None:
                    logger.info("TableScan attempt %s produced invalid JSON table.", attempt_index + 1)
                else:
                    validated = validate_table_data(table)
                    confidence_payload = validated.get("metadata", {}).get("confidence") or {}
                    confidence_score = float(confidence_payload.get("score") or 0.0)
                    warnings = validated.get("metadata", {}).get("validation_warnings", [])
                    validated["provider_used"] = "anthropic"
                    validated["provider_model"] = model_used or current_model
                    validated["requested_model"] = normalize_anthropic_model_name(requested_model)
                    validated["selected_model"] = normalize_anthropic_model_name(requested_model) or selected_model
                    validated["ai_applied"] = True
                    validated["raw_text"] = raw
                    if confidence_score >= best_score:
                        best_score = confidence_score
                        best_table = dict(validated)
                    if not warnings and confidence_score >= confidence_threshold:
                        table = validated
                        break
                    logger.info(
                        "TableScan attempt %s confidence=%s warnings=%s model=%s",
                        attempt_index + 1,
                        confidence_score,
                        len(warnings),
                        model_used,
                    )

                next_model = get_next_anthropic_model_upgrade(current_model)
                if next_model is None:
                    break
                current_model = next_model

            if best_table is None:
                if last_parsed is not None:
                    raise ValueError("AI response JSON did not match the expected table schema after retry.")
                raise ValueError("AI response was not valid JSON after retry.")

            if usage_summaries:
                best_table["token_usage"] = merge_anthropic_usage_summaries(
                    best_table.get("provider_model"),
                    usage_summaries,
                )
            if last_debug_response:
                best_table["debug_response"] = last_debug_response
            return validate_table_data(best_table)

        parsed = _extract_json_candidate(raw)
        table = _normalize_table(parsed) if parsed is not None else None
        if table is not None:
            return validate_table_data(table)
        if parsed is not None:
            raise ValueError("AI response JSON did not match the expected table schema.")
        raise ValueError("AI response was not valid JSON after retry.")

    for provider in providers:
        if not _has_provider_key(provider):
            logger.warning("TableScan provider %s skipped (missing credentials or local OCR unavailable).", provider)
            failures.append(f"{provider}:missing_key")
            continue
        try:
            table = _run_provider(provider)
            if failures:
                logger.info("TableScan fallback succeeded with %s after %s", provider, failures)
            return validate_table_data(table)
        except Exception as error:  # pylint: disable=broad-except
            last_error = error
            logger.warning("TableScan provider failed (%s): %s", provider, error, exc_info=True)
            failures.append(provider)

    if last_error:
        raise ValueError(f"TableScan failed for providers: {', '.join(failures)}") from last_error
    raise ValueError("TableScan failed: no available providers.")


def build_table_excel_bytes(
    table: dict[str, Any],
    *,
    sheet_name: str = "Table",
    metadata: dict[str, Any] | None = None,
    include_totals: bool = True,
) -> bytes:
    headers, rows = _normalize_excel_table(table)
    raw_rows = table.get("rows") if isinstance(table.get("rows"), list) else []
    metadata_rows: list[list[Any]] = []
    for raw_row in raw_rows:
        row_values = raw_row if isinstance(raw_row, list) else [raw_row]
        next_row = list(row_values[: len(headers)])
        if len(next_row) < len(headers):
            next_row.extend([None] * (len(headers) - len(next_row)))
        metadata_rows.append(next_row)
    footer_metadata = OrderedDict(
        [
            ("Generated At", datetime.now().isoformat(timespec="seconds")),
            ("Total Rows", len(rows)),
            ("Total Columns", len(headers)),
        ]
    )
    if metadata:
        for key, value in metadata.items():
            footer_metadata[str(key)] = value

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    review_metadata_rows: list[tuple[int, int, str, float | None, str | None]] = []

    current_row = 1
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    current_row += 1
    for row_index, row in enumerate(rows):
        for col_index, value in enumerate(row, start=1):
            cell = ws.cell(row=current_row, column=col_index, value=value if value != "" else None)
            cell.font = TEXT_FONT
            cell.alignment = NUMBER_ALIGN if _is_display_numeric(value) else TEXT_ALIGN
            cell.border = THIN_BORDER
            metadata_value = metadata_rows[row_index][col_index - 1] if row_index < len(metadata_rows) else None
            confidence, source = _apply_review_style(cell, metadata_value)
            if confidence is not None or source:
                review_metadata_rows.append(
                    (
                        current_row,
                        col_index,
                        str(value or ""),
                        confidence,
                        source,
                    )
                )
        current_row += 1

    totals_row = _build_totals_row(headers, rows) if include_totals else None
    if totals_row:
        for col_index, value in enumerate(totals_row, start=1):
            cell = ws.cell(row=current_row, column=col_index, value=value if value != "" else None)
            cell.fill = TOTAL_FILL
            cell.font = TOTAL_FONT
            cell.alignment = NUMBER_ALIGN if isinstance(value, (int, float)) else TEXT_ALIGN
            cell.border = THIN_BORDER
        current_row += 1

    if footer_metadata:
        current_row += 1
        for key, value in footer_metadata.items():
            label_cell = ws.cell(row=current_row, column=1, value=str(key))
            label_cell.font = FOOTER_LABEL_FONT
            label_cell.alignment = TEXT_ALIGN
            value_cell = ws.cell(row=current_row, column=2, value=_excel_safe_value(value))
            value_cell.font = FOOTER_VALUE_FONT
            value_cell.alignment = NUMBER_ALIGN if _is_display_numeric(value) else TEXT_ALIGN
            current_row += 1

    ws.freeze_panes = "A2"
    _apply_widths(ws)
    if len(headers) >= 2:
        ws.column_dimensions[get_column_letter(2)].width = max(ws.column_dimensions[get_column_letter(2)].width or 10, 18)
    _append_metadata_sheet(wb, review_metadata_rows)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_report_excel_bytes(
    report_input: dict[str, Any],
    *,
    sheet_name: str = "Report",
) -> bytes:
    title = str(report_input.get("title") or "OCR Extraction")
    metadata = report_input.get("metadata") if isinstance(report_input.get("metadata"), dict) else {}
    totals = report_input.get("totals") if isinstance(report_input.get("totals"), dict) else {}
    raw_tables = report_input.get("tables") if isinstance(report_input.get("tables"), list) else []

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    current_row = 1
    ws.cell(row=current_row, column=1, value=title).font = Font(name="Arial", size=14, bold=True)
    ws.cell(row=current_row, column=1).alignment = TEXT_ALIGN
    current_row += 2

    max_width_by_column: dict[int, int] = {}
    table_count = 0
    row_count = 0
    max_columns = 0
    footer_metadata = OrderedDict()
    for key, value in metadata.items():
        footer_metadata[str(key)] = _excel_safe_value(value)

    for index, raw_table in enumerate(raw_tables, start=1):
        if not isinstance(raw_table, dict):
            continue
        headers, rows = _normalize_excel_table(raw_table)
        table_title = str(raw_table.get("title") or f"Section {index}")
        if not headers and not rows:
            continue
        table_count += 1
        row_count += len(rows)
        max_columns = max(max_columns, len(headers))

        title_cell = ws.cell(row=current_row, column=1, value=table_title)
        title_cell.font = Font(name="Arial", size=12, bold=True)
        title_cell.alignment = TEXT_ALIGN
        current_row += 1

        for col_index, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_index, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER
            max_width_by_column[col_index] = max(max_width_by_column.get(col_index, 10), len(str(header)) + 2)
        current_row += 1

        for row in rows:
            for col_index, value in enumerate(row, start=1):
                cell = ws.cell(row=current_row, column=col_index, value=value if value != "" else None)
                cell.font = TEXT_FONT
                cell.alignment = NUMBER_ALIGN if _is_display_numeric(value) else TEXT_ALIGN
                cell.border = THIN_BORDER
                if value not in {"", None}:
                    max_width_by_column[col_index] = max(max_width_by_column.get(col_index, 10), len(str(value)) + 2)
            current_row += 1

        totals_row = _build_totals_row(headers, rows)
        if totals_row:
            for col_index, value in enumerate(totals_row, start=1):
                cell = ws.cell(row=current_row, column=col_index, value=value if value != "" else None)
                cell.fill = TOTAL_FILL
                cell.font = TOTAL_FONT
                cell.alignment = NUMBER_ALIGN if isinstance(value, (int, float)) else TEXT_ALIGN
                cell.border = THIN_BORDER
            current_row += 1

        current_row += 1

    footer_metadata.setdefault("Table Count", table_count or int(totals.get("table_count") or 0))
    footer_metadata.setdefault("Row Count", row_count or int(totals.get("row_count") or 0))
    footer_metadata.setdefault("Column Count", max_columns or int(totals.get("column_count") or 0))
    for key, value in totals.items():
        label = str(key).replace("_", " ").title()
        footer_metadata.setdefault(label, _excel_safe_value(value))

    for key, value in footer_metadata.items():
        label_cell = ws.cell(row=current_row, column=1, value=str(key))
        label_cell.font = FOOTER_LABEL_FONT
        label_cell.alignment = TEXT_ALIGN
        value_cell = ws.cell(row=current_row, column=2, value=value if value != "" else None)
        value_cell.font = FOOTER_VALUE_FONT
        value_cell.alignment = NUMBER_ALIGN if _is_display_numeric(value) else TEXT_ALIGN
        current_row += 1

    ws.freeze_panes = "A3"
    for col_index, width in max_width_by_column.items():
        ws.column_dimensions[get_column_letter(col_index)].width = min(max(width, 10), 50)
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 10, 16)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 10, 18)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_excel_from_sections(
    payload: dict[str, Any],
    *,
    sheet_name: str = "Report",
) -> tuple[bytes, dict[str, Any]]:
    from backend.services.ocr_document_pipeline import transform_sections_to_report_input

    report_input = transform_sections_to_report_input(payload)
    return build_report_excel_bytes(report_input, sheet_name=sheet_name), report_input
