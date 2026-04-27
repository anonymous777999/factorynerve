from __future__ import annotations

import ast
import base64
import json
import logging
import os
import re
from datetime import datetime
from io import BytesIO
from typing import Any

from anthropic import Anthropic
import requests
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image, ImageEnhance

MAX_WIDTH = 1000
DEFAULT_ANTHROPIC_MODEL = "claude-3-5-haiku-20241022"
logger = logging.getLogger(__name__)
BYTEZ_API_BASE = "https://api.bytez.com/models/v2"
BYTEZ_DEFAULT_MODEL = "google/gemma-7b"

SYSTEM_PROMPT = (
    "You are a financial ledger data extraction expert.\n"
    "Your job is to extract all rows from ledger/account images \n"
    "into strict JSON format.\n\n"
    "RULES YOU MUST FOLLOW:\n"
    "- Return ONLY valid JSON, no explanation, no markdown, no backticks\n"
    "- Every row must follow this format exactly:\n"
    "  {\"particular\": \"string\", \"dr\": number_or_null, \"cr\": number_or_null}\n"
    "- Numbers must be plain integers with NO commas, NO symbols, NO spaces\n"
    "- Example: 860000 NOT 8,60,000 and NOT ₹860000\n"
    "- If a cell is empty or blank, use null\n"
    "- Do NOT skip any row from the image\n"
    "- Do NOT add rows that don't exist in the image\n"
    "- Preserve the exact spelling of particulars from the image\n"
    "- The final JSON must be an array of objects like this:\n\n"
    "[\n"
    "  {\"particular\": \"Cash in hand\", \"dr\": 26000, \"cr\": null},\n"
    "  {\"particular\": \"Sundry Creditors\", \"dr\": null, \"cr\": 430000}\n"
    "]"
)

USER_MESSAGE = (
    "Extract all rows from this ledger image into JSON format "
    "following the rules exactly."
)

RETRY_MESSAGE = (
    "Your previous response was not valid JSON. "
    "Return ONLY the JSON array, nothing else."
)

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

ALT_FILL = PatternFill("solid", fgColor="F2F2F2")
LOW_CONF_FILL = PatternFill("solid", fgColor="FFFF00")

DR_FONT = Font(name="Arial", size=10, color="00008B")
CR_FONT = Font(name="Arial", size=10, color="8B0000")
TEXT_FONT = Font(name="Arial", size=10)

TOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FONT = Font(name="Arial", size=10, bold=True)

BALANCED_FILL = PatternFill("solid", fgColor="00B050")
UNBALANCED_FILL = PatternFill("solid", fgColor="C00000")
BALANCE_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")


def preprocess_image_bytes(image_bytes: bytes, *, profile: str | None = None) -> str:
    """Return base64 JPEG after resizing and enhancement."""
    chosen = (profile or os.getenv("IMAGE_PREPROCESS_PROFILE") or "standard").strip().lower()
    profiles = {"standard", "high_contrast", "binarize"}
    if chosen not in profiles:
        logger.warning("Unknown preprocess profile '%s'; using standard.", chosen)
        chosen = "standard"

    with Image.open(BytesIO(image_bytes)) as img:
        if img.mode != "RGB":
            img = img.convert("RGB")

        width, height = img.size
        if width > MAX_WIDTH:
            ratio = MAX_WIDTH / float(width)
            new_height = int(height * ratio)
            img = img.resize((MAX_WIDTH, new_height), Image.LANCZOS)

        if chosen == "high_contrast":
            img = ImageEnhance.Contrast(img).enhance(2.2)
            img = ImageEnhance.Sharpness(img).enhance(2.3)
        elif chosen == "binarize":
            threshold_raw = os.getenv("IMAGE_BINARIZE_THRESHOLD", "170").strip()
            try:
                threshold = max(0, min(255, int(threshold_raw)))
            except ValueError:
                threshold = 170
            gray = img.convert("L")
            gray = ImageEnhance.Contrast(gray).enhance(2.4)
            bw = gray.point(lambda p: 255 if p > threshold else 0)
            img = bw.convert("RGB")
        else:
            img = ImageEnhance.Contrast(img).enhance(1.5)
            img = ImageEnhance.Sharpness(img).enhance(2.0)

        buffer = BytesIO()
        img.save(buffer, format="JPEG", quality=85, optimize=True)

    logger.info("Image preprocess profile=%s size=%sx%s", chosen, width, height)
    return base64.b64encode(buffer.getvalue()).decode("utf-8")


def _get_client() -> Anthropic:
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        raise ValueError("ANTHROPIC_API_KEY is not set. Cannot use Anthropic as ledger scan provider.")
    return Anthropic(api_key=api_key)


def _is_mock_enabled() -> bool:
    return os.getenv("LEDGER_SCAN_MOCK", "").strip().lower() in {"1", "true", "yes", "on"}


def _mock_rows() -> list[dict]:
    return [
        {"particular": "Cash in hand", "dr": 26000, "cr": None},
        {"particular": "Sundry Creditors", "dr": None, "cr": 430000},
        {"particular": "Purchase - Steel", "dr": 125000, "cr": None},
        {"particular": "Sales", "dr": None, "cr": 240000},
        {"particular": "?", "dr": None, "cr": None},
    ]


def _ledger_scan_provider() -> str:
    configured = os.getenv("LEDGER_SCAN_PROVIDER", "").strip().lower()
    if configured:
        return configured
    if _has_provider_key("anthropic"):
        return "anthropic"
    if _has_provider_key("bytez"):
        return "bytez"
    return "tesseract"


def _normalize_provider(provider: str | None) -> str:
    key = (provider or "").strip().lower()
    if key in {"claude", "anthropic"}:
        return "anthropic"
    if key in {"local", "ocr", "tesseract"}:
        return "tesseract"
    if key == "bytez":
        return "bytez"
    return ""


def _ledger_scan_provider_chain(primary: str | None) -> list[str]:
    chain_env = os.getenv("LEDGER_SCAN_PROVIDER_CHAIN")
    if chain_env:
        raw = [item.strip() for item in chain_env.split(",")]
        seen: set[str] = set()
        chain: list[str] = []
        for item in raw:
            normalized = _normalize_provider(item)
            if normalized and normalized not in seen:
                chain.append(normalized)
                seen.add(normalized)
        if chain:
            return chain
    first = _normalize_provider(primary) or _normalize_provider(_ledger_scan_provider()) or "tesseract"
    fallback_order = {
        "anthropic": ["tesseract", "bytez"],
        "bytez": ["tesseract", "anthropic"],
        "tesseract": ["anthropic", "bytez"],
    }
    chain = [first]
    for candidate in fallback_order.get(first, ["tesseract", "anthropic", "bytez"]):
        if candidate not in chain:
            chain.append(candidate)
    return chain


def _has_provider_key(provider: str) -> bool:
    if provider == "tesseract":
        return True
    if provider == "anthropic":
        return bool((os.getenv("ANTHROPIC_API_KEY") or "").strip())
    if provider == "bytez":
        return bool((os.getenv("BYTEZ_API_KEY") or "").strip())
    return False


def _anthropic_model_name() -> str:
    return (
        os.getenv("LEDGER_SCAN_ANTHROPIC_MODEL")
        or os.getenv("OCR_ANTHROPIC_MODEL")
        or os.getenv("ANTHROPIC_MODEL")
        or DEFAULT_ANTHROPIC_MODEL
    ).strip()


def _bytez_auth_header(api_key: str) -> str:
    lowered = api_key.lower()
    if lowered.startswith("key ") or lowered.startswith("bearer "):
        return api_key
    return api_key


def _resolve_prompts(system_prompt: str | None, user_message: str | None) -> tuple[str, str]:
    if system_prompt is None:
        system_prompt = os.getenv("LEDGER_SCAN_SYSTEM_PROMPT") or SYSTEM_PROMPT
    if user_message is None:
        user_message = os.getenv("LEDGER_SCAN_USER_MESSAGE") or USER_MESSAGE
    return system_prompt, user_message


def _call_bytez(base64_image: str, *, system_prompt: str, user_message: str) -> Any:
    api_key = os.getenv("BYTEZ_API_KEY")
    if not api_key:
        raise RuntimeError("BYTEZ_API_KEY is not set in the environment.")

    model_id = os.getenv("BYTEZ_MODEL_ID", BYTEZ_DEFAULT_MODEL).strip() or BYTEZ_DEFAULT_MODEL
    provider_key = os.getenv("BYTEZ_PROVIDER_KEY")

    headers = {
        "Authorization": _bytez_auth_header(api_key),
        "Content-Type": "application/json",
    }
    if provider_key:
        headers["provider-key"] = provider_key

    image_url = os.getenv("BYTEZ_IMAGE_URL")
    if image_url:
        image_url = image_url.strip()
    if image_url and not image_url.lower().startswith(("http://", "https://")):
        logger.warning("BYTEZ_IMAGE_URL ignored (must start with http/https).")
        image_url = None
    if image_url:
        image_kind = "url"
        payload_mode = "content_blocks_url"
        payload = {
            "messages": [
                {"role": "system", "content": system_prompt},
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": user_message},
                        {"type": "image", "url": image_url},
                    ],
                },
            ],
            "stream": False,
            "params": {"temperature": 0, "max_length": 2048},
        }
    else:
        image_kind = "base64"
        payload_mode = "top_level_base64_with_messages"
        combined_text = f"{system_prompt}\n\n{user_message}"
        payload = {
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": combined_text},
            ],
            "text": combined_text,
            "base64": base64_image,
            "stream": False,
            "params": {"temperature": 0, "max_length": 2048},
        }
    # Log only payload shape/metadata; avoid logging image contents or keys.
    logger.info(
        "Bytez request: model=%s image_kind=%s payload_mode=%s base64_len=%s has_provider_key=%s",
        model_id,
        image_kind,
        payload_mode,
        len(base64_image) if image_kind == "base64" else "n/a",
        bool(provider_key),
    )

    response = requests.post(f"{BYTEZ_API_BASE}/{model_id}", headers=headers, json=payload, timeout=180)
    if response.status_code >= 400:
        raise RuntimeError(f"Bytez request failed ({response.status_code}): {response.text}")
    data = response.json()
    error = data.get("error")
    if error:
        raise RuntimeError(f"Bytez error: {error}")
    return data.get("output")


def _maybe_rows_from_output(output: Any) -> list[dict] | None:
    if isinstance(output, list) and output and all(isinstance(item, dict) for item in output):
        if all({"particular", "dr", "cr"}.issubset(item.keys()) for item in output):
            return output
    return None


def _extract_text_from_bytez_output(output: Any) -> str | None:
    if not isinstance(output, dict):
        return None
    content = output.get("content")
    if isinstance(content, str):
        return content
    if isinstance(content, list):
        parts: list[str] = []
        for block in content:
            if isinstance(block, str):
                parts.append(block)
            elif isinstance(block, dict):
                text = block.get("text")
                if isinstance(text, str):
                    parts.append(text)
        if parts:
            return "".join(parts)
    text = output.get("text")
    if isinstance(text, str):
        return text
    return None


def _extract_json_candidate(raw: str) -> Any | None:
    if not raw:
        return None
    raw = raw.strip()
    if not raw:
        return None
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        pass
    try:
        return ast.literal_eval(raw)
    except (ValueError, SyntaxError):
        pass

    for opener, closer in (("[", "]"), ("{", "}")):
        depth = 0
        start_idx: int | None = None
        for idx, ch in enumerate(raw):
            if ch == opener:
                if depth == 0:
                    start_idx = idx
                depth += 1
            elif ch == closer and depth > 0:
                depth -= 1
                if depth == 0 and start_idx is not None:
                    candidate = raw[start_idx : idx + 1]
                    try:
                        return json.loads(candidate)
                    except json.JSONDecodeError:
                        try:
                            return ast.literal_eval(candidate)
                        except (ValueError, SyntaxError):
                            start_idx = None
        # try next opener
    return None


def _rows_from_parsed(parsed: Any) -> list[dict] | None:
    if isinstance(parsed, list):
        return _maybe_rows_from_output(parsed)
    if isinstance(parsed, dict):
        rows = parsed.get("rows")
        if isinstance(rows, list):
            return _maybe_rows_from_output(rows)
    return None


def _extract_text(response) -> str:
    parts = []
    for block in getattr(response, "content", []) or []:
        if getattr(block, "type", None) == "text":
            parts.append(block.text)
    if not parts and hasattr(response, "content"):
        raw = response.content
        if isinstance(raw, str):
            parts.append(raw)
    return "".join(parts).strip()


def _call_local_tesseract(base64_image: str) -> list[dict]:
    from backend.ocr_utils import extract_table_from_image as local_extract_table_from_image

    image_bytes = base64.b64decode(base64_image)
    local_result = local_extract_table_from_image(
        image_bytes,
        columns=3,
        language="auto",
    )
    rows: list[dict] = []
    for row in local_result.rows or []:
        particular = str(row[0] if len(row) > 0 else "").strip()
        dr_raw = row[1] if len(row) > 1 else None
        cr_raw = row[2] if len(row) > 2 else None
        rows.append(
            {
                "particular": particular,
                "dr": _normalize_amount(dr_raw),
                "cr": _normalize_amount(cr_raw),
            }
        )
    return rows


def _call_claude(
    base64_image: str,
    *,
    system_prompt: str,
    user_message: str,
    extra_message: str | None = None,
) -> str:
    client = _get_client()
    text_message = user_message
    if extra_message:
        text_message = f"{user_message}\n\n{extra_message}"

    response = client.messages.create(
        model=_anthropic_model_name(),
        max_tokens=2048,
        temperature=0,
        system=system_prompt,
        messages=[
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": text_message},
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": "image/jpeg",
                            "data": base64_image,
                        },
                    },
                ],
            }
        ],
    )
    return _extract_text(response)


def extract_data_from_image(
    base64_image: str,
    *,
    force_mock: bool = False,
    system_prompt: str | None = None,
    user_message: str | None = None,
) -> list[dict]:
    if force_mock or _is_mock_enabled():
        logger.info("LedgerScan mock mode enabled; returning sample rows.")
        return _mock_rows()
    system_prompt, user_message = _resolve_prompts(system_prompt, user_message)
    providers = _ledger_scan_provider_chain(_ledger_scan_provider())
    failures: list[str] = []
    last_error: Exception | None = None

    def _run_provider(provider: str) -> list[dict]:
        if provider == "tesseract":
            return _call_local_tesseract(base64_image)
        if provider == "bytez":
            output = _call_bytez(base64_image, system_prompt=system_prompt, user_message=user_message)
            rows = _maybe_rows_from_output(output)
            if rows is not None:
                return rows
            raw = _extract_text_from_bytez_output(output)
            if raw is None:
                raw = output if isinstance(output, str) else json.dumps(output)
        else:
            raw = _call_claude(base64_image, system_prompt=system_prompt, user_message=user_message)

        parsed = _extract_json_candidate(raw)
        rows = _rows_from_parsed(parsed)
        if rows is not None:
            return rows
        if parsed is not None:
            raise ValueError("AI response JSON did not match ledger schema.")
        if not (raw or "").strip():
            logger.warning("LedgerScan received empty response from provider %s.", provider)

        if provider == "bytez":
            output = _call_bytez(base64_image, system_prompt=system_prompt, user_message=user_message)
            rows = _maybe_rows_from_output(output)
            if rows is not None:
                return rows
            raw = _extract_text_from_bytez_output(output)
            if raw is None:
                raw = output if isinstance(output, str) else json.dumps(output)
        else:
            raw = _call_claude(
                base64_image,
                system_prompt=system_prompt,
                user_message=user_message,
                extra_message=RETRY_MESSAGE,
            )

        parsed = _extract_json_candidate(raw)
        rows = _rows_from_parsed(parsed)
        if rows is not None:
            return rows
        if parsed is not None:
            raise ValueError("AI response JSON did not match ledger schema after retry.")
        if not (raw or "").strip():
            logger.warning("LedgerScan received empty response after retry for %s.", provider)
        raise ValueError("AI response was not valid JSON after retry.")

    for provider in providers:
        if not _has_provider_key(provider):
            logger.warning("LedgerScan provider %s skipped (missing credentials or local OCR unavailable).", provider)
            failures.append(f"{provider}:missing_key")
            continue
        try:
            rows = _run_provider(provider)
            if failures:
                logger.info("LedgerScan fallback succeeded with %s after %s", provider, failures)
            return rows
        except Exception as error:  # pylint: disable=broad-except
            last_error = error
            logger.warning("LedgerScan provider failed (%s): %s", provider, error, exc_info=True)
            failures.append(provider)

    if last_error:
        raise ValueError(f"LedgerScan failed for providers: {', '.join(failures)}") from last_error
    raise ValueError("LedgerScan failed: no available providers.")


def _normalize_amount(value) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        cleaned = re.sub(r"[^0-9\-]", "", stripped)
        if cleaned in {"", "-"}:
            return None
        try:
            return int(cleaned)
        except ValueError:
            return None
    return None


def _normalize_particular(value) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _is_low_confidence(particular: str, dr: int | None, cr: int | None) -> bool:
    if dr is None and cr is None:
        return True
    if len(particular) < 2:
        return True
    if "?" in particular or "�" in particular:
        return True
    return False


def validate_data(rows: list[dict]) -> dict:
    cleaned_rows: list[dict] = []
    low_confidence_rows: list[int] = []

    for index, row in enumerate(rows or []):
        if not isinstance(row, dict):
            row = {}
        particular = _normalize_particular(row.get("particular"))
        dr = _normalize_amount(row.get("dr"))
        cr = _normalize_amount(row.get("cr"))

        cleaned = {"particular": particular, "dr": dr, "cr": cr}
        cleaned_rows.append(cleaned)

        if _is_low_confidence(particular, dr, cr):
            low_confidence_rows.append(index)

    total_dr = sum(value for value in (row["dr"] for row in cleaned_rows) if value is not None)
    total_cr = sum(value for value in (row["cr"] for row in cleaned_rows) if value is not None)
    difference = abs(total_dr - total_cr)
    balanced = total_dr == total_cr

    return {
        "rows": cleaned_rows,
        "metadata": {
            "total_dr": total_dr,
            "total_cr": total_cr,
            "balanced": balanced,
            "difference": difference,
            "low_confidence_rows": low_confidence_rows,
            "total_rows": len(cleaned_rows),
        },
    }


def build_excel_bytes(validated_data: dict) -> bytes:
    rows = validated_data.get("rows", [])
    metadata = validated_data.get("metadata", {})
    low_confidence_rows = set(metadata.get("low_confidence_rows", []))

    wb = Workbook()
    ws = wb.active
    ws.title = "Ledger"

    ws["A1"] = "Particulars"
    ws["B1"] = "Dr. (₹)"
    ws["C1"] = "Cr. (₹)"
    for cell in (ws["A1"], ws["B1"], ws["C1"]):
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18

    for idx, row in enumerate(rows):
        excel_row = idx + 2
        fill = ALT_FILL if idx % 2 == 1 else None
        if idx in low_confidence_rows:
            fill = LOW_CONF_FILL

        ws.cell(row=excel_row, column=1, value=row.get("particular") or "")
        ws.cell(row=excel_row, column=2, value=row.get("dr"))
        ws.cell(row=excel_row, column=3, value=row.get("cr"))

        ws.cell(row=excel_row, column=1).font = TEXT_FONT
        ws.cell(row=excel_row, column=1).alignment = Alignment(horizontal="left", indent=1)
        ws.cell(row=excel_row, column=2).font = DR_FONT
        ws.cell(row=excel_row, column=3).font = CR_FONT
        ws.cell(row=excel_row, column=2).alignment = Alignment(horizontal="right")
        ws.cell(row=excel_row, column=3).alignment = Alignment(horizontal="right")
        ws.cell(row=excel_row, column=2).number_format = "#,##0"
        ws.cell(row=excel_row, column=3).number_format = "#,##0"

        if fill:
            for col in range(1, 4):
                ws.cell(row=excel_row, column=col).fill = fill

        if idx in low_confidence_rows:
            comment = Comment("Low confidence - please verify", "LedgerScan")
            ws.cell(row=excel_row, column=1).comment = comment

    total_row = len(rows) + 2
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=2, value=f"=SUM(B2:B{total_row - 1})")
    ws.cell(row=total_row, column=3, value=f"=SUM(C2:C{total_row - 1})")

    for col in range(1, 4):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        if col > 1:
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")

    balance_row = total_row + 1
    balanced = bool(metadata.get("balanced"))
    difference = metadata.get("difference", 0)
    status_text = "✓ Balanced" if balanced else f"✗ Difference: ₹{difference}"

    ws.merge_cells(start_row=balance_row, start_column=1, end_row=balance_row, end_column=3)
    cell = ws.cell(row=balance_row, column=1, value=status_text)
    cell.font = BALANCE_FONT
    cell.fill = BALANCED_FILL if balanced else UNBALANCED_FILL
    cell.alignment = Alignment(horizontal="left")

    summary = wb.create_sheet("Summary")
    summary["A1"] = "Extraction timestamp"
    summary["B1"] = datetime.now().isoformat(timespec="seconds")
    summary["A2"] = "Total rows extracted"
    summary["B2"] = metadata.get("total_rows", 0)
    summary["A3"] = "Total Dr amount"
    summary["B3"] = metadata.get("total_dr", 0)
    summary["A4"] = "Total Cr amount"
    summary["B4"] = metadata.get("total_cr", 0)
    summary["A5"] = "Balance status"
    summary["B5"] = "Balanced" if balanced else f"Difference: ₹{difference}"
    summary["A6"] = "Low confidence rows"
    summary["B6"] = len(low_confidence_rows)

    for row in range(1, 7):
        summary[f"A{row}"].font = Font(name="Arial", size=10, bold=True)
        summary[f"B{row}"].font = Font(name="Arial", size=10)

    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 32

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
