# backend/services/ocr_export_generators.py
from __future__ import annotations

from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.units import mm, inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import json
from io import BytesIO
from datetime import datetime
import os

# For Excel generation
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    # We'll handle this gracefully in the functions

def generate_gst_invoice_pdf(data: dict) -> bytes:
    """Generate legally compliant GST Invoice PDF"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=15*mm, leftMargin=15*mm, topMargin=20*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()
    story = []
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=6,
        alignment=TA_CENTER
    )
    
    normal_style = styles['Normal']
    normal_style.fontSize = 9
    
    # Company header (would come from settings in real app)
    company_name = "ABC Steel Ltd"
    company_address = "123 Industrial Area, Mumbai - 400001"
    company_gstin = "24AABCS1234K1Z5"
    
    # Invoice header
    story.append(Paragraph(company_name, title_style))
    story.append(Paragraph(company_address, normal_style))
    story.append(Paragraph(f"GSTIN: {company_gstin}", normal_style))
    story.append(Spacer(1, 12))
    
    # Invoice title
    story.append(Paragraph("TAX INVOICE", subtitle_style))
    story.append(Spacer(1, 12))
    
    # Invoice details
    invoice_data = [
        ["Invoice No:", data.get("invoice_header", {}).get("invoice_number", ""), "Date:", data.get("invoice_header", {}).get("invoice_date", "")],
        ["Buyer (Bill To):", "", "", ""],
        [data.get("invoice_header", {}).get("recipient", {}).get("name", ""), "", "", ""],
        [data.get("invoice_header", {}).get("recipient", {}).get("address", ""), "", "", ""],
        [f"GSTIN/UIN: {data.get('invoice_header', {}).get('recipient', {}).get('gstin', 'UNREGISTERED')}", "", "", ""],
        ["Place of Supply:", data.get("invoice_header", {}).get("place_of_supply", ""), "", ""]
    ]
    
    # Create table for invoice details
    table = Table(invoice_data, colWidths=[30*mm, 60*mm, 20*mm, 60*mm])
    table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('BACKGROUND', (0,0), (0,-1), colors.lightgrey),
        ('BACKGROUND', (2,0), (2,-1), colors.lightgrey),
    ]))
    story.append(table)
    story.append(Spacer(1, 12))
    
    # Line items table
    line_items_data = [["Sr. No", "Description", "HSN Code", "Qty", "Unit", "Rate", "Amount"]]
    total_amount = 0
    for i, item in enumerate(data.get("line_items", []), 1):
        amount = float(item.get("qty", 0)) * float(item.get("rate", 0))
        total_amount += amount
        line_items_data.append([
            str(i),
            item.get("description", ""),
            item.get("hsn_code", ""),
            str(item.get("qty", "")),
            item.get("unit", ""),
            f"₹{float(item.get('rate', 0)):,.2f}",
            f"₹{amount:,.2f}"
        ])
    
    # Add total row
    line_items_data.append(["", "", "", "", "", "Total:", f"₹{total_amount:,.2f}"])
    
    items_table = Table(line_items_data, colWidths=[15*mm, 60*mm, 25*mm, 20*mm, 20*mm, 25*mm, 25*mm])
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('BACKGROUND', (0,1), (-1,-2), colors.beige),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
    ]))
    story.append(items_table)
    story.append(Spacer(1, 12))
    
    # Tax details
    tax_data = [
        ["Description", "Rate%", "Amount"],
        ["CGST", f"{data.get('tax_summary', {}).get('cgst_rate', 0)}%", f"₹{data.get('tax_summary', {}).get('cgst', 0):,.2f}"],
        ["SGST", f"{data.get('tax_summary', {}).get('sgst_rate', 0)}%", f"₹{data.get('tax_summary', {}).get('sgst', 0):,.2f}"],
        ["IGST", f"{data.get('tax_summary', {}).get('igst_rate', 0)}%", f"₹{data.get('tax_summary', {}).get('igst', 0):,.2f}"],
        ["CESS", f"{data.get('tax_summary', {}).get('cess_rate', 0)}%", f"₹{data.get('tax_summary', {}).get('cess', 0):,.2f}"],
        ["Total Tax", "", f"₹{data.get('totals', {}).get('total_tax', 0):,.2f}"]
    ]
    
    tax_table = Table(tax_data, colWidths=[80*mm, 30*mm, 40*mm])
    tax_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.black),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
    ]))
    story.append(tax_table)
    story.append(Spacer(1, 12))
    
    # Amount in words
    amount_in_words = f"Rupees {total_amount:,.2f} Only"  # Simplified - in real app would use proper conversion
    story.append(Paragraph(f"Amount Chargeable (in words): {amount_in_words}", normal_style))
    story.append(Spacer(1, 12))
    
    # Bank details (if any)
    if data.get("invoice_header", {}).get("bank_details"):
        story.append(Paragraph("Bank Details:", normal_style))
        story.append(Paragraph(data["invoice_header"]["bank_details"], normal_style))
        story.append(Spacer(1, 12))
    
    # Terms and conditions
    story.append(Paragraph("Terms & Conditions:", normal_style))
    story.append(Paragraph("1. Payment due within 30 days of invoice date.", normal_style))
    story.append(Paragraph("2. Interest @ 2% per month will be charged on late payments.", normal_style))
    story.append(Paragraph("3. Subject to Mumbai jurisdiction.", normal_style))
    
    # Build PDF
    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes

def generate_gst_invoice_excel(data: dict) -> bytes:
    """Generate GST Invoice Excel file"""
    if not EXCEL_AVAILABLE:
        # Fallback to CSV if openpyxl not available
        output = BytesIO()
        output.write(b"Excel generation requires openpyxl package. Installing it would enable this feature.")
        return output.getvalue()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    
    # Styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_aligned = Alignment(horizontal="center", vertical="center")
    
    # Company info
    ws['A1'] = "ABC Steel Ltd"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "123 Industrial Area, Mumbai - 400001"
    ws['A3'] = f"GSTIN: 24AABCS1234K1Z5"
    ws.merge_cells('A1:E1')
    ws.merge_cells('A2:E2')
    ws.merge_cells('A3:E3')
    
    # Invoice title
    ws['A5'] = "TAX INVOICE"
    ws['A5'].font = Font(bold=True, size=16)
    ws.merge_cells('A5:E5')
    
    # Invoice details
    row = 7
    ws[f'A{row}'] = "Invoice No:"
    ws[f'B{row}'] = data.get("invoice_header", {}).get("invoice_number", "")
    ws[f'D{row}'] = "Date:"
    ws[f'E{row}'] = data.get("invoice_header", {}).get("invoice_date", "")
    row += 1
    
    ws[f'A{row}'] = "Buyer (Bill To):"
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    ws[f'A{row}'] = data.get("invoice_header", {}).get("recipient", {}).get("name", "")
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    ws[f'A{row}'] = data.get("invoice_header", {}).get("recipient", {}).get("address", "")
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    ws[f'A{row}'] = f"GSTIN/UIN: {data.get('invoice_header', {}).get('recipient', {}).get('gstin', 'UNREGISTERED')}"
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    ws[f'A{row}'] = "Place of Supply:"
    ws[f'B{row}'] = data.get("invoice_header", {}).get("place_of_supply", "")
    row += 2
    
    # Column headers
    headers = ["Sr. No", "Description", "HSN Code", "Qty", "Unit", "Rate", "Amount"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_aligned
    row += 1
    
    # Line items
    total_amount = 0
    for i, item in enumerate(data.get("line_items", []), 1):
        amount = float(item.get("qty", 0)) * float(item.get("rate", 0))
        total_amount += amount
        vals = [
            i,
            item.get("description", ""),
            item.get("hsn_code", ""),
            item.get("qty", ""),
            item.get("unit", ""),
            float(item.get("rate", 0)),
            amount
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            if col in [6, 7]:  # Rate and Amount columns
                cell.number_format = '#,##0.00'
        row += 1
    
    # Total row
    ws.cell(row=row, column=5, value="Total:").font = Font(bold=True)
    ws.cell(row=row, column=6, value=total_amount).font = Font(bold=True)
    ws.cell(row=row, column=6).number_format = '#,##0.00'
    ws.cell(row=row, column=6).fill = header_fill
    row += 2
    
    # Tax details
    tax_labels = ["CGST", "SGST", "IGST", "CESS"]
    tax_values = [
        data.get("tax_summary", {}).get("cgst", 0),
        data.get("tax_summary", {}).get("sgst", 0),
        data.get("tax_summary", {}).get("igst", 0),
        data.get("tax_summary", {}).get("cess", 0)
    ]
    tax_rates = [
        data.get("tax_summary", {}).get("cgst_rate", 0),
        data.get("tax_summary", {}).get("sgst_rate", 0),
        data.get("tax_summary", {}).get("igst_rate", 0),
        data.get("tax_summary", {}).get("cess_rate", 0)
    ]
    
    for label, rate, value in zip(tax_labels, tax_rates, tax_values):
        ws.cell(row=row, column=1, value=f"{label} @ {rate}%:")
        ws.cell(row=row, column=6, value=value)
        ws.cell(row=row, column=6).number_format = '#,##0.00'
        row += 1
    
    # Total tax
    total_tax = sum(tax_values)
    ws.cell(row=row, column=1, value="Total Tax:").font = Font(bold=True)
    ws.cell(row=row, column=6, value=total_tax).font = Font(bold=True)
    ws.cell(row=row, column=6).number_format = '#,##0.00'
    ws.cell(row=row, column=6).fill = header_fill
    row += 2
    
    # Invoice total
    invoice_total = total_amount + total_tax
    ws.cell(row=row, column=1, value="Invoice Total:").font = Font(bold=True)
    ws.cell(row=row, column=6, value=invoice_total).font = Font(bold=True)
    ws.cell(row=row, column=6).number_format = '#,##0.00'
    ws.cell(row=row, column=6).fill = header_fill
    
    # Adjust column widths
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Save to bytes
    excel_bytes = BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    return excel_bytes.getvalue()

def generate_weighbridge_slip_pdf(data: dict) -> bytes:
    """Generate weighbridge slip PDF"""
    buffer = BytesIO()
    # Use a smaller page size for weighbridge slip (like 100mm x 150mm)
    doc = SimpleDocTemplate(buffer, pagesize=(100*mm, 150*mm), rightMargin=10*mm, leftMargin=10*mm, topMargin=10*mm, bottomMargin=10*mm)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=14,
        spaceAfter=6,
        alignment=TA_CENTER
    )
    
    normal_style = styles['Normal']
    normal_style.fontSize = 8
    
    # Header
    story.append(Paragraph("WEIGHBRIDGE SLIP", title_style))
    story.append(Spacer(1, 4))
    
    # Details in two columns
    data_left = [
        ["Slip No:", data.get("slip_no", "")],
        ["Date:", data.get("date", "")],
        ["Time:", data.get("time", "")],
        ["Vehicle No:", data.get("vehicle_no", "")],
        ["Driver:", data.get("driver_name", "")]
    ]
    
    data_right = [
        ["Material:", data.get("material", "")],
        ["Party:", data.get("party_name", "")],
        ["Gross Weight:", f"{data.get('gross_weight', 0)} kg"],
        ["Tare Weight:", f"{data.get('tare_weight', 0)} kg"],
        ["Net Weight:", f"{data.get('net_weight', 0)} kg"]
    ]
    
    # Create a table with two columns
    combined_data = []
    max_rows = max(len(data_left), len(data_right))
    for i in range(max_rows):
        left_row = data_left[i] if i < len(data_left) else ["", ""]
        right_row = data_right[i] if i < len(data_right) else ["", ""]
        combined_data.append(left_row + [""] + right_row)  # Empty column in between for spacing
    
    # Add rate and amount if available
    if data.get("rate") or data.get("amount"):
        combined_data.append(["Rate:", f"{data.get('rate', 0)} ₹/kg", "", "Amount:", f"{data.get('amount', 0)} ₹"])
    
    table = Table(combined_data, colWidths=[30*mm, 20*mm, 10*mm, 30*mm, 20*mm])
    table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('BACKGROUND', (0,0), (0,-1), colors.lightgrey),
        ('BACKGROUND', (3,0), (3,-1), colors.lightgrey),
    ]))
    story.append(table)
    
    # Build PDF
    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes

def generate_generic_excel(data: dict) -> bytes:
    """Generic Excel export for tabular data"""
    if not EXCEL_AVAILABLE:
        output = BytesIO()
        output.write(b"Excel generation requires openpyxl package. Installing it would enable this feature.")
        return output.getvalue()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Add headers
    if data.get("headers"):
        for col, header in enumerate(data["headers"], 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
    
    # Add rows
    if data.get("rows"):
        for row_idx, row in enumerate(data["rows"], 2):  # Start at row 2
            for col_idx, cell_value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)
    
    # Auto-adjust column widths
    for col in range(1, ws.max_column + 1):
        max_length = 0
        column = get_column_letter(col)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 50)
    
    # Save to bytes
    excel_bytes = BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    return excel_bytes.getvalue()

def generate_generic_pdf(data: dict) -> bytes:
    """Simple PDF export for generic tabular data"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=12,
        alignment=TA_CENTER
    )
    story.append(Paragraph("OCR Export", title_style))
    story.append(Spacer(1, 12))
    
    # Table data
    table_data = []
    if data.get("headers"):
        table_data.append(data["headers"])
    if data.get("rows"):
        table_data.extend(data["rows"])
    
    if table_data:
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('BOTTOMPADDING', (0,0), (-1,0), 6),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ]))
        story.append(table)
    
    # Build PDF
    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes

# Export format mapping
EXPORT_FORMATS = {
    "gst_invoice": {
        "pdf": generate_gst_invoice_pdf,
        "excel": generate_gst_invoice_excel
    },
    "weighbridge_slip": {
        "pdf": generate_weighbridge_slip_pdf,
        "excel": generate_generic_excel  # Fallback to generic for now
    },
    "delivery_note": {
        "pdf": generate_generic_pdf,  # Would implement specific one later
        "excel": generate_generic_excel
    },
    "generic_table": {
        "pdf": generate_generic_pdf,
        "excel": generate_generic_excel
    }
}

def get_export_function(document_type: str, format_type: str):
    """Get the appropriate export function for a document type and format"""
    if document_type in EXPORT_FORMATS and format_type in EXPORT_FORMATS[document_type]:
        return EXPORT_FORMATS[document_type][format_type]
    # Fallback to generic
    if format_type in EXPORT_FORMATS["generic_table"]:
        return EXPORT_FORMATS["generic_table"][format_type]
    # Final fallback
    return generate_generic_pdf if format_type == "pdf" else generate_generic_excel