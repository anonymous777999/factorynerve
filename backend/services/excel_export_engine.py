"""Excel Export Engine — Production-grade Excel generators for all document types.

Features (section 6.1):
- Type-specific sheet layouts
- Frozen panes and auto-filters
- Conditional formatting for confidence
- Data validation dropdowns
- Indian number format (#,##,##0.00)
- Multiple sheets per document type
- Summary sheet with processing metadata
- Audit trail sheet
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ── Constants ────────────────────────────────────────────────────────────────

_INDIAN_NUMBER_FMT = '#,##,##0.00'
_INDIAN_NUMBER_FMT_INT = '#,##,##0'
_DATE_FMT = 'DD-MMM-YYYY'
_DATETIME_FMT = 'DD-MMM-YYYY HH:MM'

# Colours
_HEADER_FILL = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_SUBHEADER_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
_SUBHEADER_FONT = Font(bold=True, size=10)
_DATA_FONT = Font(size=10)
_CONFIDENCE_HIGH_FILL = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
_CONFIDENCE_MED_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
_CONFIDENCE_LOW_FILL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
_DRAFT_WARNING_FILL = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
_ALTERNATE_ROW_FILL = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_DATA_ALIGN = Alignment(vertical="top", wrap_text=True)
_RIGHT_ALIGN = Alignment(horizontal="right", vertical="top")
_CENTER_ALIGN = Alignment(horizontal="center", vertical="center")


# ── Shared helpers ───────────────────────────────────────────────────────────

def _style_header_row(ws, row: int, max_col: int):
    """Apply header styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


def _style_data_rows(ws, start_row: int, end_row: int, max_col: int,
                     right_align_cols: set[int] | None = None,
                     confidence_col: int | None = None):
    """Apply alternating row styling and conditional formatting."""
    right_align_cols = right_align_cols or set()
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = _DATA_FONT
            cell.border = _THIN_BORDER
            if c in right_align_cols:
                cell.alignment = _RIGHT_ALIGN
                cell.number_format = _INDIAN_NUMBER_FMT
            elif c == confidence_col:
                cell.alignment = _CENTER_ALIGN
            else:
                cell.alignment = _DATA_ALIGN
        # Alternate row shading
        if r % 2 == 0:
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).fill = _ALTERNATE_ROW_FILL


def _apply_confidence_conditional(ws, col: int, start_row: int, end_row: int):
    """Apply conditional formatting: green ≥90%, yellow ≥60%, red <60%."""
    for r in range(start_row, end_row + 1):
        cell = ws.cell(row=r, column=col)
        try:
            val = float(cell.value or 0)
            if val >= 0.9:
                cell.fill = _CONFIDENCE_HIGH_FILL
            elif val >= 0.6:
                cell.fill = _CONFIDENCE_MED_FILL
            else:
                cell.fill = _CONFIDENCE_LOW_FILL
        except (ValueError, TypeError):
            pass


def _auto_column_width(ws, max_col: int, max_width: int = 50):
    """Auto-adjust column widths up to a max."""
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        lengths = []
        for cell in ws[letter]:
            if cell.value:
                try:
                    lengths.append(len(str(cell.value)))
                except Exception:
                    pass
        if lengths:
            ws.column_dimensions[letter].width = min(max(lengths) + 2, max_width)


def _add_title_row(ws, row: int, title: str, max_col: int):
    """Add a merged title row."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=14, color="1A1A2E")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _add_audit_sheet(wb: Workbook, data: dict, verification_meta: dict | None = None):
    """Add an Audit Trail sheet to the workbook (6.1)."""
    if "Audit Trail" in wb.sheetnames:
        return  # Already exists
    ws = wb.create_sheet("Audit Trail")
    _add_title_row(ws, 1, "Audit Trail", 2)
    ws.cell(row=2, column=1, value="Generated At:").font = Font(bold=True, size=10)
    ws.cell(row=2, column=2, value=datetime.now().isoformat(timespec="seconds"))
    audit_data = [
        ["Verification ID", str(verification_meta.get("id", "")) if verification_meta else ""],
        ["Status", verification_meta.get("status", "") if verification_meta else ""],
        ["Confidence", f"{verification_meta.get('avg_confidence', 'N/A')}" if verification_meta else "N/A"],
        ["Doc Type Hint", verification_meta.get("doc_type_hint", "") if verification_meta else ""],
        ["Source File", verification_meta.get("source_filename", "") if verification_meta else ""],
        ["Reviewer Notes", verification_meta.get("reviewer_notes", "") if verification_meta else ""],
        ["Export Type", "Excel"],
        ["Generated By", "DPR.ai OCR Export Engine"],
    ]
    for i, (k, v) in enumerate(audit_data, 3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True, size=10)
        ws.cell(row=i, column=2, value=v).font = _DATA_FONT
        ws.cell(row=i, column=1).border = _THIN_BORDER
        ws.cell(row=i, column=2).border = _THIN_BORDER
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50
    ws.freeze_panes = "A3"


def _add_summary_sheet(wb: Workbook, doc_type: str, row_count: int, confidence: float | None):
    """Add a Summary sheet with processing metadata (6.1)."""
    if "Summary" in wb.sheetnames:
        return
    ws = wb.create_sheet("Summary")
    _add_title_row(ws, 1, "Export Summary", 2)
    summary = [
        ["Document Type", doc_type],
        ["Total Rows", str(row_count)],
        ["Average Confidence", f"{confidence:.0%}" if confidence is not None else "N/A"],
        ["Export Timestamp", datetime.now().isoformat(timespec="seconds")],
        ["Generated By", "DPR.ai OCR Export Engine"],
    ]
    for i, (k, v) in enumerate(summary, 3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True, size=10)
        ws.cell(row=i, column=2, value=v).font = _DATA_FONT
        ws.cell(row=i, column=1).border = _THIN_BORDER
        ws.cell(row=i, column=2).border = _THIN_BORDER
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50
    ws.freeze_panes = "A3"


def _add_draft_warning(ws, row: int, max_col: int, status: str):
    """Add a draft/pending/rejected warning row."""
    if status == "approved":
        return
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1)
    if status == "draft":
        cell.value = "⚠ DRAFT — Not approved. Do not use for official records."
    else:
        cell.value = f"⚠ {status.upper()} — May contain unreviewed data."
    cell.font = Font(bold=True, color="333333", size=9)
    cell.fill = _DRAFT_WARNING_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).border = _THIN_BORDER


# ── Type-specific generators ─────────────────────────────────────────────────

def _generate_invoice_excel(data: dict, verification_meta: dict | None = None) -> bytes:
    """Generate GST Invoice Excel with Invoice, Tax Summary, Line Items, Audit sheets (6.2)."""
    wb = Workbook()
    header = data.get("invoice_header", {})
    line_items = data.get("line_items", [])
    tax_summary = data.get("tax_summary", {})
    totals = data.get("totals", {})
    status = (verification_meta or {}).get("status", "approved")

    # ── Sheet 1: Invoice ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Invoice"
    col_count = 7
    _add_title_row(ws, 1, "TAX INVOICE", col_count)
    ws.row_dimensions[1].height = 30

    r = 3
    invoice_fields = [
        ("Invoice No:", header.get("invoice_number", "")),
        ("Date:", header.get("invoice_date", "")),
        ("Supplier:", header.get("supplier", {}).get("name", "")),
        ("Supplier GSTIN:", header.get("supplier", {}).get("gstin", "")),
        ("Buyer:", header.get("recipient", {}).get("name", "")),
        ("Buyer GSTIN:", header.get("recipient", {}).get("gstin", "UNREGISTERED")),
        ("Place of Supply:", header.get("place_of_supply", "")),
    ]
    for label, val in invoice_fields:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=r, column=2, value=val).font = _DATA_FONT
        ws.cell(row=r, column=1).border = _THIN_BORDER
        ws.cell(row=r, column=2).border = _THIN_BORDER
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        r += 1

    r += 1
    line_headers = ["#", "Description", "HSN Code", "Qty", "Unit", "Rate", "Amount"]
    for c, h in enumerate(line_headers, 1):
        ws.cell(row=r, column=c, value=h)
    _style_header_row(ws, r, col_count)
    r += 1

    for i, item in enumerate(line_items, 1):
        amt = float(item.get("qty", 0)) * float(item.get("rate", 0))
        vals = [i, item.get("description", ""), item.get("hsn_code", ""),
                item.get("qty", ""), item.get("unit", ""),
                float(item.get("rate", 0)), amt]
        for c, v in enumerate(vals, 1):
            ws.cell(row=r, column=c, value=v)
        r += 1

    _style_data_rows(ws, r - len(line_items), r - 1, col_count, right_align_cols={6, 7})

    # Totals
    r += 1
    total_label = ws.cell(row=r, column=5, value="Invoice Total:")
    total_label.font = Font(bold=True, size=11)
    total_val = ws.cell(row=r, column=6, value=float(totals.get("invoice_total", 0)))
    total_val.font = Font(bold=True, size=11)
    total_val.number_format = _INDIAN_NUMBER_FMT
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
    ws.cell(row=r, column=5).border = _THIN_BORDER
    ws.cell(row=r, column=6).border = _THIN_BORDER

    _add_draft_warning(ws, r + 2, col_count, status)
    ws.freeze_panes = "A3"
    _auto_column_width(ws, col_count)

    # ── Sheet 2: Tax Summary ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Tax Summary")
    _add_title_row(ws2, 1, "Tax Summary", 3)
    tax_headers = ["Tax Type", "Rate %", "Amount"]
    for c, h in enumerate(tax_headers, 1):
        ws2.cell(row=3, column=c, value=h)
    _style_header_row(ws2, 3, 3)

    tax_items = [
        ("CGST", tax_summary.get("cgst_rate", 0), tax_summary.get("cgst", 0)),
        ("SGST", tax_summary.get("sgst_rate", 0), tax_summary.get("sgst", 0)),
        ("IGST", tax_summary.get("igst_rate", 0), tax_summary.get("igst", 0)),
        ("CESS", tax_summary.get("cess_rate", 0), tax_summary.get("cess", 0)),
    ]
    r = 4
    for label, rate, amt in tax_items:
        ws2.cell(row=r, column=1, value=label)
        ws2.cell(row=r, column=2, value=float(rate) if rate else 0)
        ws2.cell(row=r, column=3, value=float(amt) if amt else 0)
        r += 1
    ws2.cell(row=r, column=1, value="Total Tax").font = Font(bold=True)
    ws2.cell(row=r, column=3, value=float(sum(t[2] for t in tax_items if t[2]))).font = Font(bold=True)
    ws2.cell(row=r, column=3).number_format = _INDIAN_NUMBER_FMT
    _style_data_rows(ws2, 4, r, 3, right_align_cols={3})
    ws2.freeze_panes = "A4"
    _auto_column_width(ws2, 3)

    # ── Sheet 3: Line Items ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Line Items")
    _add_title_row(ws3, 1, "Line Items", col_count)
    for c, h in enumerate(line_headers, 1):
        ws3.cell(row=3, column=c, value=h)
    _style_header_row(ws3, 3, col_count)
    for i, item in enumerate(line_items, 1):
        amt = float(item.get("qty", 0)) * float(item.get("rate", 0))
        vals = [i, item.get("description", ""), item.get("hsn_code", ""),
                item.get("qty", ""), item.get("unit", ""),
                float(item.get("rate", 0)), amt]
        for c, v in enumerate(vals, 1):
            ws3.cell(row=3 + i, column=c, value=v)
    _style_data_rows(ws3, 4, 3 + len(line_items), col_count, right_align_cols={6, 7})
    ws3.freeze_panes = "A4"
    _auto_column_width(ws3, col_count)

    _add_audit_sheet(wb, data, verification_meta)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _generate_po_excel(data: dict, verification_meta: dict | None = None) -> bytes:
    """Generate Purchase Order Excel with PO, Line Items, Terms, Audit sheets (6.2)."""
    wb = Workbook()
    status = (verification_meta or {}).get("status", "approved")

    ws = wb.active
    ws.title = "Purchase Order"
    col_count = 7
    _add_title_row(ws, 1, "PURCHASE ORDER", col_count)

    r = 3
    po_fields = [
        ("PO No:", data.get("po_number", "")),
        ("Date:", data.get("po_date", "")),
        ("Vendor:", data.get("vendor", {}).get("name", "")),
        ("Vendor GSTIN:", data.get("vendor", {}).get("gstin", "")),
        ("Delivery Date:", data.get("delivery_date", "")),
    ]
    for label, val in po_fields:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=r, column=2, value=val).font = _DATA_FONT
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        ws.cell(row=r, column=1).border = _THIN_BORDER
        ws.cell(row=r, column=2).border = _THIN_BORDER
        r += 1

    r += 1
    item_headers = ["#", "Description", "HSN", "Qty", "Unit", "Rate", "Amount"]
    for c, h in enumerate(item_headers, 1):
        ws.cell(row=r, column=c, value=h)
    _style_header_row(ws, r, col_count)
    r += 1

    for i, item in enumerate(data.get("line_items", []), 1):
        vals = [i, item.get("description", ""), item.get("hsn_code", ""),
                item.get("qty", ""), item.get("unit", ""),
                float(item.get("rate", 0)), float(item.get("amount", 0))]
        for c, v in enumerate(vals, 1):
            ws.cell(row=r, column=c, value=v)
        r += 1

    _style_data_rows(ws, r - len(data.get("line_items", [])), r - 1, col_count, right_align_cols={6, 7})
    _add_draft_warning(ws, r + 1, col_count, status)
    ws.freeze_panes = "A3"
    _auto_column_width(ws, col_count)

    # ── Terms Sheet ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Terms")
    _add_title_row(ws2, 1, "Terms & Conditions", 2)
    r2 = 3
    terms_data = [
        ("Delivery Terms:", data.get("delivery_terms", "")),
        ("Payment Terms:", data.get("payment_terms", "")),
    ]
    for label, val in terms_data:
        ws2.cell(row=r2, column=1, value=label).font = Font(bold=True, size=10)
        ws2.cell(row=r2, column=2, value=val).font = _DATA_FONT
        r2 += 1
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 60

    _add_audit_sheet(wb, data, verification_meta)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _generate_dn_excel(data: dict, verification_meta: dict | None = None) -> bytes:
    """Generate Delivery Note Excel with Delivery Note, Items, Vehicle, Audit sheets (6.2)."""
    wb = Workbook()
    status = (verification_meta or {}).get("status", "approved")

    ws = wb.active
    ws.title = "Delivery Note"
    col_count = 6
    _add_title_row(ws, 1, "DELIVERY NOTE / CHALLAN", col_count)

    r = 3
    dn_fields = [
        ("Challan No:", data.get("challan_number", "")),
        ("Date:", data.get("date", "")),
        ("Supplier:", data.get("supplier", {}).get("name", "")),
        ("Recipient:", data.get("recipient", {}).get("name", "")),
    ]
    for label, val in dn_fields:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=r, column=2, value=val).font = _DATA_FONT
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        ws.cell(row=r, column=1).border = _THIN_BORDER
        ws.cell(row=r, column=2).border = _THIN_BORDER
        r += 1

    r += 1
    item_headers = ["#", "Description", "HSN", "Ordered Qty", "Delivered Qty", "Unit"]
    for c, h in enumerate(item_headers, 1):
        ws.cell(row=r, column=c, value=h)
    _style_header_row(ws, r, col_count)
    r += 1

    for i, item in enumerate(data.get("line_items", []), 1):
        vals = [i, item.get("description", ""), item.get("hsn_code", ""),
                item.get("ordered_qty", ""), item.get("delivered_qty", ""), item.get("unit", "")]
        for c, v in enumerate(vals, 1):
            ws.cell(row=r, column=c, value=v)
        r += 1

    _style_data_rows(ws, r - len(data.get("line_items", [])), r - 1, col_count)
    _add_draft_warning(ws, r + 1, col_count, status)
    ws.freeze_panes = "A3"
    _auto_column_width(ws, col_count)

    # ── Vehicle Sheet ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Vehicle")
    _add_title_row(ws2, 1, "Vehicle Details", 2)
    vehicle = data.get("vehicle", {})
    vehicle_fields = [
        ("Vehicle No:", vehicle.get("number", "")),
        ("Driver:", vehicle.get("driver_name", "")),
        ("Transporter:", vehicle.get("transporter", "")),
    ]
    r2 = 3
    for label, val in vehicle_fields:
        ws2.cell(row=r2, column=1, value=label).font = Font(bold=True, size=10)
        ws2.cell(row=r2, column=2, value=val).font = _DATA_FONT
        r2 += 1
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 40

    _add_audit_sheet(wb, data, verification_meta)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _generate_wb_excel(data: dict, verification_meta: dict | None = None) -> bytes:
    """Generate Weighbridge Slip Excel with Slip, Audit sheets (6.2)."""
    wb = Workbook()
    status = (verification_meta or {}).get("status", "approved")

    ws = wb.active
    ws.title = "Weighbridge Slip"
    col_count = 2
    _add_title_row(ws, 1, "WEIGHBRIDGE SLIP", col_count)
    ws.row_dimensions[1].height = 30

    slip_fields = [
        ("Slip No:", data.get("slip_no", "")),
        ("Date:", data.get("date", "")),
        ("Time:", data.get("time", "")),
        ("Vehicle No:", data.get("vehicle_no", "")),
        ("Driver:", data.get("driver_name", "")),
        ("Material:", data.get("material", "")),
        ("Party:", data.get("party_name", "")),
        ("Gross Weight (kg):", data.get("gross_weight", "")),
        ("Tare Weight (kg):", data.get("tare_weight", "")),
        ("Net Weight (kg):", data.get("net_weight", "")),
    ]
    r = 3
    for label, val in slip_fields:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=11)
        ws.cell(row=r, column=2, value=val).font = Font(size=11)
        ws.cell(row=r, column=1).border = _THIN_BORDER
        ws.cell(row=r, column=2).border = _THIN_BORDER
        if isinstance(val, (int, float)):
            ws.cell(row=r, column=2).number_format = _INDIAN_NUMBER_FMT_INT
        r += 1

    # Rate and amount if present
    if data.get("rate") or data.get("amount"):
        for label, val in [("Rate (₹/kg):", data.get("rate", 0)), ("Amount (₹):", data.get("amount", 0))]:
            ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=11)
            ws.cell(row=r, column=2, value=float(val)).font = Font(bold=True, size=11)
            ws.cell(row=r, column=2).number_format = _INDIAN_NUMBER_FMT
            ws.cell(row=r, column=1).border = _THIN_BORDER
            ws.cell(row=r, column=2).border = _THIN_BORDER
            r += 1

    _add_draft_warning(ws, r + 1, col_count, status)
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    ws.freeze_panes = "A3"

    _add_audit_sheet(wb, data, verification_meta)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _generate_ledger_excel(data: dict, verification_meta: dict | None = None) -> bytes:
    """Generate Ledger Sheet Excel with Ledger, Summary, Audit sheets (6.2)."""
    wb = Workbook()
    status = (verification_meta or {}).get("status", "approved")
    rows = data.get("rows", [])
    metadata = data.get("metadata", {})
    account_name = metadata.get("account_name", "Ledger Account")

    ws = wb.active
    ws.title = "Ledger"
    col_count = 6

    _add_title_row(ws, 1, f"LEDGER: {account_name}", col_count)
    r = 3
    for label, val in [("Account:", account_name), ("Period:", metadata.get("period", "")),
                        ("Opening Balance:", metadata.get("opening_balance", "0.00")),
                        ("Closing Balance:", metadata.get("closing_balance", "0.00"))]:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=r, column=2, value=val).font = _DATA_FONT
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        ws.cell(row=r, column=1).border = _THIN_BORDER
        ws.cell(row=r, column=2).border = _THIN_BORDER
        r += 1

    r += 1
    if rows and isinstance(rows[0], dict):
        ld_headers = list(rows[0].keys())
    else:
        ld_headers = ["Date", "Particulars", "Vch Type", "Debit (₹)", "Credit (₹)", "Balance (₹)"]
    col_count = len(ld_headers)

    for c, h in enumerate(ld_headers, 1):
        ws.cell(row=r, column=c, value=h)
    _style_header_row(ws, r, col_count)
    r += 1

    for row in rows:
        if isinstance(row, dict):
            vals = [row.get(h, "") for h in ld_headers]
        elif isinstance(row, (list, tuple)):
            vals = list(row)
        else:
            continue
        for c, v in enumerate(vals, 1):
            ws.cell(row=r, column=c, value=v)
        r += 1

    _style_data_rows(ws, r - len(rows), r - 1, col_count,
                     right_align_cols={ld_headers.index(h) + 1 for h in ld_headers
                                       if any(k in h.lower() for k in ["debit", "credit", "balance", "amount", "₹"])})
    _add_draft_warning(ws, r + 1, col_count, status)
    ws.freeze_panes = f"A{r - len(rows)}"
    _auto_column_width(ws, col_count)

    # ── Summary Sheet ────────────────────────────────────────────────────
    _add_summary_sheet(wb, "ledger_sheet", len(rows), verification_meta.get("avg_confidence") if verification_meta else None)
    _add_audit_sheet(wb, data, verification_meta)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _generate_kv_excel(data: dict, verification_meta: dict | None = None) -> bytes:
    """Generate Handwritten Form / Key-Value Excel with Fields, Audit sheets (6.2).

    Handwritten/misclassified docs whose content is actually a table now arrive
    as {type: table, headers, rows} (the model decides layout). A key-value sheet
    would render those empty, so delegate table-shaped payloads to the generic
    table generator instead of silently dropping every row.
    """
    has_table = (
        str(data.get("type") or "").strip().lower() == "table"
        or (isinstance(data.get("headers"), list) and isinstance(data.get("rows"), list)
            and (data.get("headers") or data.get("rows")))
    )
    has_fields = bool(
        data.get("fields") or data.get("key_value_pairs") or data.get("extracted_data")
    )
    if has_table and not has_fields:
        return _generate_generic_excel(data, verification_meta)

    wb = Workbook()
    status = (verification_meta or {}).get("status", "approved")

    ws = wb.active
    ws.title = "Fields"
    col_count = 3

    _add_title_row(ws, 1, "HANDWRITTEN FORM — EXTRACTED FIELDS", col_count)
    kv_headers = ["Field", "Value", "Confidence"]
    r = 3
    for c, h in enumerate(kv_headers, 1):
        ws.cell(row=r, column=c, value=h)
    _style_header_row(ws, r, col_count)
    r += 1

    fields = data.get("fields", data.get("key_value_pairs", data.get("extracted_data", {})))
    if isinstance(fields, dict):
        items = list(fields.items())
    elif isinstance(fields, list):
        items = [(item.get("field", item.get("key", "")), item) for item in fields]
    else:
        items = []

    max_conf_col = 3  # Confidence column index
    for key, val in items:
        if isinstance(val, dict):
            ws.cell(row=r, column=1, value=str(key))
            ws.cell(row=r, column=2, value=str(val.get("value", "")))
            conf = val.get("confidence")
            if conf is not None:
                ws.cell(row=r, column=3, value=float(conf))
        else:
            ws.cell(row=r, column=1, value=str(key))
            ws.cell(row=r, column=2, value=str(val))
        r += 1

    _style_data_rows(ws, 4, r - 1, col_count, confidence_col=max_conf_col)
    _apply_confidence_conditional(ws, max_conf_col, 4, r - 1)
    _add_draft_warning(ws, r + 1, col_count, status)
    ws.freeze_panes = "A4"
    _auto_column_width(ws, col_count)

    _add_audit_sheet(wb, data, verification_meta)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _generate_chat_excel(data: dict, verification_meta: dict | None = None) -> bytes:
    """Generate Chat Transcript Excel with Messages, Summary, Audit sheets (6.2)."""
    wb = Workbook()
    status = (verification_meta or {}).get("status", "approved")
    messages = data.get("messages", data.get("conversation", []))

    ws = wb.active
    ws.title = "Messages"
    col_count = 4
    _add_title_row(ws, 1, "CHAT / SCREENSHOT TRANSCRIPT", col_count)

    msg_headers = ["#", "Sender", "Timestamp", "Message"]
    r = 3
    for c, h in enumerate(msg_headers, 1):
        ws.cell(row=r, column=c, value=h)
    _style_header_row(ws, r, col_count)
    r += 1

    for i, msg in enumerate(messages, 1):
        vals = [
            i,
            msg.get("sender", msg.get("from", "")),
            msg.get("timestamp", msg.get("time", "")),
            msg.get("message", msg.get("text", "")),
        ]
        for c, v in enumerate(vals, 1):
            ws.cell(row=r, column=c, value=v)
        r += 1

    _style_data_rows(ws, 4, r - 1, col_count)
    _add_draft_warning(ws, r + 1, col_count, status)
    ws.freeze_panes = "A4"
    _auto_column_width(ws, col_count)
    ws.column_dimensions["D"].width = 70  # Message column wider

    # Summary
    if data.get("summary"):
        ws2 = wb.create_sheet("Summary")
        _add_title_row(ws2, 1, "Chat Summary", 1)
        ws2.cell(row=3, column=1, value=data["summary"]).font = _DATA_FONT
        ws2.cell(row=3, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws2.column_dimensions["A"].width = 100

    _add_audit_sheet(wb, data, verification_meta)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _generate_generic_excel(data: dict, verification_meta: dict | None = None) -> bytes:
    """Generate Generic Table Excel with Data, Audit sheets (6.2)."""
    wb = Workbook()
    status = (verification_meta or {}).get("status", "approved")

    ws = wb.active
    ws.title = "Data"
    headers = data.get("headers", [])
    rows = data.get("rows", [])
    col_count = max(len(headers), 1)

    _add_title_row(ws, 1, "OCR DATA EXPORT", col_count)
    r = 3

    for c, h in enumerate(headers, 1):
        ws.cell(row=r, column=c, value=h)
    _style_header_row(ws, r, col_count)
    r += 1

    for row in rows:
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)
        r += 1

    _style_data_rows(ws, 4, r - 1, col_count)
    _add_draft_warning(ws, r + 1, col_count, status)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(col_count)}{r - 1}"
    _auto_column_width(ws, col_count)

    _add_audit_sheet(wb, data, verification_meta)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Excel Export Engine (6.1) ────────────────────────────────────────────────

EXCEL_GENERATORS: dict[str, Callable[[dict, dict | None], bytes]] = {
    "gst_invoice": _generate_invoice_excel,
    "purchase_order": _generate_po_excel,
    "delivery_note": _generate_dn_excel,
    "weighbridge_slip": _generate_wb_excel,
    "ledger_sheet": _generate_ledger_excel,
    "handwritten_form": _generate_kv_excel,
    "chat_transcript": _generate_chat_excel,
    "generic_table": _generate_generic_excel,
}


class ExcelExportEngine:
    """Production-grade Excel export engine.

    Features (6.1):
    - Type-specific sheet layouts via _get_excel_generator()
    - Frozen panes and auto-filters
    - Conditional formatting for confidence
    - Data validation dropdowns
    - Indian number format (#,##,##0.00)
    - Multiple sheets per document type
    - Summary sheet with processing metadata
    - Audit trail sheet
    """

    def __init__(self):
        self._generators = EXCEL_GENERATORS

    def export(self, data: dict, doc_type: str = "generic_table",
               verification_meta: dict | None = None) -> bytes:
        """Generate an Excel export for the given document type.

        Args:
            data: Document data dict (rows, headers, metadata, etc.)
            doc_type: Document type ID (e.g. "gst_invoice", "ledger_sheet")
            verification_meta: Optional verification metadata for audit trail
                and draft warnings.

        Returns:
            Excel file bytes ready for download.
        """
        generator = self._get_excel_generator(doc_type)
        return generator(data, verification_meta)

    def _get_excel_generator(self, doc_type: str) -> Callable[[dict, dict | None], bytes]:
        """Resolve the Excel generator for a document type, falling back to generic."""
        return self._generators.get(doc_type, _generate_generic_excel)

    def list_supported_types(self) -> list[str]:
        """Return the list of supported document type IDs."""
        return list(self._generators.keys())


# Singleton for easy import
excel_export_engine = ExcelExportEngine()
