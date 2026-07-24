"""Universal, schema-driven Document AST + renderer registry.

The OCR extractor is the source of truth for *what* a document says. This module
owns *how* that content is presented, as a separate responsibility. It models a
document as a hierarchy -- Document -> Section -> Block -> Element -- where every
block keeps its true semantic type (a form stays a form, a table stays a table,
free text stays prose) all the way to the renderer.

This stops the old failure mode: previously every block was flattened to
headers+rows and rendered as one generic blue-headed grid, so forms, tables and
paragraphs were indistinguishable. Here each block type has its own renderer in a
registry, and unknown block types compose from the generic one -- so new document
shapes render sensibly with no code change.

build_ast adapts shapes the pipeline ALREADY produces (mixed sections, form
fields, plain headers/rows); it requires no extraction/prompt changes. Rendering
is defensive: a malformed block is skipped, never raised, so one bad region can't
blank the export. The caller keeps the flat-table fallback as an outer safety net.
"""
from __future__ import annotations

import json
from dataclasses import dataclass, field
from enum import Enum
from io import BytesIO
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from backend.table_scan import (
    HEADER_FILL,
    HEADER_FONT,
    HEADER_ALIGN,
    TEXT_FONT,
    TEXT_ALIGN,
    NUMBER_ALIGN,
    TOTAL_FILL,
    TOTAL_FONT,
    THIN_BORDER,
    DRAFT_WARNING_FILL,
    DRAFT_WARNING_FONT,
    _INDIAN_NUM_FORMAT,
    _excel_safe_value,
    _is_display_numeric,
    _build_totals_row,
    _clean_excel_header,
    _make_unique_headers,
)

FORM_LABEL_FONT = Font(name="Arial", size=10, bold=True, color="1F3864")
FORM_LABEL_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
SECTION_TITLE_FONT = Font(name="Arial", size=12, bold=True, color="1F3864")
DOC_TITLE_FONT = Font(name="Arial", size=14, bold=True)
TEXT_BLOCK_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)


class BlockType(str, Enum):
    """Semantic type of a block. Drives which renderer is used."""

    FORM = "form"
    TABLE = "table"
    TEXT = "text"
    HEADING = "heading"
    UNKNOWN = "unknown"


@dataclass
class Element:
    """A leaf node: a single cell, field, or line of text."""

    value: Any = ""
    label: "str | None" = None
    confidence: "float | None" = None
    numeric: bool = False
    hints: dict = field(default_factory=dict)


@dataclass
class Block:
    """A semantic block. Only fields relevant to its type are populated."""

    type: BlockType = BlockType.UNKNOWN
    title: "str | None" = None
    headers: list = field(default_factory=list)
    rows: list = field(default_factory=list)
    fields: list = field(default_factory=list)
    lines: list = field(default_factory=list)
    confidence: "float | None" = None
    hints: dict = field(default_factory=dict)


@dataclass
class Section:
    """A logical region. Rendered independently -- never merged into one grid."""

    title: "str | None" = None
    blocks: list = field(default_factory=list)


@dataclass
class Document:
    """The root of the AST."""

    title: str = "OCR Extraction"
    doc_type: "str | None" = None
    sections: list = field(default_factory=list)
    metadata: dict = field(default_factory=dict)


# -- Adapter: existing pipeline shapes -> AST --------------------------------

_FORM_LABEL_KEYS = ("label", "field", "name", "key")
_FORM_VALUE_KEYS = ("value", "val", "text")
_SECTION_META_KEYS = {"title", "name", "label", "heading", "section_title", "type", "kind"}
_MAX_SECTIONS = 60
_MAX_ROWS = 2000
_MAX_COLS = 60
_MAX_FIELDS = 200
_MAX_LINES = 400


def _s(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return str(value).strip()


def _field_from_dict(item: dict):
    label = next((_s(item[k]) for k in _FORM_LABEL_KEYS if item.get(k) not in (None, "")), None)
    value = next((item[k] for k in _FORM_VALUE_KEYS if k in item), None)
    if label is None and value is None:
        if len(item) == 1:
            (k, v), = item.items()
            label, value = _s(k), v
        else:
            return None
    return Element(label=label or "", value=_s(value))


def _classify_and_build_block(section: dict):
    """Turn one raw section dict into a typed Block, preserving its semantics."""
    raw_type = _s(section.get("type") or section.get("kind")).lower()

    fields_raw = section.get("fields")
    if raw_type == "form" or isinstance(fields_raw, list):
        fields = []
        for item in (fields_raw or [])[:_MAX_FIELDS]:
            if isinstance(item, dict):
                el = _field_from_dict(item)
                if el is not None:
                    fields.append(el)
            elif isinstance(item, (list, tuple)) and len(item) >= 2:
                fields.append(Element(label=_s(item[0]), value=_s(item[1])))
        if fields:
            return Block(type=BlockType.FORM, fields=fields)

    lines_raw = section.get("lines")
    if raw_type == "text" or isinstance(lines_raw, list):
        lines = [_s(line) for line in (lines_raw or [])[:_MAX_LINES] if _s(line)]
        if lines:
            return Block(type=BlockType.TEXT, lines=lines)

    headers_raw = section.get("headers")
    rows_raw = section.get("rows")
    if isinstance(headers_raw, list) or isinstance(rows_raw, list):
        headers = [_clean_excel_header(h, i + 1) for i, h in enumerate(headers_raw or [])][:_MAX_COLS]
        rows = []
        for raw_row in (rows_raw or [])[:_MAX_ROWS]:
            cells = raw_row if isinstance(raw_row, (list, tuple)) else [raw_row]
            row_els = []
            for c in list(cells)[:_MAX_COLS]:
                safe = _excel_safe_value(c)
                row_els.append(Element(value=safe, numeric=_is_display_numeric(safe)))
            rows.append(row_els)
        if headers or rows:
            return Block(type=BlockType.TABLE, headers=headers, rows=rows)

    scalar_pairs = [
        Element(label=_s(k), value=_s(v))
        for k, v in section.items()
        if k not in _SECTION_META_KEYS and not isinstance(v, (dict, list))
    ]
    if scalar_pairs:
        return Block(type=BlockType.FORM, fields=scalar_pairs)
    return None


def _section_title_of(section: Any, index: int) -> str:
    if isinstance(section, dict):
        for key in ("title", "name", "label", "heading", "section_title"):
            val = section.get(key)
            if val:
                return _s(val)
    return f"Section {index}"


def build_ast(payload: Any, *, doc_type=None, title=None) -> Document:
    """Adapt whatever the pipeline produced into a Document AST."""
    doc = Document(
        title=title or (_s(payload.get("title")) if isinstance(payload, dict) else "") or "OCR Extraction",
        doc_type=doc_type,
    )
    if isinstance(payload, dict) and isinstance(payload.get("metadata"), dict):
        doc.metadata = {
            str(k): _s(v)
            for k, v in payload["metadata"].items()
            if not isinstance(v, (dict, list))
        }

    sections_raw = None
    if isinstance(payload, dict):
        sections_raw = payload.get("sections") or payload.get("tables")
    elif isinstance(payload, list):
        sections_raw = payload

    if isinstance(sections_raw, list) and sections_raw:
        for i, sec in enumerate(sections_raw[:_MAX_SECTIONS], start=1):
            if not isinstance(sec, dict):
                continue
            block = _classify_and_build_block(sec)
            if block is None:
                continue
            block.title = _section_title_of(sec, i)
            doc.sections.append(Section(title=block.title, blocks=[block]))
        if doc.sections:
            return doc

    if isinstance(payload, dict) and (payload.get("headers") or payload.get("rows")):
        block = _classify_and_build_block(payload)
        if block is not None:
            doc.sections.append(Section(title=doc.title, blocks=[block]))
    return doc


# -- Excel renderer registry -------------------------------------------------

def _note_width(widths: dict, col: int, value: Any) -> None:
    widths[col] = max(widths.get(col, 10), len(str(value)) + 2)


def _row_width(block: Block) -> int:
    return max((len(r) for r in block.rows), default=0)


def _render_table_block(ws, block: Block, row: int, widths: dict, status) -> int:
    headers = _make_unique_headers(list(block.headers), max(len(block.headers), _row_width(block)))
    for c, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        _note_width(widths, c, header)
    row += 1
    plain_rows = []
    for row_els in block.rows:
        plain_rows.append([el.value for el in row_els])
        for c, el in enumerate(row_els, start=1):
            cell = ws.cell(row=row, column=c, value=el.value if el.value != "" else None)
            cell.font = TEXT_FONT
            cell.alignment = NUMBER_ALIGN if el.numeric else TEXT_ALIGN
            cell.border = THIN_BORDER
            if isinstance(el.value, (int, float)):
                cell.number_format = _INDIAN_NUM_FORMAT
            _note_width(widths, c, el.value)
        row += 1
    if len(plain_rows) > 1:
        totals = _build_totals_row(headers, plain_rows)
        if totals:
            for c, value in enumerate(totals, start=1):
                cell = ws.cell(row=row, column=c, value=value if value != "" else None)
                cell.fill = TOTAL_FILL
                cell.font = TOTAL_FONT
                cell.alignment = NUMBER_ALIGN if isinstance(value, (int, float)) else TEXT_ALIGN
                cell.border = THIN_BORDER
            row += 1
    return row


def _render_form_block(ws, block: Block, row: int, widths: dict, status) -> int:
    """Render key-value pairs as a two-column label/value layout -- NOT a table."""
    for el in block.fields:
        label_cell = ws.cell(row=row, column=1, value=_s(el.label))
        label_cell.font = FORM_LABEL_FONT
        label_cell.alignment = FORM_LABEL_ALIGN
        label_cell.border = THIN_BORDER
        _note_width(widths, 1, _s(el.label))

        value = _excel_safe_value(el.value)
        value_cell = ws.cell(row=row, column=2, value=value if value != "" else None)
        value_cell.font = TEXT_FONT
        value_cell.alignment = NUMBER_ALIGN if _is_display_numeric(value) else TEXT_ALIGN
        value_cell.border = THIN_BORDER
        if isinstance(value, (int, float)):
            value_cell.number_format = _INDIAN_NUM_FORMAT
        _note_width(widths, 2, value)
        row += 1
    return row


def _render_text_block(ws, block: Block, row: int, widths: dict, status) -> int:
    """Render free text as wrapped paragraphs, not a 1-column grid."""
    for line in block.lines:
        cell = ws.cell(row=row, column=1, value=line)
        cell.font = TEXT_FONT
        cell.alignment = TEXT_BLOCK_ALIGN
        _note_width(widths, 1, line[:80])
        row += 1
    return row


_BLOCK_EXCEL_RENDERERS = {
    BlockType.TABLE: _render_table_block,
    BlockType.FORM: _render_form_block,
    BlockType.TEXT: _render_text_block,
}


def _render_block(ws, block: Block, row: int, widths: dict, status) -> int:
    renderer = _BLOCK_EXCEL_RENDERERS.get(block.type, _render_table_block)
    return renderer(ws, block, row, widths, status)


def document_has_content(doc: Document) -> bool:
    """True if any block carries real data -- used to decide AST vs flat fallback."""
    for section in doc.sections:
        for block in section.blocks:
            if block.rows or block.fields or block.lines:
                return True
    return False


def render_document_to_excel(doc: Document, *, sheet_name="Extracted Data", verification_status=None) -> bytes:
    """Walk the AST and render each section/block by its true type."""
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name or "Extracted Data")[:31]
    widths = {}
    row = 1

    title_cell = ws.cell(row=row, column=1, value=doc.title)
    title_cell.font = DOC_TITLE_FONT
    title_cell.alignment = TEXT_ALIGN
    row += 1

    if verification_status in ("draft", "pending", "rejected"):
        banner = ws.cell(row=row, column=1)
        banner.value = (
            "⚠ DRAFT -- Not approved. Do not use for official records."
            if verification_status == "draft"
            else f"⚠ {verification_status.upper()} -- May contain unreviewed data."
        )
        banner.font = DRAFT_WARNING_FONT
        banner.fill = DRAFT_WARNING_FILL
        banner.alignment = TEXT_ALIGN
        row += 1
    row += 1

    for section in doc.sections:
        if section.title:
            sc = ws.cell(row=row, column=1, value=section.title)
            sc.font = SECTION_TITLE_FONT
            sc.alignment = TEXT_ALIGN
            _note_width(widths, 1, section.title)
            row += 1
        for block in section.blocks:
            try:
                row = _render_block(ws, block, row, widths, verification_status)
            except Exception:
                continue
        row += 1

    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(max(width, 10), 60)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
