"""OCR API router for logbook extraction."""

from __future__ import annotations

import base64
import json
import logging
import mimetypes
import os
import re
import shutil
import subprocess
import time
import uuid
from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path

import requests
from fastapi import APIRouter, Depends, File, Form, HTTPException, Response, UploadFile, status, Request
from fastapi.responses import FileResponse, JSONResponse
from itsdangerous import BadSignature, SignatureExpired, URLSafeTimedSerializer
from openpyxl import Workbook
from openpyxl.styles import Font
from PIL import Image, UnidentifiedImageError
from pydantic import BaseModel, Field
from sqlalchemy import false
from sqlalchemy.orm import Session
from anthropic import AuthenticationError, BadRequestError

from backend.database import SessionLocal, get_db, hash_ip_address
from backend.ledger_scan import (
    build_excel_bytes as ledger_build_excel_bytes,
    extract_data_from_image as ledger_extract_data,
    preprocess_image_bytes,
    validate_data as ledger_validate_data,
)
from backend.table_scan import build_table_excel_bytes
from backend.models.report import AuditLog
from backend.models.ocr_template import OcrTemplate
from backend.models.ocr_verification import OcrVerification
from backend.models.factory import Factory
from backend.ocr_utils import (
    analyze_image_quality,
    detect_column_centers,
    extract_table_from_image,
    warp_perspective,
)
from backend.security import get_current_user
from backend.rbac import is_manager_or_admin, require_any_role
from backend.models.user import User, UserRole
from backend.utils import PROJECT_ROOT, sanitize_text
from backend.ocr_limits import check_rate_limit, check_and_record_usage, check_and_record_org_usage, get_org_plan_for_usage
from backend.plans import has_plan_feature, min_plan_for_feature, org_has_ocr_access
from backend.services.background_jobs import (
    create_job,
    get_job as get_background_job,
    read_job_file,
    register_retry_handler,
    start_job,
    update_job,
    write_job_file,
)
from backend.services.ocr_document_pipeline import (
    build_structured_ocr_result,
    find_reusable_verification,
    serialize_reused_ocr_result,
)
from backend.services.ocr_normalization import normalize_structured_payload
from backend.tenancy import resolve_factory_id, resolve_org_id


logger = logging.getLogger(__name__)
router = APIRouter(tags=["OCR"])


OCR_VERIFICATION_DIR = PROJECT_ROOT / "exports" / "ocr_verifications"
_ALLOWED_VERIFICATION_STATUSES = {"draft", "pending", "approved", "rejected"}
_FILENAME_SAFE_RE = re.compile(r"[^A-Za-z0-9._-]+")
_OCR_SHARE_MAX_AGE_SECONDS = 7 * 24 * 60 * 60
_IMAGE_MAGIC_SIGNATURES: tuple[bytes, ...] = (
    b"\x89PNG\r\n\x1a\n",
    b"\xff\xd8\xff",
    b"GIF87a",
    b"GIF89a",
    b"BM",
    b"II*\x00",
    b"MM\x00*",
    b"RIFF",
)
_TABLE_EXCEL_SUPPORTED_MIME_TYPES = {
    "image/png",
    "image/jpeg",
    "image/webp",
    "image/gif",
}
_TABLE_EXCEL_NATIVE_MIME_TYPES = {"image/png", "image/jpeg"}
_TABLE_EXCEL_FORMAT_TO_MIME = {
    "PNG": "image/png",
    "JPEG": "image/jpeg",
    "WEBP": "image/webp",
    "GIF": "image/gif",
}
_TABLE_EXCEL_MODEL_HAIKU = "claude-haiku-4-5"
_TABLE_EXCEL_MODEL_SONNET = "claude-sonnet-5"
_TABLE_EXCEL_MODEL_OPUS = "claude-opus-4-7"
_TABLE_EXCEL_TIER_TO_MODEL = {
    "fast": _TABLE_EXCEL_MODEL_HAIKU,
    "balanced": _TABLE_EXCEL_MODEL_SONNET,
    "best": _TABLE_EXCEL_MODEL_OPUS,
}
_TABLE_EXCEL_MODEL_TO_TIER = {
    _TABLE_EXCEL_MODEL_HAIKU: "fast",
    _TABLE_EXCEL_MODEL_SONNET: "balanced",
    _TABLE_EXCEL_MODEL_OPUS: "best",
}
_TABLE_EXCEL_TIER_COSTS = {
    "fast": {"actual_cost_usd": 0.0003, "cost_saved_usd": 0.0137},
    "balanced": {"actual_cost_usd": 0.0030, "cost_saved_usd": 0.0110},
    "best": {"actual_cost_usd": 0.0150, "cost_saved_usd": 0.0},
}
_TABLE_EXCEL_MODEL_FALLBACKS = {
    "fast": [
        "claude-haiku-4-5",
        "claude-3-5-haiku-20241022",
        "claude-sonnet-5",
        "claude-3-5-sonnet-20241022",
    ],
    "balanced": [
        "claude-sonnet-5",
        "claude-3-5-sonnet-20241022",
        "claude-opus-4-7",
    ],
    "best": [
        "claude-opus-4-7",
        "claude-sonnet-5",
    ],
}
_TABLE_EXCEL_PROMPT = """You are a precise data extraction engine. Extract ALL data from this image into a structured JSON format suitable for Excel export.

Rules:
- If the image contains a table: return { "type": "table", "headers": [...], "rows": [[...], [...]] }
- If the image contains a form: return { "type": "form", "fields": [{ "label": "...", "value": "..." }] }
- If the image contains mixed content: return { "type": "mixed", "sections": [...] }
- If the image contains plain text/paragraphs: return { "type": "text", "lines": [...] }
- Preserve all numbers, dates, currencies exactly as they appear
- Do NOT add commentary, explanations, or markdown - return ONLY valid JSON"""
_ANTHROPIC_MESSAGES_URL = "https://api.anthropic.com/v1/messages"


class TableExcelRouteError(Exception):
    def __init__(self, status_code: int, payload: dict[str, object]):
        super().__init__(str(payload.get("error") or "Table Excel route failed."))
        self.status_code = status_code
        self.payload = payload


class OcrVerificationUpdatePayload(BaseModel):
    template_id: int | None = None
    source_filename: str | None = Field(default=None, max_length=255)
    columns: int | None = Field(default=None, ge=1, le=12)
    language: str | None = Field(default=None, max_length=20)
    avg_confidence: float | None = Field(default=None, ge=0, le=100)
    warnings: list[str] | None = None
    scan_quality: dict | None = None
    document_hash: str | None = Field(default=None, max_length=128)
    doc_type_hint: str | None = Field(default=None, max_length=80)
    routing_meta: dict | None = None
    raw_text: str | None = Field(default=None, max_length=50000)
    headers: list[str] | None = None
    original_rows: list[list[str]] | None = None
    reviewed_rows: list[list[str]] | None = None
    raw_column_added: bool | None = None
    reviewer_notes: str | None = Field(default=None, max_length=5000)


class OcrVerificationSubmitPayload(BaseModel):
    reviewer_notes: str | None = Field(default=None, max_length=5000)


class OcrVerificationDecisionPayload(BaseModel):
    reviewer_notes: str | None = Field(default=None, max_length=5000)
    rejection_reason: str | None = Field(default=None, max_length=5000)


def _safe_file_name(filename: str | None, default_stem: str = "ocr-source") -> str:
    raw = sanitize_text(filename, max_length=120, preserve_newlines=False) or default_stem
    safe = _FILENAME_SAFE_RE.sub("_", Path(raw).name).strip("._")
    return safe or default_stem


def _ocr_share_serializer() -> URLSafeTimedSerializer:
    secret = os.getenv("AUTH_RESET_SECRET") or os.getenv("JWT_SECRET_KEY")
    if not secret:
        raise RuntimeError("AUTH_RESET_SECRET or JWT_SECRET_KEY must be configured for OCR share links.")
    return URLSafeTimedSerializer(secret_key=secret, salt="ocr-share")


def _reject_mock_ocr() -> None:
    app_env = (os.getenv("APP_ENV") or "development").strip().lower()
    if app_env != "production":
        return
    if os.getenv("ALLOW_OCR_MOCK", "").strip().lower() not in {"1", "true", "yes", "on"}:
        raise HTTPException(status_code=403, detail="Mock OCR mode is disabled.")


def _validate_image_bytes(image_bytes: bytes) -> None:
    if not image_bytes:
        raise HTTPException(status_code=400, detail="Upload a valid image file.")
    if image_bytes.startswith(b"RIFF") and image_bytes[8:12] == b"WEBP":
        return
    if any(image_bytes.startswith(signature) for signature in _IMAGE_MAGIC_SIGNATURES):
        return
    if b"ftypheic" in image_bytes[:32] or b"ftypheif" in image_bytes[:32]:
        return
    raise HTTPException(status_code=400, detail="Upload a valid image file.")


async def _read_validated_image_upload(file: UploadFile) -> bytes:
    if not file.filename:
        raise HTTPException(status_code=400, detail="File is required.")
    if not (file.content_type or "").startswith("image/"):
        raise HTTPException(status_code=400, detail="Only image files are supported.")
    image_bytes = await file.read()
    if len(image_bytes) > 5_242_880:
        raise HTTPException(status_code=413, detail="Max 5MB image size exceeded.")
    _validate_image_bytes(image_bytes)
    return image_bytes


async def _read_image_upload_for_mock(file: UploadFile) -> bytes:
    if not file.filename:
        raise HTTPException(status_code=400, detail="File is required.")
    image_bytes = await file.read()
    if len(image_bytes) > 5_242_880:
        raise HTTPException(status_code=413, detail="Max 5MB image size exceeded.")
    if not image_bytes:
        raise HTTPException(status_code=400, detail="File is required.")
    return image_bytes


def _table_excel_error(status_code: int, message: str, **extra: object) -> TableExcelRouteError:
    payload: dict[str, object] = {"error": message}
    payload.update(extra)
    return TableExcelRouteError(status_code=status_code, payload=payload)


def _pick_table_excel_upload(file: UploadFile | None, image: UploadFile | None) -> UploadFile:
    upload = next(
        (candidate for candidate in (file, image) if candidate is not None and candidate.filename),
        file or image,
    )
    if upload is None:
        raise _table_excel_error(400, "Image file is required.")
    return upload


async def _read_table_excel_upload(
    file: UploadFile | None,
    image: UploadFile | None,
) -> tuple[UploadFile, bytes]:
    upload = _pick_table_excel_upload(file, image)
    try:
        image_bytes = await _read_validated_image_upload(upload)
    except HTTPException as error:
        detail = error.detail if isinstance(error.detail, str) else "Invalid image upload."
        raise _table_excel_error(int(error.status_code), detail) from error
    return upload, image_bytes


def _table_excel_timeout_seconds() -> float:
    raw = (
        os.getenv("TABLE_EXCEL_PROVIDER_TIMEOUT_SECONDS")
        or os.getenv("TABLE_SCAN_PROVIDER_TIMEOUT_SECONDS")
        or os.getenv("OCR_PROVIDER_TIMEOUT_SECONDS")
        or "45"
    ).strip()
    try:
        value = float(raw)
    except ValueError:
        value = 45.0
    return max(5.0, min(120.0, value))


def _require_anthropic_api_key() -> str:
    api_key = (os.getenv("ANTHROPIC_API_KEY") or "").strip()
    if not api_key:
        logger.error("[OCR] API key: missing")
        raise _table_excel_error(
            500,
            "ANTHROPIC_API_KEY is not configured. Structured OCR requires a valid Anthropic API key.",
        )
    logger.info("[OCR] API key: present")
    return api_key


def _validate_table_excel_model_name(selected_model: str) -> str:
    normalized = sanitize_text(selected_model, max_length=80, preserve_newlines=False) or ""
    allowed_models = {
        candidate
        for candidates in _TABLE_EXCEL_MODEL_FALLBACKS.values()
        for candidate in candidates
    }
    if normalized not in allowed_models:
        logger.error("[OCR] Error: unsupported Anthropic model %s", selected_model)
        raise _table_excel_error(500, f"Unsupported Anthropic model configured for OCR: {selected_model}")
    return normalized


def _normalize_table_excel_mime_type(content_type: str | None, filename: str | None) -> str | None:
    normalized = (content_type or "").strip().lower()
    if normalized == "image/jpg":
        return "image/jpeg"
    if normalized in _TABLE_EXCEL_SUPPORTED_MIME_TYPES:
        return normalized
    guessed, _ = mimetypes.guess_type(filename or "")
    guessed_normalized = (guessed or "").strip().lower()
    if guessed_normalized == "image/jpg":
        return "image/jpeg"
    return guessed_normalized or None


def _inspect_table_excel_image(
    image_bytes: bytes,
    *,
    content_type: str | None,
    filename: str | None,
) -> dict[str, object]:
    try:
        with Image.open(BytesIO(image_bytes)) as img:
            width, height = img.size
            mime_type = _TABLE_EXCEL_FORMAT_TO_MIME.get((img.format or "").upper())
    except (UnidentifiedImageError, OSError) as error:
        raise _table_excel_error(400, "Upload a supported PNG, JPG, JPEG, WEBP, or GIF image.") from error

    mime_type = mime_type or _normalize_table_excel_mime_type(content_type, filename)
    if mime_type not in _TABLE_EXCEL_SUPPORTED_MIME_TYPES:
        raise _table_excel_error(400, "Upload a supported PNG, JPG, JPEG, WEBP, or GIF image.")

    # Relaxed scoring to allow modern compressed formats and small crops
    if len(image_bytes) < 2 * 1024:
        score = 5  # Truly invalid or near-empty
    elif len(image_bytes) < 20 * 1024:
        score = 20
    elif width < 150 or height < 150:
        score = 15
    elif width < 200 or height < 200:
        score = 30
    elif width > 600 or height > 600:
        score = 90
    else:
        score = 60

    if mime_type not in _TABLE_EXCEL_NATIVE_MIME_TYPES:
        score -= 5

    score = max(0, min(100, score))
    logger.info(
        "[OCR] Image inspection: score=%s size=%s width=%s height=%s mime=%s",
        score,
        len(image_bytes),
        width,
        height,
        mime_type,
    )

    return {
        "image_quality_score": score,
        "image_mime_type": mime_type,
        "width": width,
        "height": height,
        "size_bytes": len(image_bytes),
    }


def _select_table_excel_model(image_quality_score: int) -> str:
    if image_quality_score >= 60:
        return _TABLE_EXCEL_MODEL_HAIKU
    return _TABLE_EXCEL_MODEL_SONNET


def _normalize_force_model_tier(force_model: str | None) -> str | None:
    normalized = sanitize_text(force_model, max_length=20, preserve_newlines=False)
    if normalized:
        normalized = normalized.lower()
    if normalized in {"fast", "balanced", "best"}:
        return normalized
    return None


def _select_table_preview_model(image_quality_score: int, *, force_model: str | None) -> tuple[str, bool]:
    forced_tier = _normalize_force_model_tier(force_model)
    if forced_tier:
        return _TABLE_EXCEL_TIER_TO_MODEL[forced_tier], True
    return _select_table_excel_model(image_quality_score), False


def _table_excel_model_candidates(selected_model: str) -> list[str]:
    selected_model = _validate_table_excel_model_name(selected_model)
    tier = _TABLE_EXCEL_MODEL_TO_TIER.get(selected_model)
    if tier:
        candidates = _TABLE_EXCEL_MODEL_FALLBACKS.get(tier, [])
    else:
        candidates = []
        for options in _TABLE_EXCEL_MODEL_FALLBACKS.values():
            if selected_model in options:
                candidates = options
                break
    ordered = [selected_model, *candidates]
    seen: set[str] = set()
    unique: list[str] = []
    for model_name in ordered:
        if model_name and model_name not in seen:
            seen.add(model_name)
            unique.append(model_name)
    return unique


def _is_model_selection_error(message: str) -> bool:
    normalized = (message or "").strip().lower()
    return "model" in normalized and any(
        token in normalized
        for token in {
            "not found",
            "does not exist",
            "not available",
            "not supported",
            "invalid model",
            "access",
            "permission",
        }
    )


def _table_preview_doc_type(doc_type_hint: str | None) -> str:
    normalized = sanitize_text(doc_type_hint, max_length=80, preserve_newlines=False)
    return normalized.lower() if normalized else "table"


def _table_preview_title(doc_type_hint: str | None, template: OcrTemplate | None) -> str:
    if template and template.name:
        return template.name
    doc_type = _table_preview_doc_type(doc_type_hint)
    if not doc_type:
        return "OCR Extraction"
    return doc_type.replace("-", " ").replace("_", " ").title()


def _flatten_preview_rows(rows: list[list[str]]) -> str | None:
    parts = [" | ".join(cell for cell in row if cell) for row in rows]
    joined = "\n".join(part for part in parts if part.strip())
    return joined or None


def _table_excel_prompt_text(system_prompt: str | None, user_message: str | None) -> str:
    extras: list[str] = []
    if system_prompt:
        extras.append(f"Additional caller context: {system_prompt.strip()}")
    if user_message:
        extras.append(f"Additional caller request: {user_message.strip()}")
    if not extras:
        return _TABLE_EXCEL_PROMPT
    return f"{_TABLE_EXCEL_PROMPT}\n\n" + "\n".join(extras)


def _validate_table_excel_json(data: dict | None) -> list[str]:
    if not isinstance(data, dict):
        return ["AI response is not a valid JSON object."]
    
    extracted_type = data.get("type")
    if not extracted_type:
        return ["Missing 'type' field in AI response."]
    
    errors = []
    if extracted_type == "table":
        headers = data.get("headers")
        rows = data.get("rows")
        if not isinstance(headers, list) or not headers:
            errors.append("Table 'headers' is missing or empty.")
        if not isinstance(rows, list):
            errors.append("Table 'rows' is missing.")
        elif headers:
            header_len = len(headers)
            for i, row in enumerate(rows):
                if not isinstance(row, list) or len(row) != header_len:
                    errors.append(f"Row {i+1} length mismatch (expected {header_len} columns).")
    
    elif extracted_type == "form":
        fields = data.get("fields")
        if not isinstance(fields, list) or not fields:
            errors.append("Form 'fields' is missing or empty.")
    
    return errors


def _extract_table_excel_json_text(ai_data: object) -> str:
    if not isinstance(ai_data, dict):
        raise _table_excel_error(502, "Anthropic API returned an invalid response.")
    error = ai_data.get("error")
    if isinstance(error, dict):
        raise _table_excel_error(502, str(error.get("message") or "Anthropic API request failed."))
    content = ai_data.get("content")
    if not isinstance(content, list):
        raise _table_excel_error(502, "Anthropic API returned no message content.")
    text_blocks = [
        str(block.get("text") or "")
        for block in content
        if isinstance(block, dict) and str(block.get("type") or "") == "text"
    ]
    text = "\n".join(item for item in text_blocks if item).strip()
    if not text:
        raise _table_excel_error(502, "Anthropic API returned an empty extraction result.")
    return text


def _extract_json_candidate(raw: str) -> object:
    stripped = raw.strip()
    try:
        return json.loads(stripped)
    except json.JSONDecodeError:
        pass

    opener = stripped.find("{")
    if opener != -1:
        closer = stripped.rfind("}")
        if closer > opener:
            candidate = stripped[opener : closer + 1]
            try:
                return json.loads(candidate)
            except json.JSONDecodeError:
                pass

    opener = stripped.find("[")
    if opener != -1:
        closer = stripped.rfind("]")
        if closer > opener:
            candidate = stripped[opener : closer + 1]
            try:
                return json.loads(candidate)
            except json.JSONDecodeError:
                pass

    raise _table_excel_error(502, "Anthropic API returned invalid JSON.")


def _call_table_excel_anthropic(
    image_base64: str,
    *,
    image_mime_type: str,
    selected_model: str,
    system_prompt: str | None,
    user_message: str | None,
) -> dict[str, object]:
    api_key = _require_anthropic_api_key()
    model_candidates = _table_excel_model_candidates(selected_model)
    logger.info("[OCR] Starting extraction model=%s base64_len=%s", model_candidates[0], len(image_base64))

    last_error: TableExcelRouteError | None = None
    
    # Pass 1: Extraction with cost-optimized model
    extraction_json = None
    first_model_used = None
    
    for model_name in model_candidates:
        payload = {
            "model": model_name,
            "max_tokens": 4096,
            "system": [
                {
                    "type": "text",
                    "text": _table_excel_prompt_text(system_prompt, user_message),
                    "cache_control": {"type": "ephemeral"},
                }
            ],
            "messages": [
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image",
                            "source": {
                                "type": "base64",
                                "media_type": image_mime_type,
                                "data": image_base64,
                            },
                        },
                    ],
                }
            ],
        }

        try:
            logger.info("[OCR] Pass 1: Sending request model=%s", model_name)
            response = requests.post(
                _ANTHROPIC_MESSAGES_URL,
                headers={
                    "Content-Type": "application/json",
                    "x-api-key": api_key,
                    "anthropic-version": "2023-06-01",
                },
                json=payload,
                timeout=_table_excel_timeout_seconds(),
            )
        except requests.RequestException as error:
            logger.error("[OCR] Pass 1 connection error: %s", error)
            last_error = _table_excel_error(502, f"Anthropic API request failed: {error}")
            continue

        if response.status_code >= 400:
            try:
                error_data = response.json()
                message = error_data.get("error", {}).get("message") or "Anthropic API request failed."
            except ValueError:
                message = f"Anthropic API returned status {response.status_code}."
            
            logger.error("[OCR] Pass 1 error status=%s message=%s", response.status_code, message)
            
            # Resilient fallback: Retry on any 4xx/5xx that isn't fatal (auth or too large)
            if response.status_code in {400, 404, 429, 500, 502, 503, 504}:
                # 413 (Payload Too Large) is NOT retried as subsequent models won't help
                if response.status_code == 413:
                    raise _table_excel_error(413, "Image too large for Anthropic API.")
                
                last_error = _table_excel_error(response.status_code, message, model=model_name)
                continue
            
            # Non-retryable error (e.g. 401, 403)
            raise _table_excel_error(response.status_code, message)

        try:
            ai_data = response.json()
            raw_text = _extract_table_excel_json_text(ai_data)
            extraction_json = _extract_json_candidate(raw_text)
            first_model_used = model_name
            break # Success
        except (ValueError, TableExcelRouteError) as error:
            logger.error("[OCR] Pass 1 parse error: %s", error)
            last_error = _table_excel_error(502, f"Failed to parse AI response: {error}")
            continue

    if not extraction_json:
        if last_error:
            raise last_error
        raise _table_excel_error(502, "Extraction failed for all configured models.")

    # Validation & Correction Pass (Pass 2)
    validation_errors = _validate_table_excel_json(extraction_json)
    if validation_errors:
        logger.warning("[OCR] Validation failed for %s: %s", first_model_used, validation_errors)
        
        # Multi-turn retry using Opus (Text-only correction)
        correction_prompt = (
            f"Your previous response had structural inconsistencies:\n"
            f"- " + "\n- ".join(validation_errors) + "\n\n"
            "Fix the JSON structure. Do not change values. Only correct formatting and alignment. "
            "Return ONLY the fixed JSON object."
        )
        
        retry_payload = {
            "model": _TABLE_EXCEL_MODEL_OPUS,
            "max_tokens": 4096,
            "messages": [
                {"role": "user", "content": "Extract data from an image (provided in previous context)."},
                {"role": "assistant", "content": json.dumps(extraction_json)},
                {"role": "user", "content": correction_prompt}
            ]
        }
        
        try:
            logger.info("[OCR] Pass 2: Sending correction request model=%s", _TABLE_EXCEL_MODEL_OPUS)
            response = requests.post(
                _ANTHROPIC_MESSAGES_URL,
                headers={
                    "Content-Type": "application/json",
                    "x-api-key": api_key,
                    "anthropic-version": "2023-06-01",
                },
                json=retry_payload,
                timeout=_table_excel_timeout_seconds(),
            )
            if response.status_code == 200:
                ai_data = response.json()
                raw_text = _extract_table_excel_json_text(ai_data)
                corrected_json = _extract_json_candidate(raw_text)
                if isinstance(corrected_json, dict):
                    logger.info("[OCR] Correction pass successful model=%s", _TABLE_EXCEL_MODEL_OPUS)
                    corrected_json.setdefault("_provider_model", _TABLE_EXCEL_MODEL_OPUS)
                    corrected_json.setdefault("_correction_applied", True)
                    return corrected_json
                logger.error("[OCR] Correction pass returned invalid JSON")
            else:
                logger.error("[OCR] Correction pass failed status=%s", response.status_code)
        except Exception as error:
            logger.error("[OCR] Correction pass failed unexpectedly: %s", error)

    # Return Pass 1 result if correction was not triggered or failed
    extraction_json.setdefault("_provider_model", first_model_used)
    return extraction_json


def _normalize_table_excel_value(value: object) -> str:
    return "" if value is None else str(value)


def _normalize_table_excel_extracted_json(extracted_json: dict[str, object]) -> dict[str, object]:
    extracted_type = str(extracted_json.get("type") or "").strip().lower()
    if extracted_type not in {"table", "form", "mixed", "text"}:
        raise _table_excel_error(502, "Anthropic API returned an unsupported extraction type.")

    if extracted_type == "table":
        raw_headers = extracted_json.get("headers")
        raw_rows = extracted_json.get("rows")
        headers = [
            _normalize_table_excel_value(header)
            for header in raw_headers
        ] if isinstance(raw_headers, list) else []
        rows: list[list[str]] = []
        max_columns = len(headers)
        if isinstance(raw_rows, list):
            for row in raw_rows:
                if not isinstance(row, list):
                    continue
                normalized_row = [_normalize_table_excel_value(cell) for cell in row]
                rows.append(normalized_row)
                max_columns = max(max_columns, len(normalized_row))
        if max_columns == 0:
            raise _table_excel_error(502, "Anthropic API did not return any extractable table data.")
        if not headers:
            headers = [f"Column {index}" for index in range(1, max_columns + 1)]
        for row in rows:
            if len(row) < len(headers):
                row.extend([""] * (len(headers) - len(row)))
        return {"type": "table", "headers": headers, "rows": rows}

    if extracted_type == "form":
        raw_fields = extracted_json.get("fields")
        fields: list[dict[str, str]] = []
        if isinstance(raw_fields, list):
            for field in raw_fields:
                if not isinstance(field, dict):
                    continue
                fields.append(
                    {
                        "label": _normalize_table_excel_value(field.get("label")),
                        "value": _normalize_table_excel_value(field.get("value")),
                    }
                )
        return {"type": "form", "fields": fields}

    if extracted_type == "text":
        raw_lines = extracted_json.get("lines")
        lines = [_normalize_table_excel_value(line) for line in raw_lines] if isinstance(raw_lines, list) else []
        return {"type": "text", "lines": lines}

    raw_sections = extracted_json.get("sections")
    sections = list(raw_sections) if isinstance(raw_sections, list) else []
    return {"type": "mixed", "sections": sections}


def _build_table_preview_payload(
    extracted_json: dict[str, object],
    *,
    template: OcrTemplate | None,
    doc_type_hint: str | None,
) -> dict[str, object]:
    normalized = _normalize_table_excel_extracted_json(extracted_json)
    extracted_type = str(normalized["type"])

    if extracted_type == "table":
        headers = [str(value) for value in normalized.get("headers", [])]
        rows = [[str(cell) for cell in row] for row in normalized.get("rows", [])]
    elif extracted_type == "form":
        headers = ["Field", "Value"]
        rows = [
            [
                str(field.get("label") or ""),
                str(field.get("value") or ""),
            ]
            for field in normalized.get("fields", [])
            if isinstance(field, dict)
        ]
    elif extracted_type == "text":
        headers = ["Text"]
        rows = [[str(line)] for line in normalized.get("lines", [])]
    else:
        headers = ["Section"]
        rows = [
            [json.dumps(section, ensure_ascii=True)]
            for section in normalized.get("sections", [])
        ]

    fallback_headers = list(template.column_names or []) if template and template.column_names else headers
    structured = normalize_structured_payload(
        {
            "type": _table_preview_doc_type(doc_type_hint),
            "title": _table_preview_title(doc_type_hint, template),
            "headers": headers,
            "rows": rows,
        },
        fallback_headers=fallback_headers,
        fallback_rows=rows,
        fallback_type=_table_preview_doc_type(doc_type_hint),
        fallback_title=_table_preview_title(doc_type_hint, template),
    )
    structured["warnings"] = [
        str(value).strip()
        for value in extracted_json.get("warnings", [])
        if str(value).strip()
    ] if isinstance(extracted_json.get("warnings"), list) else []
    if not structured.get("rows"):
        raise _table_excel_error(502, "Anthropic API did not return any extractable data for this scan.")
    if not structured.get("raw_text"):
        structured["raw_text"] = _flatten_preview_rows(structured.get("rows") or [])
    return structured


def _auto_fit_openpyxl_columns(sheet) -> None:
    for column_cells in sheet.columns:
        max_len = 10
        for cell in column_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)) + 2)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max_len, 50)


def _build_table_excel_workbook(extracted_json: dict[str, object]) -> tuple[bytes, dict[str, object]]:
    normalized = _normalize_table_excel_extracted_json(extracted_json)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Extracted Data"

    extracted_type = str(normalized["type"])
    total_rows = 0
    total_columns = 0

    try:
        if extracted_type == "table":
            headers = [str(value) for value in normalized.get("headers", [])]
            rows = [[str(cell) for cell in row] for row in normalized.get("rows", [])]
            sheet.append(headers)
            sheet[1][0].font = Font(bold=True)
            for cell in sheet[1]:
                cell.font = Font(bold=True)
            for row in rows:
                sheet.append(row)
            total_rows = len(rows)
            total_columns = len(headers)
        elif extracted_type == "form":
            sheet.append(["Field", "Value"])
            for cell in sheet[1]:
                cell.font = Font(bold=True)
            fields = normalized.get("fields", [])
            for field in fields:
                label = str(field.get("label") or "") if isinstance(field, dict) else ""
                value = str(field.get("value") or "") if isinstance(field, dict) else ""
                sheet.append([label, value])
            total_rows = len(fields)
            total_columns = 2
        elif extracted_type == "text":
            lines = [str(line) for line in normalized.get("lines", [])]
            for line in lines:
                sheet.append([line])
            total_rows = len(lines)
            total_columns = 1
        else:
            sections = normalized.get("sections", [])
            for section in sections:
                sheet.append([json.dumps(section, ensure_ascii=False)])
            total_rows = len(sections)
            total_columns = 1

        if sheet.max_row == 1 and sheet.max_column == 1 and sheet["A1"].value is None:
            sheet["A1"] = "No extractable data found"
            total_rows = 0
            total_columns = 1

        _auto_fit_openpyxl_columns(sheet)
        output = BytesIO()
        workbook.save(output)
        excel_bytes = output.getvalue()
        if not excel_bytes:
            raise RuntimeError("Workbook writer returned empty output.")
    except TableExcelRouteError:
        raise
    except Exception as error:
        logger.error("[OCR] Error: %s", error)
        raise _table_excel_error(500, f"Excel generation failed: {error}") from error

    return excel_bytes, {
        "total_rows": total_rows,
        "total_columns": total_columns,
        "extracted_type": extracted_type,
    }


def _run_table_excel_pipeline(
    image_bytes: bytes,
    *,
    content_type: str | None,
    filename: str | None,
    system_prompt: str | None,
    user_message: str | None,
) -> tuple[bytes, dict[str, object]]:
    started_at = time.perf_counter()
    logger.info("[OCR] Flow: upload -> /ocr/table-excel -> _run_table_excel_pipeline")
    inspection = _inspect_table_excel_image(image_bytes, content_type=content_type, filename=filename)
    image_quality_score = int(inspection["image_quality_score"])
    if image_quality_score < 10:
        raise _table_excel_error(
            400,
            "Image too vague or low quality to process",
            imageQualityScore=image_quality_score,
        )

    selected_model = _select_table_excel_model(image_quality_score)
    logger.info("[OCR] Using AI")
    logger.info(
        "Table Excel model selected model=%s quality_score=%s mime=%s size_bytes=%s",
        selected_model,
        image_quality_score,
        inspection["image_mime_type"],
        inspection["size_bytes"],
    )

    # Harden vision payload with resizing
    image_base64 = preprocess_image_bytes(image_bytes)

    extracted_json = _call_table_excel_anthropic(
        image_base64,
        image_mime_type="image/jpeg", # preprocess_image_bytes always returns JPEG
        selected_model=selected_model,
        system_prompt=system_prompt,
        user_message=user_message,
    )
    used_model = str(extracted_json.get("_provider_model") or selected_model)
    excel_bytes, metadata = _build_table_excel_workbook(extracted_json)
    elapsed_ms = int((time.perf_counter() - started_at) * 1000)
    logger.info(
        "Table Excel completed model=%s quality_score=%s elapsed_ms=%s extracted_type=%s",
        used_model,
        image_quality_score,
        elapsed_ms,
        metadata.get("extracted_type"),
    )
    metadata.update(
        {
            "image_quality_score": image_quality_score,
            "model_used": used_model,
            "time_taken_ms": elapsed_ms,
            "image_mime_type": inspection["image_mime_type"],
        }
    )
    return excel_bytes, metadata


def _run_table_preview_pipeline(
    image_bytes: bytes,
    *,
    content_type: str | None,
    filename: str | None,
    template: OcrTemplate | None,
    doc_type_hint: str | None,
    force_model: str | None,
    language: str,
) -> dict[str, object]:
    started_at = time.perf_counter()
    logger.info("[OCR] Flow: upload -> /ocr/logbook -> _run_table_preview_pipeline")
    inspection = _inspect_table_excel_image(image_bytes, content_type=content_type, filename=filename)
    image_quality_score = int(inspection["image_quality_score"])
    if image_quality_score < 10:
        raise _table_excel_error(
            400,
            "Image too vague or low quality to process",
            imageQualityScore=image_quality_score,
        )

    selected_model, forced = _select_table_preview_model(image_quality_score, force_model=force_model)
    logger.info("[OCR] Using AI")
    model_tier = _TABLE_EXCEL_MODEL_TO_TIER.get(selected_model, "balanced")
    logger.info(
        "Structured table preview model selected model=%s quality_score=%s mime=%s size_bytes=%s",
        selected_model,
        image_quality_score,
        inspection["image_mime_type"],
        inspection["size_bytes"],
    )

    # Harden vision payload with resizing
    image_base64 = preprocess_image_bytes(image_bytes)

    extracted_json = _call_table_excel_anthropic(
        image_base64,
        image_mime_type="image/jpeg", # preprocess_image_bytes always returns JPEG
        selected_model=selected_model,
        system_prompt=None,
        user_message=None,
    )
    used_model = str(extracted_json.get("_provider_model") or selected_model)
    structured = _build_table_preview_payload(
        extracted_json,
        template=template,
        doc_type_hint=doc_type_hint,
    )

    elapsed_ms = int((time.perf_counter() - started_at) * 1000)
    cost_meta = _TABLE_EXCEL_TIER_COSTS.get(model_tier, {"actual_cost_usd": 0.0, "cost_saved_usd": 0.0})
    headers = structured.get("headers") or []
    rows = structured.get("rows") or []
    logger.info(
        "Structured table preview completed model=%s quality_score=%s elapsed_ms=%s rows=%s columns=%s",
        used_model,
        image_quality_score,
        elapsed_ms,
        len(rows),
        len(headers),
    )
    return {
        "type": structured.get("type") or _table_preview_doc_type(doc_type_hint),
        "title": structured.get("title") or _table_preview_title(doc_type_hint, template),
        "headers": headers,
        "rows": rows,
        "raw_text": structured.get("raw_text"),
        "language": language,
        "confidence": float(image_quality_score),
        "warnings": structured.get("warnings") or [],
        "routing": {
            "clarity_score": float(image_quality_score),
            "score_reason": f"Claude vision extracted structured {structured.get('type') or 'table'} content directly from the uploaded image.",
            "model_tier": model_tier,
            "forced": forced,
            "scorer_used": True,
            "actual_cost_usd": float(cost_meta["actual_cost_usd"]),
            "cost_saved_usd": float(cost_meta["cost_saved_usd"]),
            "provider_used": "anthropic",
            "provider_model": used_model,
            "ai_applied": True,
            "ai_attempted": True,
            "ai_degraded_to_base": False,
        },
        "columns": max(len(headers), max((len(row) for row in rows), default=0), 1),
        "avg_confidence": float(image_quality_score),
        "cell_confidence": [],
        "cell_boxes": [],
        "used_language": language,
        "fallback_used": False,
        "raw_column_added": False,
        "reused": False,
        "reused_verification_id": None,
    }


def _table_excel_response_headers(metadata: dict[str, object], *, filename: str) -> dict[str, str]:
    return {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Cache-Control": "no-store",
        "Pragma": "no-cache",
        "X-Total-Rows": str(metadata.get("total_rows", 0)),
        "X-Total-Columns": str(metadata.get("total_columns", 0)),
        "X-Image-Quality-Score": str(metadata.get("image_quality_score", "")),
        "X-Model-Used": str(metadata.get("model_used", "")),
    }


def _build_ocr_share_token(verification: OcrVerification) -> str:
    return _ocr_share_serializer().dumps(
        {
            "verification_id": verification.id,
            "org_id": verification.org_id,
            "factory_id": verification.factory_id,
        }
    )


def _read_ocr_share_token(token: str) -> dict:
    try:
        return _ocr_share_serializer().loads(token, max_age=_OCR_SHARE_MAX_AGE_SECONDS)
    except SignatureExpired as error:
        raise HTTPException(status_code=410, detail="Share link expired.") from error
    except BadSignature as error:
        raise HTTPException(status_code=404, detail="Share link invalid.") from error


def _parse_json_value(raw: str | None, *, field_name: str):
    if raw in (None, ""):
        return None
    try:
        return json.loads(raw)
    except json.JSONDecodeError as error:
        raise HTTPException(status_code=400, detail=f"Invalid JSON for {field_name}.") from error


def _normalize_string_list(values: list | None, *, field_name: str) -> list[str] | None:
    if values is None:
        return None
    if not isinstance(values, list):
        raise HTTPException(status_code=400, detail=f"{field_name} must be a list.")
    cleaned: list[str] = []
    for value in values:
        text = sanitize_text(str(value) if value is not None else "", max_length=255, preserve_newlines=False) or ""
        cleaned.append(text)
    return cleaned


def _normalize_rows(values: list | None, *, field_name: str) -> list[list[str]] | None:
    if values is None:
        return None
    if not isinstance(values, list):
        raise HTTPException(status_code=400, detail=f"{field_name} must be a list of rows.")
    normalized: list[list[str]] = []
    max_columns = 0
    for row in values:
        if not isinstance(row, list):
            raise HTTPException(status_code=400, detail=f"{field_name} must be a list of rows.")
        cleaned_row = [
            sanitize_text(str(cell) if cell is not None else "", max_length=2000, preserve_newlines=False) or ""
            for cell in row
        ]
        normalized.append(cleaned_row)
        max_columns = max(max_columns, len(cleaned_row))
    for row in normalized:
        if len(row) < max_columns:
            row.extend([""] * (max_columns - len(row)))
    return normalized


def _normalize_scan_quality(values: dict | None, *, field_name: str) -> dict | None:
    if values is None:
        return None
    if not isinstance(values, dict):
        raise HTTPException(status_code=400, detail=f"{field_name} must be an object.")

    band = sanitize_text(str(values.get("confidence_band") or ""), max_length=20, preserve_newlines=False) or "unknown"
    if band not in {"high", "medium", "low", "unknown"}:
        band = "unknown"

    quality_signals = _normalize_string_list(values.get("quality_signals"), field_name=f"{field_name}.quality_signals") or []
    auto_processing = _normalize_string_list(values.get("auto_processing"), field_name=f"{field_name}.auto_processing") or []
    notes = sanitize_text(str(values.get("notes") or ""), max_length=500, preserve_newlines=False) or None
    outcome = sanitize_text(str(values.get("outcome") or ""), max_length=20, preserve_newlines=False) or "success"
    if outcome not in {"success", "partial", "failed"}:
        outcome = "success"
    next_action = sanitize_text(str(values.get("next_action") or ""), max_length=40, preserve_newlines=False) or None
    fallback_used = bool(values.get("fallback_used"))

    def _safe_int(key: str, default: int = 0, minimum: int = 0, maximum: int = 999) -> int:
        raw = values.get(key)
        try:
            parsed = int(raw)
        except (TypeError, ValueError):
            return default
        return max(minimum, min(maximum, parsed))

    raw_boxes = values.get("cell_boxes")
    cell_boxes: list[list[dict | None]] = []
    if isinstance(raw_boxes, list):
        for row in raw_boxes:
            if not isinstance(row, list):
                continue
            normalized_row: list[dict | None] = []
            for box in row:
                if not isinstance(box, dict):
                    normalized_row.append(None)
                    continue
                try:
                    x = float(box.get("x", 0.0))
                    y = float(box.get("y", 0.0))
                    width = float(box.get("width", 0.0))
                    height = float(box.get("height", 0.0))
                except (TypeError, ValueError):
                    normalized_row.append(None)
                    continue
                normalized_row.append(
                    {
                        "x": max(0.0, min(1.0, x)),
                        "y": max(0.0, min(1.0, y)),
                        "width": max(0.0, min(1.0, width)),
                        "height": max(0.0, min(1.0, height)),
                    }
                )
            cell_boxes.append(normalized_row)

    return {
        "confidence_band": band,
        "quality_signals": quality_signals,
        "auto_processing": auto_processing,
        "fallback_used": fallback_used,
        "correction_count": _safe_int("correction_count"),
        "page_count": _safe_int("page_count", default=1, minimum=1, maximum=100),
        "adjustment_count": _safe_int("adjustment_count"),
        "retake_count": _safe_int("retake_count"),
        "manual_review_recommended": bool(values.get("manual_review_recommended")),
        "outcome": outcome,
        "next_action": next_action,
        "notes": notes,
        "cell_boxes": cell_boxes or None,
    }


def _normalize_document_hash(value: str | None) -> str | None:
    normalized = sanitize_text(value, max_length=128, preserve_newlines=False)
    return normalized.lower() if normalized else None


def _normalize_doc_type_hint(value: str | None) -> str | None:
    normalized = sanitize_text(value, max_length=80, preserve_newlines=False)
    return normalized.lower() if normalized else None


def _normalize_routing_meta(values: dict | None, *, field_name: str) -> dict | None:
    if values is None:
        return None
    if not isinstance(values, dict):
        raise HTTPException(status_code=400, detail=f"{field_name} must be an object.")

    model_tier = sanitize_text(str(values.get("model_tier") or ""), max_length=20, preserve_newlines=False) or "fast"
    if model_tier not in {"fast", "balanced", "best"}:
        model_tier = "fast"
    score_reason = sanitize_text(str(values.get("score_reason") or ""), max_length=500, preserve_newlines=False) or None
    provider_used = sanitize_text(str(values.get("provider_used") or ""), max_length=40, preserve_newlines=False) or None
    provider_model = sanitize_text(str(values.get("provider_model") or ""), max_length=120, preserve_newlines=False) or None
    ai_failure_reason = sanitize_text(
        str(values.get("ai_failure_reason") or ""),
        max_length=80,
        preserve_newlines=False,
    ) or None

    def _safe_float(key: str, default: float = 0.0, minimum: float = 0.0, maximum: float = 100.0) -> float:
        raw = values.get(key)
        try:
            parsed = float(raw)
        except (TypeError, ValueError):
            return default
        return max(minimum, min(maximum, parsed))

    return {
        "clarity_score": _safe_float("clarity_score"),
        "score_reason": score_reason,
        "model_tier": model_tier,
        "forced": bool(values.get("forced")),
        "scorer_used": bool(values.get("scorer_used")),
        "actual_cost_usd": _safe_float("actual_cost_usd", maximum=1000.0),
        "cost_saved_usd": _safe_float("cost_saved_usd", maximum=1000.0),
        "provider_used": provider_used,
        "provider_model": provider_model,
        "ai_applied": bool(values.get("ai_applied")),
        "ai_attempted": bool(values.get("ai_attempted")),
        "ai_degraded_to_base": bool(values.get("ai_degraded_to_base")),
        "ai_failure_reason": ai_failure_reason,
    }


def _save_verification_source(filename: str | None, image_bytes: bytes) -> tuple[str, str]:
    OCR_VERIFICATION_DIR.mkdir(parents=True, exist_ok=True)
    safe_name = _safe_file_name(filename)
    suffix = Path(safe_name).suffix or ".png"
    stored_name = f"{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S_%f')}_{uuid.uuid4().hex[:8]}{suffix}"
    destination = OCR_VERIFICATION_DIR / stored_name
    destination.write_bytes(image_bytes)
    return safe_name, str(destination)


def _verification_query(db: Session, current_user: User):
    query = db.query(OcrVerification)
    org_id = resolve_org_id(current_user)
    factory_id = _active_factory_id(db, current_user)

    if current_user.role in {UserRole.ADMIN, UserRole.OWNER}:
        if org_id:
            return query.filter(OcrVerification.org_id == org_id)
        return query.filter(OcrVerification.user_id == current_user.id)

    if current_user.role == UserRole.MANAGER:
        if factory_id:
            return query.filter(OcrVerification.factory_id == factory_id)
        if org_id:
            return query.filter(OcrVerification.org_id == org_id)
        return query.filter(OcrVerification.user_id == current_user.id)

    return query.filter(OcrVerification.user_id == current_user.id)


def _serialize_verification(db: Session, verification: OcrVerification) -> dict:
    template_name = None
    if verification.template_id:
        template = db.query(OcrTemplate).filter(OcrTemplate.id == verification.template_id).first()
        if template:
            template_name = template.name
    actor_ids = {
        actor_id
        for actor_id in {verification.user_id, verification.approved_by, verification.rejected_by}
        if actor_id
    }
    actor_map = {}
    if actor_ids:
        actor_map = {
            actor.id: actor.name
            for actor in db.query(User).filter(User.id.in_(actor_ids)).all()
        }
    trusted_export = verification.status == "approved"
    export_source = "approved_review" if trusted_export else f"{verification.status}_review"
    return {
        "id": verification.id,
        "org_id": verification.org_id,
        "factory_id": verification.factory_id,
        "user_id": verification.user_id,
        "created_by_name": actor_map.get(verification.user_id),
        "template_id": verification.template_id,
        "template_name": template_name,
        "source_filename": verification.source_filename,
        "has_source_image": bool(verification.source_image_path),
        "source_image_url": f"/ocr/verifications/{verification.id}/source-image" if verification.source_image_path else None,
        "columns": verification.columns,
        "language": verification.language,
        "avg_confidence": float(verification.avg_confidence or 0),
        "warnings": verification.warnings or [],
        "scan_quality": verification.scan_quality or None,
        "document_hash": verification.document_hash,
        "doc_type_hint": verification.doc_type_hint,
        "routing_meta": verification.routing_meta or None,
        "raw_text": verification.raw_text,
        "headers": verification.headers or [],
        "original_rows": verification.original_rows or [],
        "reviewed_rows": verification.reviewed_rows or [],
        "raw_column_added": bool(verification.raw_column_added),
        "status": verification.status,
        "reviewer_notes": verification.reviewer_notes,
        "rejection_reason": verification.rejection_reason,
        "submitted_at": verification.submitted_at.isoformat() if verification.submitted_at else None,
        "approved_at": verification.approved_at.isoformat() if verification.approved_at else None,
        "rejected_at": verification.rejected_at.isoformat() if verification.rejected_at else None,
        "approved_by": verification.approved_by,
        "approved_by_name": actor_map.get(verification.approved_by) if verification.approved_by else None,
        "rejected_by": verification.rejected_by,
        "rejected_by_name": actor_map.get(verification.rejected_by) if verification.rejected_by else None,
        "trusted_export": trusted_export,
        "export_source": export_source,
        "export_url": f"/ocr/verifications/{verification.id}/export",
        "created_at": verification.created_at.isoformat() if verification.created_at else None,
        "updated_at": verification.updated_at.isoformat() if verification.updated_at else None,
    }


def _get_verification_or_404(db: Session, verification_id: int, current_user: User) -> OcrVerification:
    verification = _verification_query(db, current_user).filter(OcrVerification.id == verification_id).first()
    if not verification:
        raise HTTPException(status_code=404, detail="Verification record not found.")
    return verification


def _verification_export_rows(verification: OcrVerification) -> list[list[str]]:
    rows = verification.reviewed_rows or verification.original_rows or []
    return _normalize_rows(rows, field_name="verification_export_rows") or []


def _verification_export_headers(verification: OcrVerification, rows: list[list[str]]) -> list[str]:
    column_count = max((len(row) for row in rows), default=0, )
    column_count = max(column_count, verification.columns or 0, 1)
    normalized_headers = _normalize_string_list(verification.headers or [], field_name="verification_headers") or []
    if len(normalized_headers) < column_count:
        normalized_headers.extend([f"Column {index}" for index in range(len(normalized_headers) + 1, column_count + 1)])
    return normalized_headers[:column_count]


def _verification_export_source(verification: OcrVerification) -> str:
    return "approved_review" if verification.status == "approved" else f"{verification.status}_review"


def _verification_export_filename(verification: OcrVerification) -> str:
    base = _safe_file_name(verification.source_filename, default_stem=f"ocr-verification-{verification.id}")
    stem = Path(base).stem or f"ocr-verification-{verification.id}"
    return f"{stem}-{verification.status}.xlsx"


def _verification_export_response(verification: OcrVerification) -> Response:
    rows = _verification_export_rows(verification)
    if not rows:
        raise HTTPException(status_code=409, detail="Verification record has no rows to export.")
    headers = _verification_export_headers(verification, rows)
    excel_bytes = build_table_excel_bytes({"headers": headers, "rows": rows})
    trusted_export = verification.status == "approved"
    export_source = _verification_export_source(verification)
    filename = _verification_export_filename(verification)
    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "no-store",
            "Pragma": "no-cache",
            "X-Ocr-Verification-Id": str(verification.id),
            "X-Ocr-Export-Source": export_source,
            "X-Ocr-Trusted-Export": str(trusted_export).lower(),
            "X-Total-Rows": str(len(rows)),
            "X-Total-Columns": str(len(headers)),
        },
    )


def _verification_row_count(verification: OcrVerification) -> int:
    return len(_verification_export_rows(verification))


def _apply_verification_payload(
    verification: OcrVerification,
    *,
    template_id: int | None = None,
    source_filename: str | None = None,
    columns: int | None = None,
    language: str | None = None,
    avg_confidence: float | None = None,
    warnings: list[str] | None = None,
    scan_quality: dict | None = None,
    document_hash: str | None = None,
    doc_type_hint: str | None = None,
    routing_meta: dict | None = None,
    raw_text: str | None = None,
    headers: list[str] | None = None,
    original_rows: list[list[str]] | None = None,
    reviewed_rows: list[list[str]] | None = None,
    raw_column_added: bool | None = None,
    reviewer_notes: str | None = None,
) -> None:
    if template_id is not None:
        verification.template_id = template_id
    if source_filename is not None:
        verification.source_filename = _safe_file_name(source_filename)
    if columns is not None:
        verification.columns = max(1, min(columns, 12))
    if language is not None:
        verification.language = sanitize_text(language, max_length=20, preserve_newlines=False) or verification.language
    if avg_confidence is not None:
        verification.avg_confidence = max(0.0, min(float(avg_confidence), 100.0))
    if warnings is not None:
        verification.warnings = warnings
    if scan_quality is not None:
        verification.scan_quality = scan_quality
    if document_hash is not None:
        verification.document_hash = _normalize_document_hash(document_hash)
    if doc_type_hint is not None:
        verification.doc_type_hint = _normalize_doc_type_hint(doc_type_hint)
    if routing_meta is not None:
        verification.routing_meta = routing_meta
    if raw_text is not None:
        verification.raw_text = sanitize_text(raw_text, max_length=50000)
    if headers is not None:
        verification.headers = headers
    if original_rows is not None:
        verification.original_rows = original_rows
    if reviewed_rows is not None:
        verification.reviewed_rows = reviewed_rows
    if raw_column_added is not None:
        verification.raw_column_added = bool(raw_column_added)
    if reviewer_notes is not None:
        verification.reviewer_notes = sanitize_text(reviewer_notes, max_length=5000)
    verification.updated_at = datetime.now(timezone.utc)


def _require_ocr_access(current_user: User) -> None:
    require_any_role(
        current_user,
        {UserRole.OPERATOR, UserRole.SUPERVISOR, UserRole.MANAGER, UserRole.ADMIN, UserRole.OWNER},
    )


def _image_too_large_response() -> JSONResponse:
    return JSONResponse(
        status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
        content={"error": "image_too_large", "message": "Image must be under 8 MB"},
    )


def _require_templates_access(db: Session, current_user: User) -> None:
    org_id = resolve_org_id(current_user)
    plan = get_org_plan_for_usage(db, org_id=org_id, user_id=current_user.id)
    if has_plan_feature(plan, "templates"):
        return
    if org_has_ocr_access(db, org_id=org_id, fallback_user_id=current_user.id):
        return
    min_plan = min_plan_for_feature("templates")
    raise HTTPException(
        status_code=402,
        detail=f"Custom templates require {min_plan.title()} plan or an OCR pack.",
    )


def _parse_json_list(raw: str | None) -> list | None:
    if not raw:
        return None
    try:
        parsed = json.loads(raw)
        if isinstance(parsed, list):
            return parsed
    except json.JSONDecodeError:
        return None
    return None


def _template_query(db: Session, current_user: User):
    query = db.query(OcrTemplate)
    org_id = resolve_org_id(current_user)
    if current_user.role in {UserRole.ADMIN, UserRole.OWNER}:
        if org_id:
            factory_ids = db.query(Factory.factory_id).filter(Factory.org_id == org_id)
            return query.filter(OcrTemplate.factory_id.in_(factory_ids))
        return query.filter(OcrTemplate.factory_id.isnot(None))
    factory_id = _active_factory_id(db, current_user)
    if not factory_id:
        return query.filter(false())
    return query.filter(OcrTemplate.factory_id == factory_id)


def _active_factory_id(db: Session, current_user: User) -> str | None:
    factory_id = resolve_factory_id(db, current_user)
    return factory_id


def _active_factory_name(db: Session, current_user: User) -> str:
    factory_id = resolve_factory_id(db, current_user)
    if factory_id:
        row = db.query(Factory.name).filter(Factory.factory_id == factory_id).first()
        if row:
            return row[0]
    return current_user.factory_name


def _is_low_confidence(result) -> bool:
    try:
        return float(result.avg_confidence) < 55
    except Exception:
        return True


def _result_score(result) -> tuple[float, int, int]:
    rows = result.rows or []
    row_count = len(rows)
    cell_count = sum(1 for row in rows for cell in row if str(cell).strip())
    return (float(result.avg_confidence or 0.0), row_count, cell_count)


def _should_retry_with_fallback_language(result) -> bool:
    rows = result.rows or []
    row_count = len(rows)
    populated_cells = sum(1 for row in rows for cell in row if str(cell).strip())
    confidence = float(getattr(result, "avg_confidence", 0.0) or 0.0)

    if row_count == 0 or populated_cells == 0:
        return True
    if row_count <= 1 and populated_cells <= 2:
        return True
    if confidence < 20 and populated_cells <= 4:
        return True
    return False


def _job_urls(job_id: str) -> dict[str, str]:
    return {
        "status_url": f"/ocr/jobs/{job_id}",
        "download_url": f"/ocr/jobs/{job_id}/download",
    }


def _job_kind(mode: str) -> str:
    if mode not in {"ledger", "table"}:
        raise RuntimeError(f"Unsupported OCR job mode: {mode}")
    return f"ocr_{mode}_excel"


def _log_ocr_job_success(
    *,
    action: str,
    user_id: int,
    org_id: str | None,
    factory_id: str | None,
    details: str,
) -> None:
    with SessionLocal() as db:
        db.add(
            AuditLog(
                user_id=user_id,
                org_id=org_id,
                factory_id=factory_id,
                action=action,
                details=details,
                ip_address=None,
                timestamp=datetime.now(timezone.utc),
            )
        )
        db.commit()


def _run_ocr_excel_job(progress, *, job_id: str) -> dict[str, object]:
    job = get_background_job(job_id)
    if not job:
        raise RuntimeError("OCR job not found.")
    context = job.get("context") or {}
    mode = str(context.get("mode") or "")
    input_file = context.get("input_file")
    if not isinstance(input_file, dict):
        raise RuntimeError("OCR input file is missing.")
    stored_path = input_file.get("stored_path")
    if not stored_path:
        raise RuntimeError("OCR input file path is missing.")
    image_bytes = Path(str(stored_path)).read_bytes()

    progress(15, "Loading OCR image")
    metadata: dict[str, object]
    output_name: str
    if mode == "ledger":
        progress(35, "Preprocessing ledger image")
        base64_image = preprocess_image_bytes(image_bytes, profile=context.get("preprocess_profile"))
        progress(60, "Extracting ledger rows")
        rows, model_meta = ledger_extract_data(
            base64_image,
            force_mock=bool(context.get("mock")),
            system_prompt=context.get("system_prompt"),
            user_message=context.get("user_message"),
        )
        validated = ledger_validate_data(rows)
        progress(82, "Building ledger Excel workbook")
        excel_bytes = ledger_build_excel_bytes(validated)
        metadata = dict(validated.get("metadata", {}))
        metadata.update(model_meta) # Merge model tracking info
        output_name = "logbook_ledger_scan.xlsx"
        action = "OCR_LEDGER_EXCEL_ASYNC"
    elif mode == "table":
        progress(35, "Scoring table image quality")
        progress(58, "Extracting structured data with Anthropic")
        excel_bytes, metadata = _run_table_excel_pipeline(
            image_bytes,
            content_type=str(input_file.get("media_type") or context.get("content_type") or ""),
            filename=str(context.get("source_filename") or "table-ocr-input.png"),
            system_prompt=context.get("system_prompt"),
            user_message=context.get("user_message"),
        )
        progress(82, "Building table Excel workbook")
        output_name = "table_scan.xlsx"
        action = "OCR_TABLE_EXCEL_ASYNC"
    else:
        raise RuntimeError("Unsupported OCR job mode.")

    file_meta = write_job_file(
        job_id,
        filename=output_name,
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    _log_ocr_job_success(
        action=action,
        user_id=int(job["owner_id"]),
        org_id=job.get("org_id"),
        factory_id=str(context.get("factory_id")) if context.get("factory_id") is not None else None,
        details=f"{action} completed size_bytes={context.get('size_bytes', 0)}",
    )
    return {
        "file": file_meta,
        "metadata": metadata,
        "source_filename": context.get("source_filename"),
        "mode": mode,
    }


def _queue_ocr_excel_job(
    *,
    mode: str,
    owner_id: int,
    org_id: str | None,
    factory_id: str | None,
    source_filename: str,
    content_type: str | None,
    size_bytes: int,
    mock: bool = False,
    system_prompt: str | None = None,
    user_message: str | None = None,
    preprocess_profile: str | None = None,
    image_bytes: bytes | None = None,
    input_file: dict[str, object] | None = None,
) -> dict[str, object]:
    job = create_job(
        kind=_job_kind(mode),
        owner_id=owner_id,
        org_id=org_id,
        message=f"Queued {mode} OCR export",
        context={
            "route": "/ocr",
            "mode": mode,
            "factory_id": factory_id,
            "source_filename": source_filename,
            "mock": bool(mock),
            "system_prompt": system_prompt,
            "user_message": user_message,
            "preprocess_profile": preprocess_profile,
            "size_bytes": size_bytes,
        },
        retry_context={
            "mode": mode,
            "owner_id": owner_id,
            "org_id": org_id,
            "factory_id": factory_id,
            "source_filename": source_filename,
            "size_bytes": size_bytes,
            "mock": bool(mock),
            "system_prompt": system_prompt,
            "user_message": user_message,
            "preprocess_profile": preprocess_profile,
        },
    )
    job_input = input_file
    if job_input is None:
        if image_bytes is None:
            raise RuntimeError("OCR job image bytes are required.")
        job_input = write_job_file(
            job["job_id"],
            filename=source_filename or f"{mode}-ocr-input.bin",
            content=image_bytes,
            media_type=content_type or "application/octet-stream",
        )
    context = dict(job.get("context") or {})
    context["input_file"] = job_input
    retry_context = {
        "mode": mode,
        "owner_id": owner_id,
        "org_id": org_id,
        "factory_id": factory_id,
        "source_filename": source_filename,
        "size_bytes": size_bytes,
        "mock": bool(mock),
        "system_prompt": system_prompt,
        "user_message": user_message,
        "preprocess_profile": preprocess_profile,
        "input_file": job_input,
    }
    update_job(job["job_id"], context=context, retry_context=retry_context)
    start_job(job["job_id"], lambda progress: _run_ocr_excel_job(progress, job_id=job["job_id"]))
    return job


def _retry_ocr_job(payload: dict[str, object], _source_job: object) -> dict[str, object]:
    mode = str(payload.get("mode") or "")
    input_file = payload.get("input_file")
    if not isinstance(input_file, dict):
        raise RuntimeError("The original OCR job is missing its uploaded image.")
    return _queue_ocr_excel_job(
        mode=mode,
        owner_id=int(payload["owner_id"]),
        org_id=str(payload["org_id"]) if payload.get("org_id") is not None else None,
        factory_id=str(payload.get("factory_id")) if payload.get("factory_id") is not None else None,
        source_filename=str(payload.get("source_filename") or f"{mode}-ocr-input.bin"),
        content_type=str(input_file.get("media_type") or "application/octet-stream"),
        size_bytes=int(payload.get("size_bytes") or 0),
        mock=bool(payload.get("mock")),
        system_prompt=str(payload.get("system_prompt")) if payload.get("system_prompt") is not None else None,
        user_message=str(payload.get("user_message")) if payload.get("user_message") is not None else None,
        preprocess_profile=str(payload.get("preprocess_profile")) if payload.get("preprocess_profile") is not None else None,
        input_file=input_file,
    )


register_retry_handler("ocr_ledger_excel", _retry_ocr_job)
register_retry_handler("ocr_table_excel", _retry_ocr_job)


def _log_ocr_event(
    db: Session,
    *,
    action: str,
    details: str,
    request: Request | None,
    user_id: int | None,
    org_id: str | None,
    factory_id: str | None,
) -> None:
    client_host = request.client.host if request and request.client else None
    user_agent = request.headers.get("user-agent") if request else None
    db.add(
        AuditLog(
            user_id=user_id,
            org_id=org_id,
            factory_id=factory_id,
            action=action,
            details=details,
            ip_address=hash_ip_address(client_host),
            user_agent=user_agent,
            timestamp=datetime.now(timezone.utc),
        )
    )


def _run_ocr_with_fallback(
    image_bytes: bytes,
    *,
    language: str,
    columns: int,
    column_centers: list[float] | None,
    column_keywords: list[list[str]] | None,
    enable_raw_column: bool,
    allow_fallback: bool = False,
) -> tuple:
    if not allow_fallback:
        error = RuntimeError("Local OCR fallback is disabled for this request.")
        logger.error("[OCR] Error: %s", error)
        raise error

    logger.info("[OCR] Using fallback")
    fallback_used = False
    used_language = language
    primary_language = language
    fallback_language = None

    if language == "auto":
        primary_language = "eng"
        fallback_language = "eng+hin+mar"
        used_language = primary_language

    try:
        result = extract_table_from_image(
            image_bytes,
            columns=columns,
            language=primary_language,
            column_centers=column_centers,
            column_keywords=column_keywords,
            enable_raw_column=enable_raw_column,
        )
    except Exception as error:  # pylint: disable=broad-except
        logger.error("[OCR] Error: %s", error)
        logger.error(
            "OCR extraction failed on primary pass: %s: %s",
            type(error).__name__,
            error,
            exc_info=True,
        )
        raise

    needs_fallback = _should_retry_with_fallback_language(result)
    if fallback_language and needs_fallback:
        fallback_used = True
        try:
            fallback_result = extract_table_from_image(
                image_bytes,
                columns=columns,
                language=fallback_language,
                column_centers=column_centers,
                column_keywords=column_keywords,
                enable_raw_column=enable_raw_column,
            )
            if _result_score(fallback_result) >= _result_score(result):
                result = fallback_result
                used_language = fallback_language
            else:
                used_language = primary_language
        except Exception as error:  # pylint: disable=broad-except
            logger.error("[OCR] Error: %s", error)
            logger.error(
                "OCR extraction fallback failed: %s: %s",
                type(error).__name__,
                error,
                exc_info=True,
            )
            used_language = primary_language
            if hasattr(result, "warnings"):
                result.warnings.append("Fallback OCR language failed; using primary result.")

    return result, used_language, fallback_used


@router.get("/status", status_code=status.HTTP_200_OK)
def ocr_status() -> dict:
    tesseract_path = shutil.which("tesseract")
    if not tesseract_path:
        for path in (
            Path(r"C:\Program Files\Tesseract-OCR\tesseract.exe"),
            Path(r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe"),
        ):
            if path.exists():
                tesseract_path = str(path)
                break
    if not tesseract_path:
        return {"installed": False, "message": "Tesseract not found on PATH."}
    env = os.environ.copy()
    local_app = os.getenv("LOCALAPPDATA")
    if local_app:
        candidate = Path(local_app) / "DPR.ai" / "tessdata"
        if candidate.exists():
            env["TESSDATA_PREFIX"] = str(candidate)
    try:
        version = subprocess.check_output([tesseract_path, "--version"], text=True, env=env).splitlines()[0]
        langs_raw = subprocess.check_output([tesseract_path, "--list-langs"], text=True, env=env)
        langs = [line.strip() for line in langs_raw.splitlines() if line and ":" not in line]
    except Exception:  # pylint: disable=broad-except
        version = "unknown"
        langs = []
    return {
        "installed": True,
        "path": tesseract_path,
        "version": version,
        "tessdata_prefix": env.get("TESSDATA_PREFIX"),
        "languages": langs,
    }


@router.get("/templates", status_code=status.HTTP_200_OK)
def list_templates(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> list[dict]:
    _require_ocr_access(current_user)
    _require_templates_access(db, current_user)
    templates = (
        _template_query(db, current_user)
        .filter(OcrTemplate.is_active.is_(True))
        .order_by(OcrTemplate.created_at.desc())
        .all()
    )
    return [
        {
            "id": template.id,
            "name": template.name,
            "columns": template.columns,
            "header_mode": template.header_mode,
            "language": template.language,
            "column_names": template.column_names or [],
            "column_keywords": template.column_keywords or [],
            "raw_column_label": template.raw_column_label or "Raw",
            "enable_raw_column": template.enable_raw_column,
            "created_at": template.created_at.isoformat(),
        }
        for template in templates
    ]


@router.post("/templates", status_code=status.HTTP_201_CREATED)
async def create_template(
    name: str = Form(...),
    columns: int = Form(default=3),
    header_mode: str = Form(default="first"),
    language: str = Form(default="eng"),
    column_names: str | None = Form(default=None),
    column_keywords: str | None = Form(default=None),
    raw_column_label: str | None = Form(default="Raw"),
    enable_raw_column: bool = Form(default=True),
    samples: list[UploadFile] = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    _require_ocr_access(current_user)
    _require_templates_access(db, current_user)
    if not samples:
        raise HTTPException(status_code=400, detail="Sample images are required.")
    if columns < 1 or columns > 8:
        raise HTTPException(status_code=400, detail="Columns must be between 1 and 8.")

    sample_bytes: list[bytes] = []
    for file in samples:
        sample_bytes.append(await _read_validated_image_upload(file))

    parsed_names = _parse_json_list(column_names)
    parsed_keywords = _parse_json_list(column_keywords)

    try:
        centers, avg_conf, warnings = detect_column_centers(sample_bytes, columns=columns, language=language)
    except RuntimeError as error:
        logger.exception("Template analysis failed.")
        raise HTTPException(status_code=400, detail=str(error)) from error
    except Exception as error:  # pylint: disable=broad-except
        logger.exception("Template analysis failed.")
        raise HTTPException(status_code=500, detail="Template analysis failed. Please verify OCR setup.") from error
    factory_id = _active_factory_id(db, current_user)
    if not factory_id:
        raise HTTPException(status_code=400, detail="Factory must be selected before creating templates.")
    template = OcrTemplate(
        factory_id=factory_id,
        factory_name=_active_factory_name(db, current_user),
        name=name.strip(),
        columns=columns,
        header_mode=header_mode,
        language=language,
        column_names=parsed_names,
        column_keywords=parsed_keywords,
        column_centers=centers,
        raw_column_label=raw_column_label or "Raw",
        enable_raw_column=enable_raw_column,
        created_by=current_user.id,
        is_active=True,
        created_at=datetime.now(timezone.utc),
        updated_at=datetime.now(timezone.utc),
    )
    db.add(template)
    db.commit()
    db.refresh(template)
    return {
        "id": template.id,
        "avg_confidence": avg_conf,
        "warnings": warnings,
        "template": {
            "id": template.id,
            "name": template.name,
            "columns": template.columns,
            "header_mode": template.header_mode,
            "language": template.language,
            "column_names": template.column_names or [],
            "column_keywords": template.column_keywords or [],
            "raw_column_label": template.raw_column_label or "Raw",
            "enable_raw_column": template.enable_raw_column,
        },
    }


@router.delete("/templates/{template_id}", status_code=status.HTTP_200_OK)
def deactivate_template(
    template_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    _require_ocr_access(current_user)
    _require_templates_access(db, current_user)
    template = (
        _template_query(db, current_user)
        .filter(OcrTemplate.id == template_id, OcrTemplate.is_active.is_(True))
        .first()
    )
    if not template:
        raise HTTPException(status_code=404, detail="Template not found.")
    template.is_active = False
    template.updated_at = datetime.now(timezone.utc)
    db.commit()
    return {"message": "Template archived."}


@router.get("/verifications", status_code=status.HTTP_200_OK)
def list_verifications(
    verification_status: str | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> list[dict]:
    _require_ocr_access(current_user)
    query = _verification_query(db, current_user)
    if verification_status:
        normalized = (sanitize_text(verification_status, max_length=20, preserve_newlines=False) or "").lower()
        if normalized not in _ALLOWED_VERIFICATION_STATUSES:
            raise HTTPException(status_code=400, detail="Invalid verification status filter.")
        query = query.filter(OcrVerification.status == normalized)
    records = query.order_by(OcrVerification.updated_at.desc(), OcrVerification.id.desc()).limit(100).all()
    return [_serialize_verification(db, record) for record in records]


@router.get("/verifications/summary", status_code=status.HTTP_200_OK)
def get_verification_summary(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    _require_ocr_access(current_user)
    records = _verification_query(db, current_user).all()

    trusted_documents = 0
    trusted_rows = 0
    pending_documents = 0
    pending_rows = 0
    rejected_documents = 0
    rejected_rows = 0
    draft_documents = 0
    draft_rows = 0
    trusted_confidence_total = 0.0
    trusted_confidence_count = 0
    last_trusted_at: datetime | None = None

    for record in records:
        row_count = _verification_row_count(record)
        if record.status == "approved":
            trusted_documents += 1
            trusted_rows += row_count
            if record.avg_confidence is not None:
                trusted_confidence_total += float(record.avg_confidence)
                trusted_confidence_count += 1
            if record.approved_at and (last_trusted_at is None or record.approved_at > last_trusted_at):
                last_trusted_at = record.approved_at
        elif record.status == "pending":
            pending_documents += 1
            pending_rows += row_count
        elif record.status == "rejected":
            rejected_documents += 1
            rejected_rows += row_count
        else:
            draft_documents += 1
            draft_rows += row_count

    decision_denominator = trusted_documents + rejected_documents
    approval_rate = (
        round((trusted_documents / decision_denominator) * 100, 1)
        if decision_denominator
        else None
    )

    return {
        "total_documents": len(records),
        "trusted_documents": trusted_documents,
        "trusted_rows": trusted_rows,
        "pending_documents": pending_documents,
        "pending_rows": pending_rows,
        "rejected_documents": rejected_documents,
        "rejected_rows": rejected_rows,
        "draft_documents": draft_documents,
        "draft_rows": draft_rows,
        "untrusted_documents": draft_documents + pending_documents + rejected_documents,
        "untrusted_rows": draft_rows + pending_rows + rejected_rows,
        "export_ready_documents": trusted_documents,
        "avg_trusted_confidence": round(trusted_confidence_total / trusted_confidence_count, 1)
        if trusted_confidence_count
        else None,
        "approval_rate": approval_rate,
        "last_trusted_at": last_trusted_at.isoformat() if last_trusted_at else None,
        "trust_note": "Only approved OCR documents count as trusted downstream data.",
    }


@router.post("/verifications", status_code=status.HTTP_201_CREATED)
async def create_verification(
    file: UploadFile | None = File(default=None),
    template_id: int | None = Form(default=None),
    source_filename: str | None = Form(default=None),
    columns: int = Form(default=3),
    language: str = Form(default="eng"),
    avg_confidence: float | None = Form(default=None),
    warnings: str | None = Form(default=None),
    scan_quality: str | None = Form(default=None),
    document_hash: str | None = Form(default=None),
    doc_type_hint: str | None = Form(default=None),
    routing_meta: str | None = Form(default=None),
    raw_text: str | None = Form(default=None),
    headers: str | None = Form(default=None),
    original_rows: str | None = Form(default=None),
    reviewed_rows: str | None = Form(default=None),
    raw_column_added: bool = Form(default=False),
    reviewer_notes: str | None = Form(default=None),
    request: Request = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    _require_ocr_access(current_user)

    template = None
    if template_id is not None:
        template = (
            _template_query(db, current_user)
            .filter(OcrTemplate.id == template_id, OcrTemplate.is_active.is_(True))
            .first()
        )
        if not template:
            raise HTTPException(status_code=404, detail="Template not found.")

    parsed_warnings = _normalize_string_list(_parse_json_value(warnings, field_name="warnings"), field_name="warnings") or []
    parsed_scan_quality = _normalize_scan_quality(
        _parse_json_value(scan_quality, field_name="scan_quality"),
        field_name="scan_quality",
    )
    parsed_routing_meta = _normalize_routing_meta(
        _parse_json_value(routing_meta, field_name="routing_meta"),
        field_name="routing_meta",
    )
    parsed_headers = _normalize_string_list(_parse_json_value(headers, field_name="headers"), field_name="headers") or []
    parsed_original_rows = _normalize_rows(_parse_json_value(original_rows, field_name="original_rows"), field_name="original_rows") or []
    parsed_reviewed_rows = _normalize_rows(_parse_json_value(reviewed_rows, field_name="reviewed_rows"), field_name="reviewed_rows") or parsed_original_rows

    if file is not None and file.filename:
        image_bytes = await _read_validated_image_upload(file)
        source_name, image_path = _save_verification_source(file.filename or source_filename, image_bytes)
    else:
        source_name = _safe_file_name(source_filename) if source_filename else None
        image_path = None

    if not parsed_reviewed_rows and not parsed_original_rows:
        raise HTTPException(status_code=400, detail="Provide OCR rows before saving a verification draft.")

    verification = OcrVerification(
        org_id=resolve_org_id(current_user),
        factory_id=_active_factory_id(db, current_user),
        user_id=current_user.id,
        template_id=template.id if template else template_id,
        source_filename=source_name,
        source_image_path=image_path,
        columns=max(1, min(columns, 12)),
        language=sanitize_text(language, max_length=20, preserve_newlines=False) or (template.language if template else "eng"),
        avg_confidence=max(0.0, min(float(avg_confidence), 100.0)) if avg_confidence is not None else None,
        warnings=parsed_warnings,
        scan_quality=parsed_scan_quality,
        document_hash=_normalize_document_hash(document_hash),
        doc_type_hint=_normalize_doc_type_hint(doc_type_hint),
        routing_meta=parsed_routing_meta,
        raw_text=sanitize_text(raw_text, max_length=50000),
        headers=parsed_headers,
        original_rows=parsed_original_rows,
        reviewed_rows=parsed_reviewed_rows,
        raw_column_added=bool(raw_column_added),
        status="draft",
        reviewer_notes=sanitize_text(reviewer_notes, max_length=5000),
    )
    db.add(verification)
    db.commit()
    db.refresh(verification)
    if request is not None:
        _log_ocr_event(
            db,
            action="OCR_VERIFICATION_CREATED",
            details=(
                f"Verification created id={verification.id} status=draft "
                f"confidence={verification.avg_confidence or 0:.1f} "
                f"tier={(verification.routing_meta or {}).get('model_tier', 'fast')} "
                f"band={(verification.scan_quality or {}).get('confidence_band', 'unknown')} "
                f"outcome={(verification.scan_quality or {}).get('outcome', 'success')}"
            ),
            request=request,
            user_id=current_user.id,
            org_id=verification.org_id,
            factory_id=verification.factory_id,
        )
        db.commit()
    return _serialize_verification(db, verification)


@router.get("/verifications/{verification_id}", status_code=status.HTTP_200_OK)
def get_verification(
    verification_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    _require_ocr_access(current_user)
    verification = _get_verification_or_404(db, verification_id, current_user)
    return _serialize_verification(db, verification)


@router.get("/verifications/{verification_id}/source-image")
def get_verification_source_image(
    verification_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> FileResponse:
    _require_ocr_access(current_user)
    verification = _get_verification_or_404(db, verification_id, current_user)
    if not verification.source_image_path:
        raise HTTPException(status_code=404, detail="Verification source image not found.")

    image_path = Path(verification.source_image_path)
    try:
        resolved_path = image_path.resolve(strict=True)
    except FileNotFoundError as error:
        raise HTTPException(status_code=404, detail="Verification source image not found.") from error

    verification_root = OCR_VERIFICATION_DIR.resolve()
    if resolved_path != verification_root and verification_root not in resolved_path.parents:
        raise HTTPException(status_code=404, detail="Verification source image not found.")

    media_type = mimetypes.guess_type(resolved_path.name)[0] or "application/octet-stream"
    return FileResponse(
        resolved_path,
        media_type=media_type,
        filename=resolved_path.name,
        headers={"Cache-Control": "private, max-age=3600"},
    )


@router.get("/verifications/{verification_id}/export", status_code=status.HTTP_200_OK)
def export_verification_excel(
    verification_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> Response:
    _require_ocr_access(current_user)
    verification = _get_verification_or_404(db, verification_id, current_user)
    rows = _verification_export_rows(verification)
    headers = _verification_export_headers(verification, rows)
    trusted_export = verification.status == "approved"
    _log_ocr_event(
        db,
        action="OCR_VERIFICATION_EXCEL_EXPORT",
        details=(
            f"Verification Excel export id={verification.id} "
            f"status={verification.status} trusted={str(trusted_export).lower()} "
            f"rows={len(rows)} columns={len(headers)}"
        ),
        request=request,
        user_id=current_user.id,
        org_id=verification.org_id,
        factory_id=verification.factory_id,
    )
    db.commit()
    return _verification_export_response(verification)


@router.post("/verifications/{verification_id}/share-link", status_code=status.HTTP_200_OK)
def create_verification_share_link(
    verification_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    _require_ocr_access(current_user)
    verification = _get_verification_or_404(db, verification_id, current_user)
    token = _build_ocr_share_token(verification)
    expires_at = datetime.now(timezone.utc) + timedelta(seconds=_OCR_SHARE_MAX_AGE_SECONDS)
    return {
      "url": f"/api/ocr/shared/{token}",
      "expires_at": expires_at.isoformat(),
    }


@router.get("/shared/{token}", status_code=status.HTTP_200_OK)
def export_shared_verification_excel(
    token: str,
    db: Session = Depends(get_db),
) -> Response:
    payload = _read_ocr_share_token(token)
    verification_id = int(payload.get("verification_id") or 0)
    verification = db.query(OcrVerification).filter(OcrVerification.id == verification_id).first()
    if not verification:
        raise HTTPException(status_code=404, detail="Verification record not found.")
    if payload.get("org_id") != verification.org_id or payload.get("factory_id") != verification.factory_id:
        raise HTTPException(status_code=404, detail="Share link invalid.")
    return _verification_export_response(verification)


@router.put("/verifications/{verification_id}", status_code=status.HTTP_200_OK)
def update_verification(
    verification_id: int,
    payload: OcrVerificationUpdatePayload,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    _require_ocr_access(current_user)
    verification = _get_verification_or_404(db, verification_id, current_user)

    template_id = payload.template_id
    if template_id is not None:
        template = (
            _template_query(db, current_user)
            .filter(OcrTemplate.id == template_id, OcrTemplate.is_active.is_(True))
            .first()
        )
        if not template:
            raise HTTPException(status_code=404, detail="Template not found.")

    headers = _normalize_string_list(payload.headers, field_name="headers")
    warnings = _normalize_string_list(payload.warnings, field_name="warnings")
    scan_quality = _normalize_scan_quality(payload.scan_quality, field_name="scan_quality")
    routing_meta = _normalize_routing_meta(payload.routing_meta, field_name="routing_meta")
    original_rows = _normalize_rows(payload.original_rows, field_name="original_rows")
    reviewed_rows = _normalize_rows(payload.reviewed_rows, field_name="reviewed_rows")

    _apply_verification_payload(
        verification,
        template_id=template_id,
        source_filename=payload.source_filename,
        columns=payload.columns,
        language=payload.language,
        avg_confidence=payload.avg_confidence,
        warnings=warnings,
        scan_quality=scan_quality,
        document_hash=payload.document_hash,
        doc_type_hint=payload.doc_type_hint,
        routing_meta=routing_meta,
        raw_text=payload.raw_text,
        headers=headers,
        original_rows=original_rows,
        reviewed_rows=reviewed_rows,
        raw_column_added=payload.raw_column_added,
        reviewer_notes=payload.reviewer_notes,
    )

    if current_user.id == verification.user_id and verification.status in {"pending", "rejected"}:
        verification.status = "draft"
        verification.submitted_at = None
        verification.approved_at = None
        verification.approved_by = None
        verification.rejected_at = None
        verification.rejected_by = None
        verification.rejection_reason = None

    db.commit()
    _log_ocr_event(
        db,
        action="OCR_VERIFICATION_UPDATED",
        details=(
            f"Verification updated id={verification.id} status={verification.status} "
            f"confidence={verification.avg_confidence or 0:.1f} "
            f"tier={(verification.routing_meta or {}).get('model_tier', 'fast')} "
            f"band={(verification.scan_quality or {}).get('confidence_band', 'unknown')} "
            f"corrections={(verification.scan_quality or {}).get('correction_count', 0)}"
        ),
        request=request,
        user_id=current_user.id,
        org_id=verification.org_id,
        factory_id=verification.factory_id,
    )
    db.commit()
    db.refresh(verification)
    return _serialize_verification(db, verification)


@router.post("/verifications/{verification_id}/submit", status_code=status.HTTP_200_OK)
def submit_verification(
    verification_id: int,
    payload: OcrVerificationSubmitPayload,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    _require_ocr_access(current_user)
    verification = _get_verification_or_404(db, verification_id, current_user)
    if not (verification.reviewed_rows or verification.original_rows):
        raise HTTPException(status_code=400, detail="Verification draft has no OCR rows to submit.")
    verification.status = "pending"
    verification.submitted_at = datetime.now(timezone.utc)
    verification.approved_at = None
    verification.approved_by = None
    verification.rejected_at = None
    verification.rejected_by = None
    verification.rejection_reason = None
    if payload.reviewer_notes is not None:
        verification.reviewer_notes = sanitize_text(payload.reviewer_notes, max_length=5000)
    verification.updated_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(verification)
    return _serialize_verification(db, verification)


@router.post("/verifications/{verification_id}/approve", status_code=status.HTTP_200_OK)
def approve_verification(
    verification_id: int,
    payload: OcrVerificationDecisionPayload,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    _require_ocr_access(current_user)
    if not is_manager_or_admin(current_user):
        raise HTTPException(status_code=403, detail="Only managers and admins can approve verification records.")
    verification = _get_verification_or_404(db, verification_id, current_user)
    verification.status = "approved"
    verification.approved_at = datetime.now(timezone.utc)
    verification.approved_by = current_user.id
    verification.rejected_at = None
    verification.rejected_by = None
    verification.rejection_reason = None
    if payload.reviewer_notes is not None:
        verification.reviewer_notes = sanitize_text(payload.reviewer_notes, max_length=5000)
    verification.updated_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(verification)
    return _serialize_verification(db, verification)


@router.post("/verifications/{verification_id}/reject", status_code=status.HTTP_200_OK)
def reject_verification(
    verification_id: int,
    payload: OcrVerificationDecisionPayload,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    _require_ocr_access(current_user)
    if not is_manager_or_admin(current_user):
        raise HTTPException(status_code=403, detail="Only managers and admins can reject verification records.")
    verification = _get_verification_or_404(db, verification_id, current_user)
    rejection_reason = sanitize_text(payload.rejection_reason, max_length=5000)
    if not rejection_reason:
        raise HTTPException(status_code=400, detail="Rejection reason is required.")
    verification.status = "rejected"
    verification.rejected_at = datetime.now(timezone.utc)
    verification.rejected_by = current_user.id
    verification.rejection_reason = rejection_reason
    verification.approved_at = None
    verification.approved_by = None
    if payload.reviewer_notes is not None:
        verification.reviewer_notes = sanitize_text(payload.reviewer_notes, max_length=5000)
    verification.updated_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(verification)
    return _serialize_verification(db, verification)


@router.post("/logbook", status_code=status.HTTP_200_OK)
async def ocr_logbook(
    file: UploadFile = File(...),
    columns: int = Form(default=3),
    language: str = Form(default="eng"),
    template_id: int | None = Form(default=None),
    doc_type_hint: str | None = Form(default=None),
    force_model: str | None = Form(default=None),
    document_hash: str | None = Form(default=None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    _require_ocr_access(current_user)
    image_bytes = await _read_validated_image_upload(file)
    requested_doc_type = _normalize_doc_type_hint(doc_type_hint) or "table"

    template = None
    if template_id is not None:
        template = (
            _template_query(db, current_user)
            .filter(OcrTemplate.id == template_id, OcrTemplate.is_active.is_(True))
            .first()
        )
        if not template:
            raise HTTPException(status_code=404, detail="Template not found.")

    reusable = find_reusable_verification(
        db,
        org_id=resolve_org_id(current_user),
        document_hash=_normalize_document_hash(document_hash),
        template_id=template.id if template else template_id,
        doc_type_hint=requested_doc_type,
    )
    if reusable is not None:
        reused_payload = {
            **serialize_reused_ocr_result(reusable, template=template),
            "template": {
                "id": template.id,
                "name": template.name,
                "columns": template.columns,
                "header_mode": template.header_mode,
                "language": template.language,
                "column_names": template.column_names or [],
                "column_keywords": template.column_keywords or [],
                "raw_column_label": template.raw_column_label or "Raw",
                "enable_raw_column": template.enable_raw_column,
            }
            if template
            else None,
        }
        return reused_payload

    requested_language = template.language if template else language

    if requested_doc_type in {"table", "sheet", "spreadsheet", "logbook", "ledger", "register"}:
        fallback_used = False
        try:
            structured = _run_table_preview_pipeline(
                image_bytes,
                content_type=file.content_type,
                filename=file.filename,
                template=template,
                doc_type_hint=requested_doc_type,
                force_model=force_model,
                language=requested_language,
            )
        except TableExcelRouteError as error:
            logger.error("Structured table preview failed: %s", error.payload, exc_info=True)
            raise HTTPException(status_code=error.status_code, detail=error.payload.get("error")) from error
        except Exception as error:  # pylint: disable=broad-except
            logger.error("Structured table preview failed unexpectedly: %s: %s", type(error).__name__, error, exc_info=True)
            raise HTTPException(status_code=500, detail=f"Structured OCR failed: {error}") from error
    else:
        try:
            result, used_language, fallback_used = _run_ocr_with_fallback(
                image_bytes,
                columns=template.columns if template else columns,
                language=requested_language,
                column_centers=template.column_centers if template else None,
                column_keywords=template.column_keywords if template else None,
                enable_raw_column=bool(template.enable_raw_column) if template else False,
                allow_fallback=True,
            )
        except RuntimeError as error:
            logger.error("OCR extraction failed: %s: %s", type(error).__name__, error, exc_info=True)
            raise HTTPException(status_code=400, detail=str(error)) from error
        except Exception as error:  # pylint: disable=broad-except
            logger.error("OCR extraction failed: %s: %s", type(error).__name__, error, exc_info=True)
            raise HTTPException(status_code=500, detail=f"OCR failed: {error}") from error

        try:
            structured = build_structured_ocr_result(
                image_bytes,
                base_result=result,
                used_language=used_language,
                fallback_used=fallback_used,
                template=template,
                doc_type_hint=requested_doc_type,
                force_model=force_model,
            )
        except RuntimeError as error:
            logger.error("Structured OCR build failed: %s: %s", type(error).__name__, error, exc_info=True)
            raise HTTPException(status_code=502, detail=str(error)) from error
        except Exception as error:  # pylint: disable=broad-except
            logger.error("Structured OCR build failed unexpectedly: %s: %s", type(error).__name__, error, exc_info=True)
            raise HTTPException(status_code=500, detail=f"Structured OCR failed: {error}") from error

    scan_quality_payload = None
    try:
        image_quality = analyze_image_quality(image_bytes)
        warning_band = "low" if image_quality.warnings else "high"
        scan_quality_payload = {
            "confidence_band": warning_band,
            "quality_signals": image_quality.warnings,
            "auto_processing": ["deskew", "compression"],
            "fallback_used": fallback_used,
            "correction_count": 0,
            "page_count": 1,
            "adjustment_count": 0,
            "retake_count": 0,
            "manual_review_recommended": bool(image_quality.warnings),
            "outcome": "partial" if image_quality.warnings else "success",
            "next_action": "upload_better_image" if image_quality.warnings else None,
            "notes": "Image quality may affect accuracy." if image_quality.warnings else None,
            "cell_boxes": structured.get("cell_boxes"),
        }
    except Exception as error:  # pylint: disable=broad-except
        logger.warning("OCR scan quality analysis failed: %s", error, exc_info=True)
        scan_quality_payload = None

    return {
        **structured,
        "scan_quality": scan_quality_payload,
        "template": {
            "id": template.id,
            "name": template.name,
            "columns": template.columns,
            "header_mode": template.header_mode,
            "language": template.language,
            "column_names": template.column_names or [],
            "column_keywords": template.column_keywords or [],
            "raw_column_label": template.raw_column_label or "Raw",
            "enable_raw_column": template.enable_raw_column,
        }
        if template
        else None,
    }


@router.post("/warp", status_code=status.HTTP_200_OK)
async def warp_document(
    file: UploadFile = File(...),
    corners: str | None = Form(default=None),
    current_user: User = Depends(get_current_user),
) -> Response:
    _require_ocr_access(current_user)
    image_bytes = await _read_validated_image_upload(file)

    parsed = _parse_json_value(corners, field_name="corners")
    used_corners = None
    if parsed is not None:
        if not isinstance(parsed, list) or len(parsed) != 4:
            raise HTTPException(status_code=400, detail="corners must be a list of 4 points.")
        used_corners = []
        for point in parsed:
            if not isinstance(point, (list, tuple)) or len(point) != 2:
                raise HTTPException(status_code=400, detail="Each corner must be [x,y].")
            used_corners.append([float(point[0]), float(point[1])])

    try:
        warped_bytes, applied = warp_perspective(image_bytes, corners=used_corners)
    except RuntimeError as error:
        raise HTTPException(status_code=400, detail=str(error)) from error
    except Exception as error:  # pylint: disable=broad-except
        logger.exception("Warp failed.")
        raise HTTPException(status_code=500, detail="Could not fix perspective.") from error

    headers = {"Cache-Control": "no-store", "Pragma": "no-cache"}
    if applied is not None:
        headers["X-Warp-Corners"] = json.dumps(applied)
    return Response(content=warped_bytes, media_type="image/png", headers=headers)


@router.post("/logbook-excel", status_code=status.HTTP_200_OK)
async def ocr_logbook_excel(
    file: UploadFile = File(...),
    mock: bool = Form(default=False),
    system_prompt: str | None = Form(default=None),
    user_message: str | None = Form(default=None),
    preprocess_profile: str | None = Form(default=None),
    request: Request = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Response:
    _require_ocr_access(current_user)
    if mock:
        _reject_mock_ocr()
        image_bytes = await _read_image_upload_for_mock(file)
    else:
        image_bytes = await _read_validated_image_upload(file)
    org_id = resolve_org_id(current_user)
    plan = get_org_plan_for_usage(db, org_id=org_id, user_id=current_user.id)
    check_rate_limit(current_user.id, plan=plan)
    if org_id:
        check_and_record_org_usage(db, org_id=org_id, image_bytes=len(image_bytes), plan=plan)
    else:
        check_and_record_usage(db, user_id=current_user.id, image_bytes=len(image_bytes), plan=plan)

    try:
        if not preprocess_profile:
            preprocess_profile = os.getenv("LEDGER_SCAN_PREPROCESS_PROFILE")
        base64_image = preprocess_image_bytes(image_bytes, profile=preprocess_profile)
        rows = ledger_extract_data(
            base64_image,
            force_mock=mock,
            system_prompt=system_prompt,
            user_message=user_message,
        )
        validated = ledger_validate_data(rows)
        excel_bytes = ledger_build_excel_bytes(validated)
    except ValueError as error:
        logger.exception("LedgerScan JSON parsing failed.")
        raise HTTPException(status_code=400, detail=str(error)) from error
    except AuthenticationError as error:
        logger.exception("LedgerScan authentication failed.")
        raise HTTPException(status_code=401, detail="Anthropic authentication failed. Check ANTHROPIC_API_KEY.") from error
    except BadRequestError as error:
        logger.exception("LedgerScan request rejected by Anthropic.")
        message = str(error)
        if "credit balance is too low" in message.lower():
            raise HTTPException(status_code=429, detail="Anthropic API credits are too low.") from error
        raise HTTPException(status_code=400, detail=message) from error
    except RuntimeError as error:
        logger.exception("LedgerScan configuration error.")
        raise HTTPException(status_code=500, detail=str(error)) from error
    except Exception as error:  # pylint: disable=broad-except
        logger.exception("LedgerScan OCR failed.")
        raise HTTPException(status_code=500, detail="LedgerScan OCR failed unexpectedly.") from error

    metadata = validated.get("metadata", {})
    headers = {
        "Content-Disposition": "attachment; filename=logbook_ledger_scan.xlsx",
        "Cache-Control": "no-store",
        "Pragma": "no-cache",
        "X-Total-Rows": str(metadata.get("total_rows", 0)),
        "X-Total-Dr": str(metadata.get("total_dr", 0)),
        "X-Total-Cr": str(metadata.get("total_cr", 0)),
        "X-Balanced": str(bool(metadata.get("balanced"))).lower(),
        "X-Difference": str(metadata.get("difference", 0)),
        "X-Low-Confidence-Rows": json.dumps(metadata.get("low_confidence_rows", [])),
    }
    if request is not None:
        safe_name = sanitize_text(file.filename, max_length=200, preserve_newlines=False) or "unknown"
        org_id = resolve_org_id(current_user)
        factory_id = resolve_factory_id(db, current_user)
        _log_ocr_event(
            db,
            action="OCR_LEDGER_EXCEL",
            details=f"Ledger Excel generated filename={safe_name} size_bytes={len(image_bytes)}",
            request=request,
            user_id=current_user.id,
            org_id=org_id,
            factory_id=factory_id,
        )
        db.commit()
    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.post("/logbook-excel-async", status_code=status.HTTP_202_ACCEPTED)
async def ocr_logbook_excel_async(
    file: UploadFile = File(...),
    mock: bool = Form(default=False),
    system_prompt: str | None = Form(default=None),
    user_message: str | None = Form(default=None),
    preprocess_profile: str | None = Form(default=None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    _require_ocr_access(current_user)
    if mock:
        _reject_mock_ocr()
        image_bytes = await _read_image_upload_for_mock(file)
    else:
        image_bytes = await _read_validated_image_upload(file)
    if not preprocess_profile:
        preprocess_profile = os.getenv("LEDGER_SCAN_PREPROCESS_PROFILE")
    org_id = resolve_org_id(current_user)
    plan = get_org_plan_for_usage(db, org_id=org_id, user_id=current_user.id)
    check_rate_limit(current_user.id, plan=plan)
    if org_id:
        check_and_record_org_usage(db, org_id=org_id, image_bytes=len(image_bytes), plan=plan)
    else:
        check_and_record_usage(db, user_id=current_user.id, image_bytes=len(image_bytes), plan=plan)
    job = _queue_ocr_excel_job(
        mode="ledger",
        owner_id=current_user.id,
        org_id=org_id,
        factory_id=resolve_factory_id(db, current_user),
        source_filename=_safe_file_name(file.filename, "ledger-ocr-input.png"),
        content_type=file.content_type,
        size_bytes=len(image_bytes),
        mock=mock,
        system_prompt=system_prompt,
        user_message=user_message,
        preprocess_profile=preprocess_profile,
        image_bytes=image_bytes,
    )
    payload = dict(job)
    payload.update(_job_urls(str(job["job_id"])))
    return payload


@router.post("/table-excel", status_code=status.HTTP_200_OK)
async def ocr_table_excel(
    file: UploadFile | None = File(default=None),
    image: UploadFile | None = File(default=None),
    system_prompt: str | None = Form(default=None),
    user_message: str | None = Form(default=None),
    preprocess_profile: str | None = Form(default=None),
    request: Request = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Response:
    _require_ocr_access(current_user)
    del preprocess_profile
    try:
        upload, image_bytes = await _read_table_excel_upload(file, image)
    except TableExcelRouteError as error:
        return JSONResponse(status_code=error.status_code, content=error.payload)
    org_id = resolve_org_id(current_user)
    plan = get_org_plan_for_usage(db, org_id=org_id, user_id=current_user.id)
    check_rate_limit(current_user.id, plan=plan)
    if org_id:
        check_and_record_org_usage(db, org_id=org_id, image_bytes=len(image_bytes), plan=plan)
    else:
        check_and_record_usage(db, user_id=current_user.id, image_bytes=len(image_bytes), plan=plan)

    try:
        excel_bytes, metadata = _run_table_excel_pipeline(
            image_bytes,
            content_type=upload.content_type,
            filename=upload.filename,
            system_prompt=system_prompt,
            user_message=user_message,
        )
    except TableExcelRouteError as error:
        logger.warning("Table Excel OCR failed status=%s error=%s", error.status_code, error.payload.get("error"))
        return JSONResponse(status_code=error.status_code, content=error.payload)
    except Exception as error:  # pylint: disable=broad-except
        logger.exception("[OCR] Error: Table Excel OCR failed unexpectedly")
        return JSONResponse(status_code=500, content={"error": f"Table Excel OCR failed: {error}"})

    headers = {
        "Content-Disposition": "attachment; filename=output.xlsx",
        "X-Total-Rows": str(metadata.get("total_rows", 0)),
        "X-Total-Columns": str(metadata.get("total_columns", 0)),
        "X-Image-Quality-Score": str(metadata.get("image_quality_score", "")),
        "X-Model-Used": str(metadata.get("model_used", "")),
        "Cache-Control": "no-store",
    }
    if request is not None:
        safe_name = sanitize_text(upload.filename, max_length=200, preserve_newlines=False) or "unknown"
        org_id = resolve_org_id(current_user)
        factory_id = resolve_factory_id(db, current_user)
        _log_ocr_event(
            db,
            action="OCR_TABLE_EXCEL",
            details=f"Table Excel generated filename={safe_name} size_bytes={len(image_bytes)}",
            request=request,
            user_id=current_user.id,
            org_id=org_id,
            factory_id=factory_id,
        )
        db.commit()

    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.post("/table-excel-async", status_code=status.HTTP_202_ACCEPTED)
async def ocr_table_excel_async(
    file: UploadFile | None = File(default=None),
    image: UploadFile | None = File(default=None),
    system_prompt: str | None = Form(default=None),
    user_message: str | None = Form(default=None),
    preprocess_profile: str | None = Form(default=None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    _require_ocr_access(current_user)
    del preprocess_profile
    try:
        upload, image_bytes = await _read_table_excel_upload(file, image)
        inspection = _inspect_table_excel_image(
            image_bytes,
            content_type=upload.content_type,
            filename=upload.filename,
        )
    except TableExcelRouteError as error:
        return JSONResponse(status_code=error.status_code, content=error.payload)
    if int(inspection["image_quality_score"]) < 30:
        return JSONResponse(
            status_code=400,
            content={
                "error": "Image too vague or low quality to process",
                "imageQualityScore": int(inspection["image_quality_score"]),
            },
        )
    org_id = resolve_org_id(current_user)
    plan = get_org_plan_for_usage(db, org_id=org_id, user_id=current_user.id)
    check_rate_limit(current_user.id, plan=plan)
    if org_id:
        check_and_record_org_usage(db, org_id=org_id, image_bytes=len(image_bytes), plan=plan)
    else:
        check_and_record_usage(db, user_id=current_user.id, image_bytes=len(image_bytes), plan=plan)
    job = _queue_ocr_excel_job(
        mode="table",
        owner_id=current_user.id,
        org_id=org_id,
        factory_id=resolve_factory_id(db, current_user),
        source_filename=_safe_file_name(upload.filename, "table-ocr-input.png"),
        content_type=upload.content_type,
        size_bytes=len(image_bytes),
        system_prompt=system_prompt,
        user_message=user_message,
        image_bytes=image_bytes,
    )
    payload = dict(job)
    payload.update(_job_urls(str(job["job_id"])))
    return payload


@router.get("/jobs/{job_id}", status_code=status.HTTP_200_OK)
def get_ocr_job(job_id: str, current_user: User = Depends(get_current_user)) -> dict:
    _require_ocr_access(current_user)
    payload = get_background_job(job_id, owner_id=current_user.id)
    if payload is None or not str(payload.get("kind", "")).startswith("ocr_"):
        raise HTTPException(status_code=404, detail="Job not found.")
    payload.update(_job_urls(job_id))
    return payload


@router.get("/jobs/{job_id}/download", status_code=status.HTTP_200_OK)
def download_ocr_job(job_id: str, current_user: User = Depends(get_current_user)) -> Response:
    _require_ocr_access(current_user)
    job = get_background_job(job_id, owner_id=current_user.id)
    if job is None or not str(job.get("kind", "")).startswith("ocr_"):
        raise HTTPException(status_code=404, detail="Job not found.")
    if job.get("status") != "succeeded":
        raise HTTPException(status_code=409, detail=f"Job is not ready (status: {job.get('status')}).")
    try:
        excel_bytes, file_meta = read_job_file(job_id, owner_id=current_user.id)
    except FileNotFoundError as error:
        raise HTTPException(status_code=404, detail="Result file missing.") from error
    metadata = (job.get("result") or {}).get("metadata") if isinstance(job.get("result"), dict) else {}
    headers = {
        "Content-Disposition": f'attachment; filename="{file_meta.get("filename") or "ocr_job_result.xlsx"}"',
        "Cache-Control": "no-store",
        "Pragma": "no-cache",
    }
    if str(job.get("kind")) == "ocr_ledger_excel":
        headers.update(
            {
                "X-Total-Rows": str((metadata or {}).get("total_rows", 0)),
                "X-Total-Dr": str((metadata or {}).get("total_dr", 0)),
                "X-Total-Cr": str((metadata or {}).get("total_cr", 0)),
                "X-Balanced": str(bool((metadata or {}).get("balanced"))).lower(),
                "X-Difference": str((metadata or {}).get("difference", 0)),
                "X-Low-Confidence-Rows": json.dumps((metadata or {}).get("low_confidence_rows", [])),
            }
        )
    else:
        headers.update(
            {
                "X-Total-Rows": str((metadata or {}).get("total_rows", 0)),
                "X-Total-Columns": str((metadata or {}).get("total_columns", 0)),
                "X-Image-Quality-Score": str((metadata or {}).get("image_quality_score", "")),
                "X-Model-Used": str((metadata or {}).get("model_used", "")),
            }
        )
    return Response(
        content=excel_bytes,
        media_type=str(file_meta.get("media_type") or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        headers=headers,
    )
