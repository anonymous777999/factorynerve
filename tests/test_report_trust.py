import json
from datetime import date, timedelta
from http import HTTPStatus
from io import BytesIO

from openpyxl import load_workbook

from tests.utils import create_entry_payload, mark_entry_approved, register_user, set_org_plan_for_user_email


def _auth_headers(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


def _verification_form() -> dict[str, str]:
    return {
        "source_filename": "owner-logbook.png",
        "columns": "3",
        "language": "eng",
        "avg_confidence": "91.2",
        "warnings": json.dumps([]),
        "headers": json.dumps(["Date", "Output", "Raw"]),
        "original_rows": json.dumps([["2026-04-12", "120", "120 / ok"]]),
        "reviewed_rows": json.dumps([["2026-04-12", "122", "122 / approved"]]),
        "raw_column_added": "true",
        "reviewer_notes": "QA verification",
    }


def _create_approved_attendance(http_client, *, manager: dict, operator: dict) -> dict:
    manager_headers = _auth_headers(manager["access_token"])
    operator_headers = _auth_headers(operator["access_token"])

    punched_in = http_client.post(
        "/attendance/punch",
        headers=operator_headers,
        json={"action": "in", "shift": "morning"},
    )
    assert punched_in.status_code == HTTPStatus.OK, punched_in.text
    payload = punched_in.json()

    regularization = http_client.post(
        "/attendance/me/regularizations",
        headers=operator_headers,
        json={
            "attendance_record_id": payload["attendance_id"],
            "request_type": "missed_punch",
            "requested_in_at": payload["punch_in_at"],
            "requested_out_at": payload["punch_in_at"],
            "reason": "Create an approved attendance record for trust output coverage.",
        },
    )
    assert regularization.status_code == HTTPStatus.CREATED, regularization.text
    regularization_payload = regularization.json()

    approved = http_client.post(
        f"/attendance/review/{payload['attendance_id']}/approve",
        headers=manager_headers,
        json={
            "regularization_id": regularization_payload["id"],
            "punch_in_at": payload["punch_in_at"],
            "punch_out_at": payload["punch_in_at"],
            "final_status": "completed",
            "note": "Approved for reporting trust",
        },
    )
    assert approved.status_code == HTTPStatus.OK, approved.text
    return approved.json()


def test_report_trust_blocks_pending_ocr_before_email_generation(http_client):
    manager = register_user(http_client, role="manager")
    headers = _auth_headers(manager["access_token"])
    report_day = create_entry_payload(index=0)["date"]
    report_start = (date.fromisoformat(report_day) - timedelta(days=1)).isoformat()

    created_entry = http_client.post("/entries", json=create_entry_payload(index=0), headers=headers)
    assert created_entry.status_code == HTTPStatus.CREATED, created_entry.text
    mark_entry_approved(created_entry.json()["id"], manager["user_id"])

    created_verification = http_client.post("/ocr/verifications", data=_verification_form(), headers=headers)
    assert created_verification.status_code == HTTPStatus.CREATED, created_verification.text
    verification_id = created_verification.json()["id"]

    submitted = http_client.post(
        f"/ocr/verifications/{verification_id}/submit",
        json={"reviewer_notes": "Waiting for approval"},
        headers=headers,
    )
    assert submitted.status_code == HTTPStatus.OK, submitted.text

    trust = http_client.get(
        f"/reports/trust-summary?start_date={report_start}&end_date={report_day}",
        headers=headers,
    )
    assert trust.status_code == HTTPStatus.OK, trust.text
    trust_payload = trust.json()
    assert trust_payload["can_send"] is False
    assert "OCR record" in trust_payload["blocking_reason"]
    assert trust_payload["ocr"]["pending_count"] == 1

    generated = http_client.post(
        f"/emails/summary/generate?start_date={report_start}&end_date={report_day}",
        headers=headers,
    )
    assert generated.status_code == HTTPStatus.CONFLICT, generated.text
    assert "pending review" in generated.text


def test_trusted_exports_include_signoff_metadata_and_owner_pdf_passes(http_client):
    manager = register_user(http_client, role="manager")
    operator = register_user(
        http_client,
        role="operator",
        factory_name=manager["factory_name"],
        company_code=manager["company_code"],
    )
    set_org_plan_for_user_email(manager["email"], "factory")

    manager_headers = _auth_headers(manager["access_token"])
    report_day = create_entry_payload(index=0)["date"]
    report_start = (date.fromisoformat(report_day) - timedelta(days=1)).isoformat()

    created_entry = http_client.post("/entries", json=create_entry_payload(index=0), headers=manager_headers)
    assert created_entry.status_code == HTTPStatus.CREATED, created_entry.text
    entry_id = created_entry.json()["id"]
    mark_entry_approved(entry_id, manager["user_id"])

    created_verification = http_client.post("/ocr/verifications", data=_verification_form(), headers=manager_headers)
    assert created_verification.status_code == HTTPStatus.CREATED, created_verification.text
    verification_id = created_verification.json()["id"]

    submitted = http_client.post(
        f"/ocr/verifications/{verification_id}/submit",
        json={"reviewer_notes": "Ready for owner reporting"},
        headers=manager_headers,
    )
    assert submitted.status_code == HTTPStatus.OK, submitted.text

    approved = http_client.post(
        f"/ocr/verifications/{verification_id}/approve",
        json={"reviewer_notes": "Approved for trusted output"},
        headers=manager_headers,
    )
    assert approved.status_code == HTTPStatus.OK, approved.text

    attendance = _create_approved_attendance(http_client, manager=manager, operator=operator)
    assert attendance["review_status"] == "approved"

    trust = http_client.get(
        f"/reports/trust-summary?start_date={report_start}&end_date={report_day}",
        headers=manager_headers,
    )
    assert trust.status_code == HTTPStatus.OK, trust.text
    trust_payload = trust.json()
    assert trust_payload["can_send"] is True
    assert trust_payload["overall_trust_score"] == 100
    assert trust_payload["approval_register"]["ocr"]
    assert trust_payload["approval_register"]["shift_entries"]
    assert trust_payload["approval_register"]["attendance"]
    assert trust_payload["approval_register"]["shift_entries"][0]["approved_by_name"]
    assert trust_payload["approval_register"]["shift_entries"][0]["approved_at"]

    weekly = http_client.get("/reports/weekly", headers=manager_headers)
    assert weekly.status_code == HTTPStatus.OK, weekly.text
    weekly_row = next(row for row in weekly.json() if row["entry_id"] == entry_id)
    assert weekly_row["approved_by_name"]
    assert weekly_row["approved_at"]

    export = http_client.get(
        f"/reports/excel-range?start_date={report_day}&end_date={report_day}",
        headers=manager_headers,
    )
    assert export.status_code == HTTPStatus.OK, export.text
    workbook = load_workbook(BytesIO(export.content))
    sheet = workbook.active
    headers = [sheet.cell(row=1, column=index).value for index in range(1, 10)]
    assert headers[7] == "Approved By"
    assert headers[8] == "Approved At"
    assert sheet["H2"].value
    assert sheet["I2"].value

    premium = http_client.get("/premium/executive-pdf?days=7", headers=manager_headers)
    assert premium.status_code == HTTPStatus.OK, premium.text
    assert premium.headers["content-type"].startswith("application/pdf")
