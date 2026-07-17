"""Tests for Excel Export Engine (Phase 6.1-6.2)."""

from __future__ import annotations

import pytest
from io import BytesIO
from openpyxl import load_workbook

from backend.services.excel_export_engine import (
    ExcelExportEngine,
    excel_export_engine,
    _generate_invoice_excel,
    _generate_po_excel,
    _generate_dn_excel,
    _generate_wb_excel,
    _generate_ledger_excel,
    _generate_kv_excel,
    _generate_chat_excel,
    _generate_generic_excel,
)


def _read_excel_sheets(data: bytes) -> dict[str, list[list]]:
    """Helper to load Excel bytes and return all sheets as {name: rows}."""
    wb = load_workbook(BytesIO(data), read_only=True)
    result = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([cell for cell in row])
        result[name] = rows
    wb.close()
    return result


class TestExcelExportEngine:
    def test_singleton(self):
        """ExcelExportEngine should be usable as singleton."""
        assert isinstance(excel_export_engine, ExcelExportEngine)

    def test_list_supported_types(self):
        types = excel_export_engine.list_supported_types()
        assert "gst_invoice" in types
        assert "purchase_order" in types
        assert "delivery_note" in types
        assert "weighbridge_slip" in types
        assert "ledger_sheet" in types
        assert "handwritten_form" in types
        assert "chat_transcript" in types
        assert "generic_table" in types

    def test_unknown_type_falls_back_to_generic(self):
        data = {"headers": ["A", "B"], "rows": [["1", "2"]]}
        result = excel_export_engine.export(data, doc_type="unknown_type")
        sheets = _read_excel_sheets(result)
        assert "Data" in sheets

    def test_generic_includes_audit_sheet(self):
        data = {"headers": ["A"], "rows": [["1"]]}
        result = excel_export_engine.export(data, doc_type="generic_table")
        sheets = _read_excel_sheets(result)
        assert "Audit Trail" in sheets
        assert "Data" in sheets


class TestInvoiceExcel:
    """Section 6.2: gst_invoice generator."""

    def test_sheets_present(self):
        data = {
            "invoice_header": {
                "invoice_number": "INV-001",
                "invoice_date": "2024-01-15",
                "supplier": {"name": "ABC Steel", "gstin": "24AABCS1234K1Z5"},
                "recipient": {"name": "XYZ Corp"},
                "place_of_supply": "Mumbai",
            },
            "line_items": [
                {"description": "Steel Rods", "hsn_code": "7214", "qty": 10, "unit": "MT", "rate": 50000},
            ],
            "tax_summary": {"cgst": 45000, "sgst": 45000},
            "totals": {"total_taxable": 500000, "total_tax": 90000, "invoice_total": 590000},
        }
        result = _generate_invoice_excel(data)
        sheets = _read_excel_sheets(result)
        assert "Invoice" in sheets
        assert "Tax Summary" in sheets
        assert "Line Items" in sheets
        assert "Audit Trail" in sheets

    def test_invoice_contains_header_data(self):
        data = {
            "invoice_header": {
                "invoice_number": "INV-001",
                "invoice_date": "2024-01-15",
                "supplier": {"name": "ABC Steel", "gstin": "24AABCS1234K1Z5"},
                "recipient": {"name": "XYZ Corp"},
                "place_of_supply": "Mumbai",
            },
            "line_items": [],
            "tax_summary": {},
            "totals": {"invoice_total": 0},
        }
        result = _generate_invoice_excel(data)
        sheets = _read_excel_sheets(result)
        invoice_sheet = str(sheets.get("Invoice", []))
        assert "INV-001" in invoice_sheet
        assert "ABC Steel" in invoice_sheet
        assert "XYZ Corp" in invoice_sheet


class TestPOExcel:
    """Section 6.2: purchase_order generator."""

    def test_sheets_present(self):
        data = {
            "po_number": "PO-001",
            "po_date": "2024-01-15",
            "vendor": {"name": "Vendor A"},
            "line_items": [{"description": "Item", "qty": 5, "rate": 100, "amount": 500}],
        }
        result = _generate_po_excel(data)
        sheets = _read_excel_sheets(result)
        assert "Purchase Order" in sheets
        assert "Terms" in sheets
        assert "Audit Trail" in sheets

    def test_po_data(self):
        data = {
            "po_number": "PO-999",
            "po_date": "2024-06-01",
            "vendor": {"name": "Test Vendor"},
            "line_items": [],
        }
        result = _generate_po_excel(data)
        po_sheet = str(_read_excel_sheets(result).get("Purchase Order", []))
        assert "PO-999" in po_sheet
        assert "Test Vendor" in po_sheet


class TestDNExcel:
    """Section 6.2: delivery_note generator."""

    def test_sheets_present(self):
        data = {
            "challan_number": "DC-001",
            "date": "2024-01-15",
            "supplier": {"name": "Sup A"},
            "recipient": {"name": "Rec B"},
            "vehicle": {"number": "MH04AB1234", "driver_name": "Raj"},
            "line_items": [{"description": "Item", "ordered_qty": 10, "delivered_qty": 10}],
        }
        result = _generate_dn_excel(data)
        sheets = _read_excel_sheets(result)
        assert "Delivery Note" in sheets
        assert "Vehicle" in sheets
        assert "Audit Trail" in sheets


class TestWBExcel:
    """Section 6.2: weighbridge_slip generator."""

    def test_sheets_present(self):
        data = {
            "slip_no": "WB-001",
            "date": "2024-01-15",
            "vehicle_no": "MH04AB1234",
            "gross_weight": 25000,
            "tare_weight": 5000,
            "net_weight": 20000,
        }
        result = _generate_wb_excel(data)
        sheets = _read_excel_sheets(result)
        assert "Weighbridge Slip" in sheets
        assert "Audit Trail" in sheets

    def test_weights_displayed(self):
        data = {
            "slip_no": "WB-001",
            "vehicle_no": "MH04AB1234",
            "gross_weight": 25000,
            "tare_weight": 5000,
            "net_weight": 20000,
        }
        result = _generate_wb_excel(data)
        ws_sheet = str(_read_excel_sheets(result).get("Weighbridge Slip", []))
        assert "25000" in ws_sheet
        assert "20000" in ws_sheet


class TestLedgerExcel:
    """Section 6.2: ledger_sheet generator."""

    def test_sheets_present(self):
        data = {
            "rows": [{"Date": "2024-01-01", "Particulars": "Opening", "Debit": 0, "Credit": 0, "Balance": 10000}],
            "metadata": {"account_name": "Test Account", "period": "Jan 2024"},
        }
        result = _generate_ledger_excel(data)
        sheets = _read_excel_sheets(result)
        assert "Ledger" in sheets
        assert "Summary" in sheets
        assert "Audit Trail" in sheets


class TestKVExcel:
    """Section 6.2: handwritten_form generator."""

    def test_sheets_present(self):
        data = {
            "key_value_pairs": {"Date": "2024-01-15", "Inspector": "Raj", "Shift": "A"},
        }
        result = _generate_kv_excel(data)
        sheets = _read_excel_sheets(result)
        assert "Fields" in sheets
        assert "Audit Trail" in sheets

    def test_key_values_present(self):
        data = {"fields": {"Name": "John", "Age": 30, "Department": "QA"}}
        result = _generate_kv_excel(data)
        fields = str(_read_excel_sheets(result).get("Fields", []))
        assert "John" in fields
        assert "Department" in fields


class TestChatExcel:
    """Section 6.2: chat_transcript generator."""

    def test_sheets_present(self):
        data = {
            "messages": [
                {"sender": "Alice", "timestamp": "10:00", "message": "Hi"},
                {"sender": "Bob", "timestamp": "10:01", "message": "Hello"},
            ],
        }
        result = _generate_chat_excel(data)
        sheets = _read_excel_sheets(result)
        assert "Messages" in sheets
        assert "Audit Trail" in sheets

    def test_messages_present(self):
        data = {
            "messages": [
                {"sender": "Alice", "timestamp": "10:00", "message": "Hi there"},
                {"sender": "Bob", "timestamp": "10:01", "message": "Hello world"},
            ],
        }
        result = _generate_chat_excel(data)
        msgs = str(_read_excel_sheets(result).get("Messages", []))
        assert "Alice" in msgs
        assert "Bob" in msgs
        assert "Hi there" in msgs


class TestGenericExcel:
    """Section 6.2: generic_table generator."""

    def test_sheets_present(self):
        data = {"headers": ["A", "B", "C"], "rows": [["1", "2", "3"], ["4", "5", "6"]]}
        result = _generate_generic_excel(data)
        sheets = _read_excel_sheets(result)
        assert "Data" in sheets
        assert "Audit Trail" in sheets

    def test_data_present(self):
        data = {"headers": ["Name", "Value"], "rows": [["Foo", "100"], ["Bar", "200"]]}
        result = _generate_generic_excel(data)
        data_sheet = str(_read_excel_sheets(result).get("Data", []))
        assert "Foo" in data_sheet
        assert "100" in data_sheet
        assert "Bar" in data_sheet

    def test_auto_filter_enabled(self):
        """Generic table should have auto-filter on headers."""
        data = {"headers": ["A", "B"], "rows": [["1", "2"]]}
        wb = load_workbook(BytesIO(_generate_generic_excel(data)))
        ws = wb["Data"]
        assert ws.auto_filter.ref is not None
        assert "A3" in ws.auto_filter.ref  # Headers at row 3, data below
        wb.close()


class TestVerificationMeta:
    """Tests that verification_metadata is correctly passed to generators."""

    def test_audit_trail_contains_meta(self):
        meta = {
            "id": 42,
            "status": "approved",
            "avg_confidence": 0.92,
            "doc_type_hint": "gst_invoice",
            "source_filename": "test.jpg",
        }
        data = {
            "invoice_header": {"invoice_number": "INV-001"},
            "line_items": [],
            "tax_summary": {},
            "totals": {"invoice_total": 0},
        }
        result = _generate_invoice_excel(data, verification_meta=meta)
        sheets = _read_excel_sheets(result)
        audit_sheet = str(sheets.get("Audit Trail", []))
        assert "42" in audit_sheet
        assert "approved" in audit_sheet
        assert "gst_invoice" in audit_sheet

    def test_draft_warning_added(self):
        """Draft status should add a warning row."""
        meta = {"status": "draft"}
        data = {"headers": ["A"], "rows": [["1"]]}
        result = _generate_generic_excel(data, verification_meta=meta)
        data_sheet = str(_read_excel_sheets(result)["Data"])
        assert "DRAFT" in data_sheet


class TestIndianNumberFormat:
    """Indian number format (#,##,##0.00) should be applied to financial cells."""

    def test_invoice_format(self):
        data = {
            "invoice_header": {"invoice_number": "INV-001"},
            "line_items": [{"description": "Item", "qty": 1, "rate": 50000}],
            "tax_summary": {},
            "totals": {"invoice_total": 59000},
        }
        wb = load_workbook(BytesIO(_generate_invoice_excel(data)))
        ws = wb["Invoice"]
        # The rate and amount columns should have Indian number format
        cell = ws.cell(row=12, column=6)  # First data row, Rate column
        fmt = cell.number_format
        assert "#,##" in fmt  # Indian format has ## after the initial comma
        wb.close()
