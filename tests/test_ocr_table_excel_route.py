from __future__ import annotations

import base64
import json
from io import BytesIO

import pytest
from openpyxl import load_workbook
from PIL import Image

from backend.routers import ocr as ocr_router
from backend.services.ocr_document_pipeline import format_for_ui, transform_sections_to_report_input
from backend.table_scan import build_table_excel_bytes


def _make_image_bytes(fmt: str, size: tuple[int, int]) -> bytes:
    image = Image.effect_noise(size, 100).convert("RGB")
    output = BytesIO()
    save_kwargs = {"quality": 90} if fmt == "JPEG" else {}
    image.save(output, format=fmt, **save_kwargs)
    return output.getvalue()


def test_inspect_table_excel_image_scores_large_jpeg_as_high_quality():
    image_bytes = _make_image_bytes("JPEG", (720, 720))

    inspection = ocr_router._inspect_table_excel_image(
        image_bytes,
        content_type="image/jpeg",
        filename="scan.jpg",
    )

    assert inspection["image_quality_score"] == 90
    assert inspection["image_mime_type"] == "image/jpeg"
    assert ocr_router._select_table_excel_model(90) == ocr_router._TABLE_EXCEL_MODEL_HAIKU
    assert ocr_router._select_table_excel_model(60) == ocr_router._TABLE_EXCEL_MODEL_SONNET
    assert ocr_router._select_table_excel_model(30) == ocr_router._TABLE_EXCEL_MODEL_OPUS


def test_run_table_excel_pipeline_rejects_low_quality_image():
    tiny_png = _make_image_bytes("PNG", (40, 40))

    with pytest.raises(ocr_router.TableExcelRouteError) as exc_info:
        ocr_router._run_table_excel_pipeline(
            tiny_png,
            content_type="image/png",
            filename="tiny.png",
            system_prompt=None,
            user_message=None,
        )

    assert exc_info.value.status_code == 400
    assert exc_info.value.payload["error"] == "Image too vague or low quality to process"
    assert exc_info.value.payload["imageQualityScore"] == 20


def test_call_table_excel_anthropic_uses_expected_payload(monkeypatch):
    image_bytes = _make_image_bytes("PNG", (720, 720))
    captured: dict[str, object] = {}

    class FakeResponse:
        status_code = 200

        def json(self):
            return {
                "content": [
                    {
                        "type": "text",
                        "text": json.dumps({"type": "table", "headers": ["Name"], "rows": [["Amit"]]}),
                    }
                ]
            }

    def fake_post(url, *, headers=None, json=None, timeout=None):
        captured["url"] = url
        captured["headers"] = headers
        captured["json"] = json
        captured["timeout"] = timeout
        return FakeResponse()

    monkeypatch.setattr(ocr_router.requests, "post", fake_post)

    extracted = ocr_router._call_table_excel_anthropic(
        image_bytes,
        image_mime_type="image/png",
        selected_model=ocr_router._TABLE_EXCEL_MODEL_SONNET,
        system_prompt=None,
        user_message=None,
    )

    assert captured["url"] == "https://api.anthropic.com/v1/messages"
    assert captured["headers"]["Content-Type"] == "application/json"
    assert captured["headers"]["x-api-key"] == "test"
    assert captured["headers"]["anthropic-version"] == "2023-06-01"
    assert captured["json"]["model"] == ocr_router._TABLE_EXCEL_MODEL_SONNET
    assert captured["json"]["max_tokens"] == 4096
    message = captured["json"]["messages"][0]["content"]
    assert message[0]["type"] == "image"
    assert message[0]["source"]["type"] == "base64"
    assert message[0]["source"]["media_type"] == "image/png"
    assert message[0]["source"]["data"] == base64.b64encode(image_bytes).decode("utf-8")
    assert message[1]["type"] == "text"
    assert "return ONLY valid JSON" in message[1]["text"]
    assert extracted["type"] == "table"
    assert extracted["headers"] == ["Name"]
    assert extracted["rows"] == [["Amit"]]
    assert extracted["_provider_model"] == ocr_router._TABLE_EXCEL_MODEL_SONNET


def test_call_table_excel_anthropic_requires_api_key(monkeypatch):
    image_bytes = _make_image_bytes("PNG", (720, 720))
    monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)

    with pytest.raises(ocr_router.TableExcelRouteError) as exc_info:
        ocr_router._call_table_excel_anthropic(
            image_bytes,
            image_mime_type="image/png",
            selected_model=ocr_router._TABLE_EXCEL_MODEL_SONNET,
            system_prompt=None,
            user_message=None,
        )

    assert exc_info.value.status_code == 500
    assert "ANTHROPIC_API_KEY is not configured" in exc_info.value.payload["error"]


def test_call_table_excel_anthropic_rejects_unknown_model(monkeypatch):
    image_bytes = _make_image_bytes("PNG", (720, 720))

    with pytest.raises(ocr_router.TableExcelRouteError) as exc_info:
        ocr_router._call_table_excel_anthropic(
            image_bytes,
            image_mime_type="image/png",
            selected_model="claude-unknown-model",
            system_prompt=None,
            user_message=None,
        )

    assert exc_info.value.status_code == 500
    assert "Unsupported Anthropic model configured for OCR" in exc_info.value.payload["error"]


def test_call_table_excel_anthropic_preserves_upstream_status(monkeypatch):
    image_bytes = _make_image_bytes("PNG", (720, 720))

    class FakeResponse:
        status_code = 401

        def json(self):
            return {"error": {"message": "invalid x-api-key"}}

    monkeypatch.setattr(ocr_router.requests, "post", lambda *args, **kwargs: FakeResponse())

    with pytest.raises(ocr_router.TableExcelRouteError) as exc_info:
        ocr_router._call_table_excel_anthropic(
            image_bytes,
            image_mime_type="image/png",
            selected_model=ocr_router._TABLE_EXCEL_MODEL_SONNET,
            system_prompt=None,
            user_message=None,
        )

    assert exc_info.value.status_code == 401
    assert exc_info.value.payload["error"] == "invalid x-api-key"


def test_call_table_excel_anthropic_retries_with_alternate_model_on_model_error(monkeypatch):
    image_bytes = _make_image_bytes("PNG", (720, 720))
    calls: list[str] = []

    class FakeResponse:
        def __init__(self, status_code, payload):
            self.status_code = status_code
            self._payload = payload

        def json(self):
            return self._payload

    responses = [
        FakeResponse(404, {"error": {"message": "model: claude-haiku-4-5 not found"}}),
        FakeResponse(
            200,
            {
                "content": [
                    {
                        "type": "text",
                        "text": json.dumps({"type": "table", "headers": ["Name"], "rows": [["Amit"]]}),
                    }
                ]
            },
        ),
    ]

    def fake_post(_url, *, headers=None, json=None, timeout=None):
        del headers, timeout
        calls.append(json["model"])
        return responses[len(calls) - 1]

    monkeypatch.setattr(ocr_router.requests, "post", fake_post)

    extracted = ocr_router._call_table_excel_anthropic(
        image_bytes,
        image_mime_type="image/png",
        selected_model=ocr_router._TABLE_EXCEL_MODEL_HAIKU,
        system_prompt=None,
        user_message=None,
    )

    assert calls == ["claude-haiku-4-5-20251001", "claude-haiku-4-5"]
    assert extracted["headers"] == ["Name"]
    assert extracted["_provider_model"] == ocr_router._TABLE_EXCEL_MODEL_HAIKU


def test_call_table_excel_anthropic_respects_explicit_requested_model_without_fallback(monkeypatch):
    image_bytes = _make_image_bytes("PNG", (720, 720))
    calls: list[str] = []

    class FakeResponse:
        status_code = 404

        def json(self):
            return {"error": {"message": "model not available"}}

    def fake_post(_url, *, headers=None, json=None, timeout=None):
        del headers, timeout
        calls.append(json["model"])
        return FakeResponse()

    monkeypatch.setattr(ocr_router.requests, "post", fake_post)

    with pytest.raises(ocr_router.TableExcelRouteError) as exc_info:
        ocr_router._call_table_excel_anthropic(
            image_bytes,
            image_mime_type="image/png",
            selected_model=ocr_router._TABLE_EXCEL_MODEL_SONNET,
            requested_model=ocr_router._TABLE_EXCEL_MODEL_SONNET,
            system_prompt=None,
            user_message=None,
        )

    assert calls == [ocr_router._TABLE_EXCEL_MODEL_SONNET]
    assert exc_info.value.status_code == 404
    assert exc_info.value.payload["error"] == "model not available"


def test_run_table_excel_pipeline_builds_excel_from_form_response(monkeypatch):
    image_bytes = _make_image_bytes("JPEG", (720, 720))

    class FakeResponse:
        status_code = 200

        def json(self):
            return {
                "content": [
                    {
                        "type": "text",
                        "text": json.dumps(
                            {
                                "type": "form",
                                "fields": [
                                    {"label": "Invoice No", "value": "INV-1001"},
                                    {"label": "Amount", "value": "12500"},
                                ],
                            }
                        ),
                    }
                ]
            }

    monkeypatch.setattr(ocr_router.requests, "post", lambda *args, **kwargs: FakeResponse())

    excel_bytes, metadata = ocr_router._run_table_excel_pipeline(
        image_bytes,
        content_type="image/jpeg",
        filename="form.jpg",
        system_prompt=None,
        user_message=None,
    )

    workbook = load_workbook(BytesIO(excel_bytes))
    sheet = workbook.active

    assert sheet.title == "Extracted Data"
    assert sheet["A1"].value == "Field"
    assert sheet["B1"].value == "Value"
    assert sheet["A2"].value == "Invoice No"
    assert sheet["B2"].value == "INV-1001"
    assert sheet["A3"].value == "Amount"
    assert sheet["B3"].value == 12500
    assert metadata["extracted_type"] == "form"
    assert metadata["total_rows"] == 2
    assert metadata["total_columns"] == 2
    assert metadata["image_quality_score"] == 90
    assert metadata["model_used"] == ocr_router._TABLE_EXCEL_MODEL_HAIKU


def test_build_table_excel_bytes_orders_dict_columns_and_adds_totals_footer():
    excel_bytes = build_table_excel_bytes(
        {
            "rows": [
                {"amount": "10.5", "item": "Widget A", "date": "2026-04-29"},
                {"amount": "+20", "item": "=cmd|' /C calc'!A0", "date": "2026-04-30"},
            ]
        },
        metadata={"Source": "OCR"},
    )

    workbook = load_workbook(BytesIO(excel_bytes))
    sheet = workbook.active

    assert [sheet["A1"].value, sheet["B1"].value, sheet["C1"].value] == ["date", "item", "amount"]
    assert sheet["A2"].alignment.horizontal == "left"
    assert sheet["C2"].alignment.horizontal == "right"
    assert sheet["B3"].data_type == "s"
    assert str(sheet["B3"].value).endswith("=cmd|' /C calc'!A0")
    assert sheet["A4"].value == "Total"
    assert float(sheet["C4"].value) == 30.5
    assert sheet["A6"].value == "Generated At"
    assert sheet["A9"].value == "Source"
    assert sheet["B9"].value == "OCR"


def test_run_table_excel_pipeline_extends_short_headers_and_adds_totals(monkeypatch):
    image_bytes = _make_image_bytes("PNG", (720, 720))

    class FakeResponse:
        status_code = 200

        def json(self):
            return {
                "content": [
                    {
                        "type": "text",
                        "text": json.dumps(
                            {
                                "type": "table",
                                "headers": ["Date"],
                                "rows": [
                                    ["2026-04-29", "1250"],
                                    ["2026-04-30", "900"],
                                ],
                            }
                        ),
                    }
                ]
            }

    monkeypatch.setattr(ocr_router.requests, "post", lambda *args, **kwargs: FakeResponse())

    excel_bytes, metadata = ocr_router._run_table_excel_pipeline(
        image_bytes,
        content_type="image/png",
        filename="table.png",
        system_prompt=None,
        user_message=None,
    )

    workbook = load_workbook(BytesIO(excel_bytes))
    sheet = workbook.active

    assert sheet["A1"].value == "Date"
    assert sheet["B1"].value == "Column 2"
    assert sheet["A4"].value == "Total"
    assert float(sheet["B4"].value) == 2150.0
    assert sheet["A9"].value == "Extracted Type"
    assert sheet["B9"].value == "Table"
    assert metadata["total_rows"] == 2
    assert metadata["total_columns"] == 2


def test_transform_sections_to_report_input_creates_dynamic_report_tables():
    report = transform_sections_to_report_input(
        {
            "title": "Invoice OCR",
            "metadata": {"source": {"model": "opus"}},
            "sections": [
                {
                    "title": "Vendor",
                    "type": "form",
                    "fields": [{"label": "Name", "value": "ACME"}],
                },
                {
                    "title": "Items",
                    "type": "table",
                    "headers": ["Item", "Qty"],
                    "rows": [["Bolt", "5"]],
                },
            ],
        }
    )

    assert report["title"] == "Invoice OCR"
    assert report["metadata"]["source.model"] == "opus"
    assert report["totals"]["table_count"] == 2
    assert report["totals"]["row_count"] == 2
    assert report["tables"][0]["headers"] == ["Field", "Value"]
    assert report["tables"][1]["headers"] == ["Item", "Qty"]


def test_format_for_ui_flattens_multi_table_sections_without_losing_report_data():
    payload = format_for_ui(
        {
            "title": "Mixed OCR",
            "metadata": {"scan": {"page": 1}},
            "tables": [
                {"title": "Header", "headers": ["Field", "Value"], "rows": [["Invoice", "INV-1"]]},
                {"title": "Items", "headers": ["Item", "Qty"], "rows": [["Bolt", "5"]]},
            ],
            "totals": {"table_count": 2, "row_count": 2, "column_count": 2},
        }
    )

    assert payload["metadata"]["scan.page"] == "1"
    assert payload["headers"] == ["Section", "Field", "Value", "Item", "Qty"]
    assert payload["rows"][0] == ["Header", "Invoice", "INV-1", "", ""]
    assert payload["rows"][1] == ["Items", "", "", "Bolt", "5"]
    assert len(payload["tables"]) == 2


def test_build_table_preview_payload_keeps_ui_flat_for_mixed_sections():
    payload = ocr_router._build_table_preview_payload(
        {
            "type": "mixed",
            "sections": [
                {"title": "Header", "type": "form", "fields": [{"label": "Invoice", "value": "INV-1001"}]},
                {"title": "Items", "type": "table", "headers": ["Item", "Qty"], "rows": [["Bolt", "5"]]},
            ],
        },
        template=None,
        doc_type_hint="table",
    )

    assert payload["title"] == "Table"
    assert payload["headers"] == ["Section", "Field", "Value", "Item", "Qty"]
    assert payload["rows"][0] == ["Header", "Invoice", "INV-1001", "", ""]
    assert payload["rows"][1] == ["Items", "", "", "Bolt", "5"]
    assert payload["tables"][0]["title"] == "Header"
    assert payload["totals"]["table_count"] == 2


def test_build_table_preview_payload_unwraps_structured_cell_objects():
    payload = ocr_router._build_table_preview_payload(
        {
            "type": "table",
            "headers": [{"value": "Date"}, {"text": "Amount"}],
            "rows": [[{"value": "2026-04-29", "confidence": 0.98}, {"content": "1250"}]],
        },
        template=None,
        doc_type_hint="table",
    )

    assert payload["headers"] == ["Date", "Amount"]
    assert payload["rows"] == [["2026-04-29", "1250"]]
    assert not any(cell.startswith("{") for row in payload["rows"] for cell in row)

def test_run_table_excel_pipeline_builds_excel_from_mixed_sections(monkeypatch):
    image_bytes = _make_image_bytes("PNG", (720, 720))

    class FakeResponse:
        status_code = 200

        def json(self):
            return {
                "content": [
                    {
                        "type": "text",
                        "text": json.dumps(
                            {
                                "type": "mixed",
                                "sections": [
                                    {"title": "Header", "type": "form", "fields": [{"label": "Invoice", "value": "INV-1001"}]},
                                    {"title": "Items", "type": "table", "headers": ["Item", "Qty"], "rows": [["Bolt", "5"]]},
                                ],
                            }
                        ),
                    }
                ]
            }

    monkeypatch.setattr(ocr_router.requests, "post", lambda *args, **kwargs: FakeResponse())

    excel_bytes, metadata = ocr_router._run_table_excel_pipeline(
        image_bytes,
        content_type="image/png",
        filename="mixed.png",
        system_prompt=None,
        user_message=None,
    )

    workbook = load_workbook(BytesIO(excel_bytes))
    sheet = workbook.active

    assert sheet["A1"].value == "Extracted Data"
    assert sheet["A3"].value == "Header"
    assert sheet["A4"].value == "Field"
    assert sheet["A5"].value == "Invoice"
    assert sheet["B5"].value == "INV-1001"
    assert sheet["A7"].value == "Items"
    assert sheet["A8"].value == "Item"
    assert sheet["A9"].value == "Bolt"
    assert sheet["B9"].value == 5
    assert metadata["extracted_type"] == "mixed"
    assert metadata["total_rows"] == 2
    assert metadata["total_columns"] == 2


def test_run_table_preview_pipeline_reports_requested_model_and_token_usage(monkeypatch):
    image_bytes = _make_image_bytes("JPEG", (720, 720))

    class FakeResponse:
        status_code = 200

        def json(self):
            return {
                "content": [
                    {
                        "type": "text",
                        "text": json.dumps(
                            {
                                "type": "table",
                                "headers": ["Date", "Amount"],
                                "rows": [["2026-04-29", "1250"]],
                            }
                        ),
                    }
                ],
                "usage": {
                    "input_tokens": 1200,
                    "output_tokens": 345,
                },
            }

    monkeypatch.setattr(ocr_router.requests, "post", lambda *args, **kwargs: FakeResponse())

    payload = ocr_router._run_table_preview_pipeline(
        image_bytes,
        content_type="image/jpeg",
        filename="scan.jpg",
        template=None,
        doc_type_hint="table",
        requested_model=ocr_router._TABLE_EXCEL_MODEL_SONNET,
        language="auto",
    )

    assert payload["routing"]["requested_model"] == ocr_router._TABLE_EXCEL_MODEL_SONNET
    assert payload["routing"]["selected_model"] == ocr_router._TABLE_EXCEL_MODEL_SONNET
    assert payload["routing"]["provider_model"] == ocr_router._TABLE_EXCEL_MODEL_SONNET
    assert payload["routing"]["model_tier"] == "balanced"
    assert payload["routing"]["forced"] is True
    assert payload["token_usage"]["model"] == ocr_router._TABLE_EXCEL_MODEL_SONNET
    assert payload["token_usage"]["input_tokens"] == 1200
    assert payload["token_usage"]["output_tokens"] == 345
    assert payload["token_usage"]["total_tokens"] == 1545
    assert payload["token_usage"]["estimated_cost"] == pytest.approx(0.008775)
    assert payload["debug"]["requested_model"] == ocr_router._TABLE_EXCEL_MODEL_SONNET
    assert payload["debug"]["selected_model"] == ocr_router._TABLE_EXCEL_MODEL_SONNET
    assert payload["debug"]["final_model_used"] == ocr_router._TABLE_EXCEL_MODEL_SONNET
    assert payload["debug"]["raw_api_response"]["usage"]["total_tokens"] == 1545


def test_run_table_preview_pipeline_returns_structured_rows_and_anthropic_routing(monkeypatch):
    image_bytes = _make_image_bytes("JPEG", (720, 720))

    class FakeResponse:
        status_code = 200

        def json(self):
            return {
                "content": [
                    {
                        "type": "text",
                        "text": json.dumps(
                            {
                                "type": "table",
                                "headers": ["Date", "Amount"],
                                "rows": [["2026-04-29", "1250"]],
                            }
                        ),
                    }
                ]
            }

    monkeypatch.setattr(ocr_router.requests, "post", lambda *args, **kwargs: FakeResponse())

    payload = ocr_router._run_table_preview_pipeline(
        image_bytes,
        content_type="image/jpeg",
        filename="scan.jpg",
        template=None,
        doc_type_hint="table",
        force_model="best",
        language="auto",
    )

    assert payload["type"] == "table"
    assert payload["headers"] == ["Date", "Amount"]
    assert payload["rows"] == [["2026-04-29", "1250"]]
    assert payload["sheets"] == [
        {"name": "Table", "columns": ["Date", "Amount"], "rows": [["2026-04-29", "1250"]]}
    ]
    assert payload["routing"]["provider_used"] == "anthropic"
    assert payload["routing"]["provider_model"] == ocr_router._TABLE_EXCEL_MODEL_OPUS
    assert payload["routing"]["model_tier"] == "best"
    assert payload["routing"]["forced"] is True
    assert payload["routing"]["ai_applied"] is True
    assert payload["used_language"] == "auto"


def test_run_ocr_with_fallback_requires_explicit_opt_in():
    with pytest.raises(RuntimeError) as exc_info:
        ocr_router._run_ocr_with_fallback(
            b"fake-image",
            language="eng",
            columns=3,
            column_centers=None,
            column_keywords=None,
            enable_raw_column=False,
            allow_fallback=False,
        )

    assert str(exc_info.value) == "Local OCR fallback is disabled for this request."
